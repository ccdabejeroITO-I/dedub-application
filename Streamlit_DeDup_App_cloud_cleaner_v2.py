# -*- coding: utf-8 -*-
"""
Streamlit_DeDup_App Streamlit wrapper
Single-file build generated from the user's uploaded Tkinter app.

Run:
    streamlit run Streamlit_DeDup_App.py
"""

import base64
import gzip
import io
import json
import os
import re
import sys
import tempfile
import types
import zipfile
from datetime import date, datetime
from pathlib import Path
from functools import lru_cache

try:
    import streamlit as st
except ImportError as exc:
    raise SystemExit(
        "Streamlit is not installed. Install it first with: pip install streamlit"
    ) from exc

import pandas as pd




_TKINTER_FALLBACK_BLOCK = """
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
    from tkinter import font as tkfont
except Exception:
    import types

    class _DummyWidget:
        def __init__(self, *args, **kwargs):
            self._value = kwargs.get("value", "")
        def __call__(self, *args, **kwargs):
            return _DummyWidget()
        def __getattr__(self, name):
            return _DummyWidget()
        def get(self, *args, **kwargs):
            return self._value
        def set(self, value=None, *args, **kwargs):
            self._value = "" if value is None else value
            return None
        def insert(self, *args, **kwargs):
            return None
        def delete(self, *args, **kwargs):
            return None
        def pack(self, *args, **kwargs):
            return None
        def grid(self, *args, **kwargs):
            return None
        def place(self, *args, **kwargs):
            return None
        def configure(self, *args, **kwargs):
            return None
        config = configure
        def bind(self, *args, **kwargs):
            return None
        def heading(self, *args, **kwargs):
            return None
        def column(self, *args, **kwargs):
            return None
        def yview(self, *args, **kwargs):
            return None
        def xview(self, *args, **kwargs):
            return None
        def see(self, *args, **kwargs):
            return None
        def after(self, *args, **kwargs):
            return None
        def update(self, *args, **kwargs):
            return None
        def update_idletasks(self, *args, **kwargs):
            return None
        def destroy(self, *args, **kwargs):
            return None
        def mainloop(self, *args, **kwargs):
            return None
        def title(self, *args, **kwargs):
            return None
        def geometry(self, *args, **kwargs):
            return None
        def protocol(self, *args, **kwargs):
            return None
        def withdraw(self, *args, **kwargs):
            return None
        def deiconify(self, *args, **kwargs):
            return None
        def winfo_exists(self, *args, **kwargs):
            return False
        def current(self, *args, **kwargs):
            return None
        def selection(self, *args, **kwargs):
            return ()
        def selection_set(self, *args, **kwargs):
            return None
        def __iter__(self):
            return iter(())
        def __bool__(self):
            return False

    class _DummyNamespace:
        def __getattr__(self, name):
            return _DummyWidget

    tk = _DummyNamespace()
    tk.Tk = _DummyWidget
    tk.Toplevel = _DummyWidget
    tk.Frame = _DummyWidget
    tk.Label = _DummyWidget
    tk.Button = _DummyWidget
    tk.Entry = _DummyWidget
    tk.Canvas = _DummyWidget
    tk.Scrollbar = _DummyWidget
    tk.StringVar = _DummyWidget
    tk.BooleanVar = _DummyWidget
    tk.END = "end"

    ttk = _DummyNamespace()
    ttk.Frame = _DummyWidget
    ttk.Label = _DummyWidget
    ttk.Button = _DummyWidget
    ttk.Entry = _DummyWidget
    ttk.Combobox = _DummyWidget
    ttk.Progressbar = _DummyWidget
    ttk.LabelFrame = _DummyWidget
    ttk.Scrollbar = _DummyWidget
    ttk.Treeview = _DummyWidget

    filedialog = types.SimpleNamespace(
        askopenfilename=lambda *args, **kwargs: "",
        asksaveasfilename=lambda *args, **kwargs: "",
    )
    messagebox = types.SimpleNamespace(
        showwarning=lambda *args, **kwargs: None,
        showinfo=lambda *args, **kwargs: None,
        showerror=lambda *args, **kwargs: None,
    )
    tkfont = types.SimpleNamespace(
        nametofont=lambda *args, **kwargs: _DummyWidget(),
    )
"""



_OPENPYXL_FALLBACK_BLOCK = """
try:
    from openpyxl import load_workbook
except Exception:
    def load_workbook(*args, **kwargs):
        raise ModuleNotFoundError(
            "openpyxl is required for Excel workbook operations in this app."
        )

try:
    from openpyxl.styles.numbers import is_date_format
except Exception:
    def is_date_format(*args, **kwargs):
        return False
"""


_RAPIDFUZZ_FALLBACK_BLOCK = """
try:
    from rapidfuzz import fuzz, distance
except Exception:
    import difflib

    class _SimpleFuzz:
        @staticmethod
        def ratio(a, b):
            a = str(a or "")
            b = str(b or "")
            return int(round(difflib.SequenceMatcher(None, a, b).ratio() * 100))

        @staticmethod
        def partial_ratio(a, b):
            a = str(a or "")
            b = str(b or "")
            if not a and not b:
                return 100
            if not a or not b:
                return 0

            shorter, longer = (a, b) if len(a) <= len(b) else (b, a)
            if shorter in longer:
                return 100

            n = len(shorter)
            if n == 0:
                return 100

            best = 0.0
            max_start = max(1, len(longer) - n + 1)
            for i in range(max_start):
                candidate = longer[i:i + n]
                best = max(best, difflib.SequenceMatcher(None, shorter, candidate).ratio())
                if best >= 0.999:
                    break
            return int(round(best * 100))

    class _Levenshtein:
        @staticmethod
        def distance(a, b):
            a = str(a or "")
            b = str(b or "")
            if a == b:
                return 0
            if not a:
                return len(b)
            if not b:
                return len(a)

            prev = list(range(len(b) + 1))
            for i, ca in enumerate(a, start=1):
                curr = [i]
                for j, cb in enumerate(b, start=1):
                    insert_cost = curr[j - 1] + 1
                    delete_cost = prev[j] + 1
                    replace_cost = prev[j - 1] + (ca != cb)
                    curr.append(min(insert_cost, delete_cost, replace_cost))
                prev = curr
            return prev[-1]

    class _Distance:
        Levenshtein = _Levenshtein

    fuzz = _SimpleFuzz()
    distance = _Distance()
"""

EMBEDDED_SOURCE_GZ_B64 = """H4sIAGS+5mkC/+y97XbbSJIo+F9PgUFtjckxRUtyVY1L0+q7siR3aVuWvZbc1d0qXQxAgBJKIMAmSEksle65T7Fn998+xZ79t+dsv8k+ycZHfiMBUi5Xd8+54+4SE5mRkZmRX5GRkRFhGG4cVGVdFXkaz7M0GC+KIpjG89F1kCzyIh0Gty+H3wz/NYhV/PMgXZbxJB8FST6bX2+m8TKYZXFdlcE4vx8GB3FZlfkoLjjrN0FdLWajbHNUZHG5mAq8G8flqFikWGR+n9XBeFZNgvl1Xgej63i+u7EZHC6mBaCZZ8E8ToosqBfTaTWb18H5h49H0eHH9yfHB/vnR4Pgzcc//9n8jss0gPDR6dlRdPDu9A2Ez4eA0IYLFjUWu/jpp+WL6XVVZnNoEbQrCybYzLy8Cu7y+XWQ3cejeVBWs0lc5D9BfanVSC3Eadcl2AtqxDDNZkiO5/ylMsgIaFZW1pTfrafEIECAWHOJLc3mcV7UQVpBbebBXxZQn/EyiGsgW2aWi3hPq6C+i6dBUV3lI4x4rSpRz2c5tOg2pj7PsduqWXA3q6DFk6qcX7+AHn2xzOIZkPIKqpAV1V2w/YoJizFxUt1mwauvEe+H6q5mQuEIKSsYDvM4AIRxicPiL4t8hjQr4vIGejor0uCqCuZVcBKPboJqHBxSq6h78jIFqn/goVTkNXT1JK9r7IlZfMeICUUdJMvgOovTbEY9RrmLKp4zqWGc4jAp8pss2P722+3hVvAi+Ib+bm/BTzzLgmtoSwEVq+NxViwRwe+yMgPCKMrLHmcSzbCaZZb+W3By/m6znmajfAzAYwSyaDnLJkCcFDGeTaAewfsP794ffdis50sYw2leTwugbnaPQ1lkn1MLRQGDIKeJgXGnxx9exCPomOUkmM6yOpvdUiGI/IgxQI8TjbAXoQMSpKpEhWB/yOscJw9NIVksk64OrvM0g+Lm2ayE2XqdFYAtGFXFYlLWw40Q1oYvgsOz7w+xItBRqZiQWNh5VcH83oGF4c3Hk5Pg/f75wXdB7/yGsPUxn5y9m+N8Vs95IOIE4NFHiwGAfZxezaAy6S6EoRNjhIxnV5gLas2ZFjQE9ACm+c29BMM8KaoRFHs14FHoTEigSXUH/ZzdZmVwdw1/dM40H4+RDL71IphkUIsUqVNx06mCek2a4bi/mlWLKUDNq6sMJuEMwAOmMUEfzKoa5qvMsxq+miTYcTqLSJW9DyimcT57QXhgkkEvMt1gCYVckGGewzKQT+ToMtYeazzbjRFdNM2nWQHlMyFH1SzbrMpiCR0yutmsxpu8AHGnbGyIQubc5bQM3WzwKi6iZDWgJ9M8hmwDoGpdwwqSVPeDYN4GD2sQo8OQLGcKvQSR8P9pytlm8TRPcf1WGSE8wCk2j8tRxkDjRTmaw1itJVAxW0SjeHSdScRVLUMzFbeA/auCUQkrjmrnNexxOCtVBFBahWfxKMO5JyPqpcKqegWqnqrmwLS/luG/LLKFqG41zcrp8r5Qta3iNLqrZjdJVd3YIENaUephuZgkOIpFjryOsMSIl5aNDejpve5/AHEGqyDQG9axWc57bTaf09halTfNxrSGRpCjd4tbBqyFi6y/uxHAv/lsyQH8l48RP1SQtq9T2HF1Gv6DxCFjQQSUlN2Psukc1jr8gWGrM0zjul6zcefUc5tYy+DjsVjm6jWyLvJoHtc3EfUPbMz0O/xf8W+vvwEzclHM66gu42l9Xc0jLGUJYG/ios42iDLRbFFGVRkhJqpFD8fjIPgXWOJqWLDifL5H4BD1Lzd3GCsoN6sAY5X8CPiuiioBmF5/CItGL8SUkMkDFFXDcjhazGZZOZcF9ZHQOnUCO4JOgoVBFQBgdl/A9rGYlTRzeqKiqnIbBAXTF+r1EKaQL9wNRAtCpghEID74zmazaiY+HzknESUvb6sbIKEu0honooQLie8SyvJXRkK7wwSnWuZByBVCfJlKHOew9xW+4ql1CHw+g+nZHM+SgsN4DHOltzWQLVs1dgWBkSwbsh9xKGiIu2vc/XCaGDWxq4hLyrAusmzaA55mu29OM6uxrfNtFufIYBqwbg2tXthoVJ06c7qYTCMxU+qeb95zY5CIdvnjEjhK6lHuT6C0NeNwuEdlhZTp9e22u8OF0bUPkO7+sNYUA5RrcTSZzpfOwvOUGWpOtGZHNJrijKqvcVhZNG4f9jYiXiGpj6ghgKCHJHeJ5Okxuxumi3mv53ZWf70FmooHdLCuR9NZdQWDqe7B8gss2FyUTCDxdFoszRUBCCfh20ewhLgIac+gyTqJ73FCAKuRl73tLQqO8WygihVVb67NXAtzWe7bLQDGYr6oe/Psfr4XhsjNIDNbQ7izLY0+hsYxqqiI4XTV3kD8Z0IOR1U5zq+4Bvin7+IVVVoHsQVqYRYpTxppn0DSWTaG3ruORtWinKu1w4NkMSWWBtZphuUq161osa9mGbJDmKfX2TkmtLHMtJXZW2vw4GJn1EhwvT3gc+F8N8/nRaZY4c7KjUvaZ/NyXME+qpnnITAcdxgLuyxMxxJ2+Ga6SDA2YgeAoh9pzeKaeQroG3XpOTV/cp/D0bLIRnPKkpewrqhO5zNvFvH5XWy5Pg7SAqQe6uQpXbzARhc9Lw4iw8pVrbWeYmN8UJAhHnuiGKhOAVh4Zsso5lKGeAieIos2tlLNdgQw2LIAVhkHY2JjTDoxJisw1tcZDH6sZJ1hz2RpJKKaaF2ItVAnTdTJStSran2dzqjOcGrssSgjgtN4FBvdiFjdpCbWLQdp0kSatCNNViNlYUdUZLdZIXBzVAlTKKqLHHAZyBtpTeyvTPTWaCT8xjfDPa53RvqQsZAWRbzX4oi05ulvJrJGmLU3y4p4nt/yl4+3SOKaE1HSuayH0duj4/f7Z2dB8EXwfnlcwnYHi8QsmGeTaTCuChRxdM9JE2NVDzE0jJOaqhMOBSMm5qdM/7EC7kBlHAR2tdci2bv358fvTvdPcFmdxyRBXiOfgo2kaFMeGBWV6JSv4OS5Pq0WcDDXRfnp4UOvjy/YXQ6eXm0zXTjaaocjZdIhwxOGmqkBzDM4gSyS3iy8+K/7m3+ON3+6JJgBDmRA3B8upjCOjF1VoJoTprWo/LuTd6/3T4Kz/TdHwdEfD46I6GuOzGh+E+E+FmWSSD0IRfPlFDYxCMFWnHgFFQ3JFmDS+6BHZKU3ToKeJLrFW0AnyD48vzHpsDVEGVs6i+9MnmOS6H25F+5Pp8HBLK6viag8ZJWgacjync6m9fsdx4ytYZpBN1XL3pPZvA6mf4ZL3KdVdkAb1x6uCfUcZv0MuAb84EzXVXXDlHS6dN317Sq7D54Hcziel+vIfk7fnUb7J++/248+HPFIH1WTKVSQRjsN9f7GwbuT6ODkaP/UB/RDHf2w+cMQ4U7fRYdH5/vHJ9H5u98fnZ7h0geL/gVNqXJRFPSLh3z4XZQ3cOgtMUi3JhgQkxmDm/Rnk3OgcDGm4Av+4b+b/ANTGVmxfIQiPRFjRFz2N95/F70/2T8/8tT/v17AlL58ePn4Q/1D+vDV4/8UGuAH796+3z8492XryXz9HuXrY0Yg+EE1mVRlUOajG7pkexHERQ6TZRJPSVBNIuhN6/5t4/T44Pen+2+PgGzRwT50yfEBrAV7gssKf6xqFD/x7wB/3c+l/V1n02sdI5As4pLi8Jegrp1v45OzTFBKipEcGHAgc2OWRgRnjEvKFnM3xaX7mdmfuf7m/NMsneFBQAQGGJhnjYiZEcMZi0VeYyT9AlRRLRi5GWFAcK5RPCuoOAoQHIaaMVVtxiki3QgC3AiCjNzvSkeIemYUhT8D+imdz3iWagjOM4snDMYBlEPGS1/UouS8IpJzj2dxOcprror+GKgPf7wLzsiy9Com4nMAT15pyoQ2YoxPSapyQfyhDBF5yqqozDiGvQUGkc5yMjSgkBnBgElW/hhPciKDCg84zFVyYz1xdpSgd5VkszkRkkMDCtkRSZUkSzNKtDQfXceiqSJIq9tN1oi6WjhwYkxez/J6Xk2veZSbnwPx2YiXkw82kXLJU4qDkGPuRomexIuyJCOu2vjAUZj/1IjzADKamxhv+XKe2PpjwB/eyKUv1s3P2GGmZzUtISI0EKF8bsVuPNJNCV9H2xoOtzsbb48PD0+OIlpn3747xBU9fFNk93htHCJXfsY3ti8CGQvB46uymmWIdX+R5vMXaZYsruQN8i3eOedFPl9unH337vto/+Ph8Xn03dHJ+6MPZ5LfRcy/z7Kpc+2srqOL5b/x/TTxW9NZdptnd/ICO1kiLxsvivl62/8BIZcqA0IRZV0uksAjcfBD8vXwjxRcjJmibbcogmNG4TYyxZSTYgQ+3P72TLHBWT7JC9g05svoDK9gsSPfYne9Z/ElsJnmWViDrwONShcRK1ggpKVxYcKpm2EDWMVZ4I+SCthzRqMahNBJF7V1l1APZ9m0AE6xF0bE6Upp0SS+yaJFmf9lkQnq10Q/eWKps4zEY1yFaoEnkgtGjQxFiRUiePuoCFAldZtPyuBCYp9h0L5eiWu+nIECsBI2m4wxFwiC0uhtKwnqOAQGLCtTBymW3orleQeacfiAQI9B70HneOw7h97FXF7WzPIKxkr+kyJsnkZi5vXgV5IWg1B71MGhaLwwvLhkpALHUlMbVhmD9gtASWQn0WX0obr7eHyIHatDH6lsMxzwh/wbnNKldvioenOExMaqaDKNIugd0UejvhGNmg4QT+mNwaUjAhGBCiLmIRV6mFFDiboxWAfGDNEPsK/dQRrz69gADsu/pai+e13BlJOdN+oaAEhUB1B0p+qA5wS0Ifr2ljV9Iki7wjvNp/YrNLu5Wjem8cXI7g4lM6BOgJMbql7h2RYIHtUogQ/7PC5gMU+zMmJ1OmvJi85IgISkc0NBb6dvLkxPGU7GIsm4IIccWEYiLZfR71CjBhHQp/jSQBQRvUOFGwSiT/HlYvpQ8WGLvvjDBTmHk68C4Q/f2nse04lPRdB3xyI98CzSQKfG6h69ofGhmkxVjj6Is2YEOLBojjAW+cYK+3eak+ZAsufMqCrnebnIGtn84/KTMq+XTS0UFmxYq+HNoR1mnn1LCQNAWumOWHFmgKFzJXuQApUcnZQ2s0eezjQXg0/pmM15mKnvmRo6M3foiGsIGD2qaCxnJoYOoABkMzVuaOysIJexmY2au1Z09nb/w3nEOpTR/sGHd6d/enumRQD7xwdnWPLr/Q9v32LgYP8DzYOz72k5ODw++R39vvuOf06O8Pfk/B3/vPnwWnLTJ7/7iHGnB4Tg9Jh/Xh/TzxnBv3vzPf68P9unHy7i/OjscD8EHtuu64d3b/dPdUUJzbH4yz9/wL/855j/yh9O/yP+/aOs3h8p8o/H8kf8Uv4/ir/HIbH6tB2ggKzAO3dgrqNpnsHsor+ab6XPVYwrA/GE5vx29j04LYS+zBQlWHyhosbahKPrRXlTB4tyXi1G16ilKhUlsiFqn9MgJQFa+sPmi93h5XNonFl1cyvEaFnUPl47BrFUuCX13zkMxpqlj6gXhAeNuJheA/s/IWXfO5g0QkFFqFjiqajOpvEshsN0PTRHpBB360qEFyz2/ut/3/zr//bX/33zr//XX//vzb/+P5fPe/9l91lbWv+/GBOqiCdJGgeT3WAypCnV2+pf7G5fSsk57PFGwvbupVwZNQqmAYb6su9rPOlFUmc0EirEdRSXaTTLroASrMQQmbp6YiPX8Q1q6yRDmUss3RFJVnvUeabqA8aSVgbEq4bYyRG1FYDoq3FngEpuBhgsqv5FwXtpYeTc8CVtGDubus34IRF9d/mwM3j8IYHxZzdzYJCib410VD1N4wJPFHBsjcuABhqcdwOqwyhGrf9GkcaI+iGBsZP/nOP/4b/bn+F/Ofyf/oOI+5/hfzn8n/7DP7c/3+P/8z5WtHNgSdIOzMVXf4yL+Kregyod/+703Yejg/2zIzmsuIWn8g1EML2exfh6gubYsw80pII/5PmzYPO3+vv4+NmqtjJov/dD/bzfu8hv74tROoE50tKWccgZHmSjdvqPKvxSNfAxfHIT3Y1HTCQ4QmcldDt2HPc38s10JLA1XXFBTId5XcY9M7pdEmBoudYR3YVjgrwVV6dqmdS5zFKVmMEyruRgAG9tfmteycmqI9tvcFlcWlkJneFInBIMDl28iShZDis+BF8ShDAz5iKNgw0emy6k8UUJSctY7WKpwvToxEiZW/wwLR1TQ5wBvJESetQk9GhlopzzmC2xUFRDOYHV+AahRTdsdGxVtd6sBHS/Gw0/UuEeA85rDnhoGhgI5OyJJKzBfGMU1pySGle4lCqgYFTWU+SuHc5VIZb8F370V3Jrdh4vn4F4+sbUlxeCKrc139bariCr2t5ET8kDbzrGmEg8remlYz0l03Erf0MqrLqK6Xg4qqbyMlSdbuikjCyqOEx3afUB2MUIhT4cGLI2l1i8bgc43vZGu62LCqwn/SdfxTaWLCAGHlQsoWX9ZJqIUXQd1/Ec1o10jLJ1pkDYHNbpmGJkf4jlo10gQUQlgjNN+2aZJhZdkoPblEK42LokElIWQUelLJXoTOytY4LjsG/N3JdG2oqBSUPYkMJceCTLo75uWLMYSVmfZJSkPLIykhh44pZtaggDUYMcqDOZRvViPM7vvcrbsovnQ/mOZVhWd6y+NcZPoOqXf/py8mUaffndl2+/PAvX1H4PozK7C+WU1s/YIvmMjfWJDDUilG2ivsDc0POhlROiGI6JBMsGPWbFZwNCPPrgaezjA+R75D2hVKJaPgtI9Nk9vkLsKZTG2F+7mOihNIvi4pRMV1BDYRMEmVdiYXXncZv+9hfyQeBkUc+DO9gbM3oWSq/zUEtgPKOHkgz7nq9T8BEbzJ9JDpsLPorEoffxWD5SDOi4ho/jCGeZ3aJa1mxRGg/TaonxeEzFzfHtHj/5wlUG3yq9wPd5+AR0SaMB1rk6vs3wCWgcKHLhM2D5fLFUFU3HOzQNW5Y0vaLtNJc0zgs82SE0/g02vteXGy+piiHxLrYuEQG9PqAbAKYpq/FzBxApZ2F/w6uOiricVw2ksmYM0TSf0dTWI9TIz9BN3WzIjJMc8tY9hsGRD4Mxqm6EdrHR0KEcLu0vIOSBGZ+dGfPzvZp3R6h05OEj3PbRqxEVExdzqXq3Yg5bWcRzD1q67XpiqnoOkuajec9tCfaXQGHXTMZip+6pmnlB8LC5mBZZT0Z0SeJ1lS7sMbFOMfjT1lUCCrtLl9HoMlXCE/TlztaAHeEjySLCJ9owB2ECqpdr+AQFvogbUI8wxLf1bkHE2W8OROSGpdgsIi3VZImRdH5H1SSpNKAZqSFdDWR/vILfEGtFFEc40vjVGqXIhMRN2GgoDVs5DL1fmaGpsytS/Jrqssp38TSa5GkK9MCX+21p/EQZn/W1AKCqJ6mJtQEkFSydDYiNorqqonxyJSPo2+7AFHh6Yj6AEnU2F0prfSM+seI3zs73Tw/3PxxGb/ffvz8+/V305vjo5BCltRc0qHshKXnhkewNPTg+hZbTXQ/RMBSCiV7IVacbEgq1AhYxozuJO7DxTkwpKvSOmJK40FDG2TMI9VvzF8Hhu9c+rAQfyTMtZQj+JL78sOqYy8Bv5acfWhyQGfaQPxp0wnfK+TRGVQuilvPtaWU8i8srgdoINyFh4t/CECZyvDfCTUghD4hYAiCe16uLtiCE0bTp1pzkBCLHewyvgJ/ES+CkaYeZ880ZfPPuNW/Nernxdv/04/5J9O4PRx8+HB8e6QHp1lrVp7+B+lZwWJnAFIIJU9My8/DYiE4a0bRa+eBFgp0D1l0ffhXdhPbiNxJEjg3gl2J+4p7KuQxRiROFL8thyURFcEYZToVKEklYcCWVH7z2qYe8Zs7kU3Lye6rYqQnHqsV71fPqGPWOovomnwLzSMguLp3IhCMFKsE/xnYBUbqYNuLKKmJbBzrFRpJ4kCStSBI/ktiHpRGJxheMWIUCTUv4c3CSv7VmWmc+N5G2FKsa9LyOpBcw3ObVPC50FisNKtmSQmW1pV1no5uWNKSJvyZJR02S1pokHTVJOmqSdNSE6gGnwIk3J3ZcWxr3QGsq1qY1lUaLkbjBs4qsv9gMjZmQNBPoeZ5ohMk3cCpXsZEJW9WIpCq1ofBUSQxOP/bYH520FBu3xCeeCuG9derLwQlJW4KnaXpuepqnEpOuxNiu4jqsvzjUv4ATPGyUJIuCPZtMeeTl1TqPHt5/OPrD8dH30Yd330cnx2+P0RrV11tbKv79/u+OorPjPx858cD4nX88i/6w/0EPSjb8FJG9JtrRxO60Tks0+2UYVnpCO94en0JlTo4Po9fHH86/i/50RDXb/hbqrNOgMRj5CriEP1pRr74W3cGLXrObKD5piX9qt33HSqmwV8Ahhi6pY7S9stYrlYN3Jx/fnkK/nJ8ffTg19CKY0d4NLjhUCtaY7fzIr6v8NiutD5U0tnKoQKSiTaSoaq0+KjitmCgjCxM98AgvlTo48flYTQ5KWP4K9Cf9HfKPDaRDkZGggpBgFUknBiwQAxIKw6qwejFTVY4nebG0vxRcYeaWv6oOhUESyhcZ+AMzq00ScVjBGupzS3Y/B141Z6Uc+JDZIcgYEVAVgL+RSFI5I6sUfdjZFWczJ1pexgXyK60S+oHPaszvTcUnGnGzIqJqLF6kujeAkcJtFSLvAs07wcD84MOQW09+aCBrTKlW+/hw1mygdQ8ZyC/8NVqGn1bLlkyAROUWZz+3iUuNm4NAOZkHgpFZniqot4R/ffqEf7628OGx2Rj77jRQnxQw2kPfVoMmokEagzyhuk0Sb2hkeaIlKh+2yi5XF9ibTKhZk4mvUditvgG4LBeTrpGA42xpDcOlMwqXnhFo4FVoqkSHDNwSaS9NqfppalffOnhbLXCP5PAtfmjxGonouXjhZ4JHI51HRXKcxo/fL9xCMDLwRUa+SE8V+VuhgqN1KxCWT5Wy+1PKEnDhMoUMyexK/gYcoESZCh86KPPc5kWhbvphS9TZYr2MyghrXVNCC6yGKcHAsPzFOIUGI8ywlYCQ9vLM19GEXgRJTfJK/KiFlj/NMCTZqPBuPR7hOZHQiU/qAw6yZEJ8BNaX1r1QyW5MJMZ622cl1oBcbKEUCsxoAyc9FVcBATXPikwlwIcTi1Bq171a8KYLv/TaEK3DXlW3wOZMspLs7c2tBE+UgKW4q4XeQq8WL+K6rkY5sYn0dFJ/BqJEGtYibA5lGYcY5fTDsDv8IS4ySgzMEmUjiVfFSSc3VhWjdlEZ4X7zVj3NZxqUPrkNzSgNr7CgQmVemEjVbpvOgBGbPauVaVPIGtiJXWkdSTbOYRfSYQfWYSdalwwe5H6Q1RBuQcvuQpadBehUH8G8zV2d36q3mk5adkmzij/NnJamVmR9GdNapbsxVoYvdIONkjgiasQYJNVoXWq6KWamypuhagEe+qGHnk4ycaxTky98qDmykMgKiUDtQFKQTDtQITqUAoEjatbxtiZd1K5ZR6XfZtf5qMgUcvEdcYRiUUyJOZ2wSGSuRej07XxGzrcAN2poR4sn554PlSFB6714gahRyygRc0kPVNc5KtMd9RMMeJJxUrIWQRY1pPqGzwxWLyTkUL1x+IBPJAB6GNGaH0WPu8EDqhNl/ccfyh/KB59hC9Q6FS8XqUj4M5vLgoXRrz1RiE/RZlJfsSZJR+HhCtUaxhHulwGVG1QjMgqaDsP2dkubXvWVqP1dnqLhRaH6ctehFNS7M59Q9klVhAxsQX7SG7tD82RoRUxgA+qyOpywGdpfbXBLFCWE/l+gegpaBa7RejuUlqG8IagrtH0el1nBeixoYX6+gRadyJYwXvecZVcVmn+FKnxL79OLlO1SrB51H+f4ljnP1h1xylQ8XpKiXpY06cpNMY2GkK6uoZ1rqOQKvShhmD66yZassy8fv+IzVqsgbQ6wacZH6p9m0sQdokNpUf1o5mhYD+rQdryNWYntlrSGblFhrWlaiGz+Xlr5sOQhG9dDHej6U1UeEY9QeSTJmfnc0dDUnMTTKZtBEu+EtSF/pefsdJjQjt4NVNB67YbINSpSIJCqLA7uIXVaX6gOIQ6IGCj5HuJzBHjDfJ5NassCYLUo1Z2dETuTaFgHmDE6FGsMEQHX0EKasgq0TRiczU7cxfTSY3dV1LAJ2wBNZll8s2GpP0FeG6XosAsgFRn+RQi/TvKvQg16OT6S9KDebTZZTJdy1Ezyak03SE0Di4uZ+nG0UbUcXXrhbdK2k/dX7IamEeesnkekoQ8YtppJ0CA5st0uWNmdT+hSAo3EiuesqK0Z1hoHa4yFzvFAC7mgDxqvH6JyfI63+sj096YDQNtvK1U0Cadpz2kV5Ar+WQD02+sli0aTuRSGvXG7tTyG/q3Rq+2YrZ6n325QHgnlyB3ABp7f7gWvdpjFEDk8Fm3tQevOHZnx8kkWlI09R+Df2Dg8erP/8eQ8+vj+5N3+YXR+9JbtYn13tH/INk4u9HWDEBwYFzNWjHV5IazP2FJ7mXVRFFaElEA5si8pX2pKCh1ZXVMQ1Txfykoa0lyO0eJR/rbk2MLijjI+EtbZfWjIglA3Z8OhWrR/crx/dmRcYmniyQubyLw4+Ty3N/E0K5Z5qmVVRg+p+zP7EuvJt2n0ZV+rGddkfMV26RkO+mIssi+92u7JnAszSs3LHNcU48bs0j/MsDR35BlAQSOm6zaq5V4qsDKrm6rLxhAn0usBz0lmWFUHbc8V2VyTSEa434FVEbwvyMpsDNMjZjGa8amEzDpKl1fkWTk3q8MxgV0rjlOfi3oOh5RZo9ZBo4hVomZHxmz2jBZER+uLxRvicIwIzAhbNn7Zur6QGPtXvJ8wYNruUF6MzCJX3GP47lv4Emb1tUbbdYQinLr4cO47Wi85uu8Foo7bgKbwv3k/8IXx8YVf9K9lPL67ABapjbKiUDJ//Aj011TWarrq5kCEFZB7k5A51wzmTYJX9OlKOLvlnUwMU7ppCELXl4s2pYlrCx398klHFNkpiFwlJF0tI20Vzn7hk4aulNW2SWWf1U+Utf5///3/8JRqxlrlOgnDthSnbCtRlY6SAqdoI8oo14rVqJtyYfrQd0QGH0Wbu31Jbt3lf/rNvnk5r0OTiaqF4t2wDtY9d7cGxsp7d3Wdrq7VU1WoySBiua7ig6mV8WkqGoamhQqQPoVQ37g0+FKsgVh+4Ye0MmrxEVgfPfriKsIkQbmhKFt9ZGmsm4mcLmnvZMQDXWWlsMoDES/0F4deCKjJizEL1wt6mz/OMPTZbk81qIo2LlB916Ej52rUcx0ayI3ycWOjcQCKDvYPvjtSYithR1G+cYzQtjvddKDiuBJ30Yd+5F4DiUcw8PKZiPQ+wJOi5npZkypa9RNx2ixe3nUcAOF7Gt/TPNdqPJudzkYLUo/tN+RlApPHV4uus3wgL2DXc5fTaGE9muXT+ZrVjiJ6XxX1rcdyGoVrYbBRVw36idUd3aVcT+gWCDsGvO7SlTUAmPWLRouHcnxo0TYZk+ZHmGSfUZXReNwooBx4iRMFQNVcET2vIVa8xVxlAkwhGcZpKvO0y8lI1kRGRKH8ajbPUuxaFDkZRXrkOHRXgHIOfDDrWkwz2kpJ1kP4xZQcC8p5GPLdDcNBNzBULxyOamL+h/dFfS8D8ncS9ltkS3QQ2rN9LchHrNoAqqeimtY4kHuIp0N8ReuFMoSJsGvKdaz+Mh7G1nKlsoijXhyLqrCzr6BlzZPXKW1LotdTjagD3SG0ZDQuEei9LZpBWrWgegebHDREaN+gUQNGDQQeB56ewGKjdMzvrPF9TgSAPfaoUaKD1L2tQZDiheZePYfOv8kygGfDvXBykq4Ps3JUoR3avXAxH2++2qzzq9Dx3dZ4lOsvnl/WtlXAximd4MLeom7fpE8auu+pCmVeQpRk2ZhwM126VBQFNOvdvl3yo2SRsTlN1honT5wDG+vUaIW80zLnsmIUC6dYbLzCftJv39zZ5n0Jkl6vGTZw8vR+IPspY0tb86wnrZuybRHdbFpi0RiNMP4jbQbwy3+3O93NAnPbJBT4xuHHUzSefRg9QH2C58G2YemB3gXQ9pSVZFWALUdtoWG3bWu/uMCUSzLYAVk0Am63XOZ4mxhLvHvBNtd+HD5g0mMUPVDSo20BWGCR1CdBuFozyA9DxLaMBNkG7JwhgwDes8u71T127DnJSxbK773aEvQ1odCPo/kJHaV3Zy5MXj/btzcxT7wYu1OUT6/9eQT4royse4bNbb1GSpszxp2ttf2b6WZlV2zswsiWXW2JqzlkAGx3xaVZ83osZw0CRQQilsM4iUzbW1sddguIluRjowOVUV0CefqNJlbWZC4ITZ/ZJ0hRu4iOp5CZBQGNFJWF4nc31r272mqumFlB5iIEEYiuomKKKGvjh8G+sfGEC7EVl2DGeJYjyLpsUumNezA1/3juG2eseDGvIpS5P9HwUzoWFh3HDbNHetfD21Pvum1sjNJU4JsK/a6TSZgx+SYHXqWCMyS5maf3Tqw5FZAV32Ipzcjsl0u8IpjFLwB8Uy5Raq+GMbNEyzGEF07/QLdrGLjoKRWPq4zFZdykdsbKTczOXVTVDb0XdVap6/5ucE2Ti7gwtzTT0fF8WogFXxpE6XuMalrl8SZhFUgss4FC1JMs7RhtNM31GRdnVInm5Rl/9C3vgjM3l+fuq5nPvrDijN5Lq2bWxgUU5+64hPJU27gyErV2r42amdTdCudw71csWOPyhKF9FyhWjhFfhjB064WIHWnmN24dGIfv5sEu0bwzEOX67g3IwsGgYdfRzOeVqLv57DdJos+bskz+6nuegem2OeJHDjeziLdcnMcjOpSfZk6W83EWLeuzQFhaxyBCFoc/Fo1I+iaIo0RiEkRMd8lD2KIIMUdhTRbeFTU7K5y166XVNWOneRlhU8uz1uxaKlCCEXZQNzwPCzBV4+b+593tc59iI3KgjhYiVdZ3VnRoccERzPEWDXBVPRKhYFktakwbfKitYJCyLxVXHdAmh7C4yn45rQzqFOJkwJaHNkyIBFROv3kXfhAjQZm+4mbaK7O9aJs6DZe7dAwVtWIDZOY6rZuo08L+5aAVtbFa+3Bbi7mLnDUSurCbK7oPvb3iu/iFHkJXAe7DVU8ZQtGgiV4mdFFHbQte4uhNowOH2CsQgVMBuYt0ZFYbh698Y1dx26aTOpCzaN6DeMRbTaM/rJ2oA7HafHzIjZ3JLUAnddVa71Eempo7WAcSY8PyIDG3s65mGtubt6VGuqexZurKUtI2aqrUthLSFfQ0Ns72Aii5rQRO7CgCt1PE3QGC2+kKENxOvesfbbONlQ9jL63dV++Sg0Auu7BCu0txU7sabcvr3dXZoVbIAkbXUK3SVWA1NmJRqncDRnG7ADMOWDJW+Jbq2qJl4b69s6HfK6Abpz8lkKFtqkUfj6guqSt3bcdMtKxNl2TJ3lbbJEyqzf1mA+yKdzEUBL7RwU8QQF+et0kPzfXEJgaBKMfqYHus+caOHGqWvEBBO10uEv5pTw5GjBRBMRR0DtuGua67afmUD/GcJkV6exrU4C3FYbyTt1Si7GZl7B5JxxcMeEli1g2fqeWLjqMzlWQXcCkuA+ez2Gf72SAkjhJZwablZCXnsIw3B88N1JfK7Zw03FaT+WbunZ5wck4mErVMhRxdmynKcQda7H/dNNnfMCM3CDyW4lSkNgdnRUm4jQ7MsQ9z3MQc+zDHQqA0raYL6iKNCjWgI34kVlttHwS2AJ+wowlXTCCs/KFK1p9GOtu3Xt0FzMfH48w4Q11civWVR4Qp3rcuBSwr3GFvfzGvgkPi8PvhJQwKE631eAcHl9+moGGqnOz/qAbS8glZDWdY2FhaulWjm0D4cAoQmS4KeXfAvN5LRIWa3p0vsjq8NM1/KxBeh7mKVK6+9sL5J0y+QyJpqcAxoOEb0ZAZS2w4IUVQzEH/voVl4vnYIbohWBWjgxwjGOOkSSE1bgToKnJKxC5JNZ5WsioQi7QxijizNBK2riZspRWT7bLN8z7A6pa1U18APKEHEPOu58KQC0KS4/4k717NKuMEIytdVGvx8BB9MczXmfR/m7n+95mDn3codkxt92VlyxyxeMgnrgONZW7F1GhOgKeOIwf/rzLF5FitittMDVPyuTAmW/nmKNoL98PGo1BHWOT3TGEMYNG15BCKh4QIcvzfb0DrQeUbDc7RxfeYd4HSyZ5alaSLIPc2vFurQOIxliqqGMbiWsshYHSd0WgkdvCXzvMuwxOZ+z6L7sLSVNi9Vu7c0Hw3cE3LiB1G4QihEPb/nuUI1H6yrW7E9DWa7TjYdTOA5bODBHQCYmqlYJlYMy5X1UyqwnRXx3FLsmZtLJ9nXrcS6CMr4PrlZZrdc23kG3RWSYvKqoyUj0ieYL67RRw4UENg7efLNk8xrnsVfWM2zeG4TtJlqdtU9x0pNzEXKrV5uqYDgHaBstumRTPyer/hiUnCUb6b1VW6YKvlp83rMNuguTgcCxFoEwvbND/13I05Bs8FIhIDN9GgzfPTxo2eZQxdIGBhThPD64a5P8v+eXAovtACukAlVrEmrqYPX5GDHXP5cjRc+V4a66IQjlyorhgYBB0omlzu2rti0XTJNJYOXx8g+TFUDpog83Cc4yuuHi50gBANb6Ca2JCdvtLCR2FTTS4fa3rubvgLUgC6OBW1TqFCvRVxyVPKZRdlJDVsGukiu4mkytFOxS2KaWd0qvMNBWU3swZS+fQQ6MingfobusCZ65BNeWEa1srqcd1LlnsKdhDcwCK2F9akvE3OhZGXpYWth74LhP8UXnSkLwNHsGD6UJK6fp3erXENlLh8tcVEJYpRkAM21FLvhflViZ7sWh0R4sBggzDxCD3hNYyBYBUxSmzcKh5F+1WaoQugobqq64Wnb36PU8/2xCb9tI1YSDMiYYxy8W1iQnYnL2GvBdi+1OrTN4EkeDIq2HCbI1wlGiB+b4kr1MBlI1vpY01duc6/O432T95/t28ZYGk2At1KGRUUuX2N5JxKxSXSW4OJoLRVx4zsps7YpM1xpKz98cHvUd0wOn8XHexDW44P9k9Ya2QygOx9txm8JDytIsLJPT2WFc/yykVRsHvFm1I8vCTuivcwshbD4Dg++JUtf8/RDx6QBSfj44rGlROuvbEpImcNrAk6VOvFsKjBlEEb/3Nc4ZTVg0lewtoHg3Pv7fHh4clRRBR6++7wCIGiWCguQjChoCBEvGURQlAqFrumNzHRqxOWR2K+Y567jbbhOtMAFu4pGsBYlz0oVGlQFdk9NnqXD0EBslZ1npLjrCQLiPZADlhgiECzLLi7zsoA3b2wpyhzVAFy8Zwh2Wqvpvhmk90m2begoK12whOF8Y+gMP6RprAMTzQ8O/lER135y3uBsJEDHby/mC1IN9brW8cSBpKRKh8Yn336Bm7lgEd6nnBLsD30dBdjwTbLUr58Whpj+/rpLsqCbRalvQK1lOW4DeouzAaWpTkDkO0JeUegW6e1h65DN5lPzqUWeOztjUas6hweuXoNX38FWn+poTXXP8QpKemvTz1zmVlJMWFIjao0QgGYb9OKhSZt0gqRbKkKIhog9shfL+UcGUjTlE5Q7J6mIvIXvXjUN3Il3lxJI1cy6m90r2Ij6LvRk1exLwDhpLrN0oDfiKkjMB5Yd41+07oFGytz5XVEVlqflEnDAjPL2ZOF8JtborC2vnH3fmlr0tz929kvv6Nq5qPkccj0zl5bTtNrdH0cBpqphILefxfxlTJwWtphcu3xjCyuZKC3RnPbjfX+5p9NN9bicA0wCvnBu7fv9w/O7UIELlWVJl81DpXr8O3+Y2D6FLeYknYiELWhL2UXYHfAvJPW855EdZ6Y7GHxzkd6dZrCZK87cAuFxyE4O7U3iAlgffdwC3EyCpnA03fR4dH5/vEJ8Ju/Pzo9w6VFYUcmcQ7nKxhmIzhTvCC0/Ihj7sv82FVlc9GwR7ESTViu1elcJKaBci4QVcmPvdTW8Ej9UtFGDeigUqEO6p70dzuk717DJ6PXwchvINuQdFEhRBn5C5fuVMWpFVGIO7XUrJfDoIEjZjkSDR0E46KK531bdNrtXr7hhx4xC7Fspw9594jWrEwPEJ9Lh60D0yWw+nBfhKZSTEupLLF8CuK+uPpiUa8j6Vvd+c0Jzx20u/VV+ghh0tHa3dqhD+gbCoaffhxdl5B5OZd921c7exMO+SynNSW5dqKson9dqmzjSEMw+PlmC/41yYG9YtG8F26/+vbbze2dzZdbIT5Ng8R5Rc6M06yYx1jhHuDsD8iwwF54GPb7okt9T4LW75lf2Duf1kNyZ2MCuutr3T5RJETXAtk6wcRirHeoWdj7IX346rF/8cOLzUsMbw923C9rO510vOtdLuEEPIHZg72L/aX3tv7AithxI146Fg5SvQD2NNr+P8D8W30DtR7RPYRWX189gehMmQEQ/z+JromO+z8gZvN4MEV64nKL5K4d1JzGs5qe+PLyIzeBniEmHVUZ3zJI/Hsy0PcpR8itkDH3/UshJ/5jLmefYKQSe0HwpVxvwwym5qTIqil98nterOmej61xRMw6n++GGG0HZCUafgZG5ZU8htZkPOMqn/e62F5rOKgJVV/sfmVom04mKv6r3W+MBLXs1Rff7L4yEzrm1Zpmx1cxhvL2WZBxbUYQe0l7h2470/XqbJajnuw63YS7E4G3PvWEKXFGED22i0BMhjo0iILpFRMC4e1Msey1n3H6G8a7aCuPh3+X6hToroRFHgKWrcIH97srRu29QQUZ6FudYDbin4PefyNkUrbPM1jZe4rGs2pCBVnTgXrHlN34x/6zZ7907CvFcWvoJ9WsNMasngQDc+APzMG+5mhW5aHaO43nxiDpGuZoDcsaz5tUVxnu9TiJ1jlBRVzn+sFvgh4BihQKY4LVcYBdmSfghZad5+IUqElVQ91n7j17BluuCk1UKNVxXXOFk6HAPcsZJOrB3nO06RBSXmGRK1RUIMHHetDR6bgvH3zR84T6hncuMbsuaO24DP7FyNnvXEoENdDdrqWI/gynGhPl2a5RnNCeGPgA5WqyIsPdrCqvIv3Y4wng1JlPgIc2roDGeZlkKIrdfrUGaJxUt1n06ut20Ed11u7RaKEFmEOWlhGfwXoTBTNphUkVTOqBMTxxIHUaTwqo7j/GJVR5e0ChRTxb0pcNM84SiN0ZUGgmgHYcoAn118sBhUbXFLYh4ilCfDWgUF5Q2MWBmL92Yn9cYBW/GVAoo6ALgMj+lQCKJQWdkhdXEPtqQKFFPacPG6TOphD77YBCcyOYoeFE+rYzVCOE2t4aULBiIPi0ocrqFqORvhCUuLZdCqfZCKORxBBUYAaNH7XGJl/BT4sc1QjGFU2V3q17cW1tnx1SwVsvq2pkb+Xle3y3gHy0NHFna3LIg21Tb4+Sdwx54LOL/4qSQFgrcb2sG3VC4D1IW1UZk7e33+e47L3N1+04Fp46mdwnFPtF8K4sUKdynM0C1ARYzJZ0IpllcRFsb+4EtB2zkcehk/WwgvP8OSpM3mazefDVJsOSfPOahfKscRIU+U0WbG1tbWNrqmAHQsOGGS1kB3b6yPGx5fotDAIJfoOWW1qI0wNUW8Fz+CBpGULvBS+3WA7WQxfGnOgWRmgVCwrthYjfGrv1JxNUgiHzLICkemDb/KDV7z/GBPGaQHviRPE8dveNffv8suMxVrNyDjyh20xQPPMIQKRkj2SEExIRbu84coj6pd32ePOn9rbTCH/ZR6M0Lz3uGLCxaieka+H65cXuy0sSSrsJ61h882Js1AqgvBL/Txj0BuVWDXpk4v5zT/iEPUEf3X+tPUGCoaiuMQ9SOpK/3O4bfcvmfVFF+WLJSpt8TnWZgL7N8Iksk7YsBGSIJuKlyJG25cBjkl2tSDje4v3LPT7fX2xdOpWSGbiGK3Kg52QBj7VrgdYF4IGDNpz2Cm2bFUJ4Ng3TUaFGDnFe8ldo26wQnzhW1GjHwM8ZVlapmaWzTjtmnehpSGTmutCvb/Aks0BT0M1otNpsx45pWx8osfMdhO8gfAfhCcUDteAnxXPJTzDX1YAZGGNhoHp5YNBrYJJioBs50J08MPpvoHqm31DBn7DnBQSgQIp/7yjyjuyv3aWCQTFuP/WiTS8I0vYl3KGnVCD2rH0WgdeAQ4qvAPPaxvEuax45DrBt4hJgIgJp2n/qmtfSfD8H8Ldo/fr98StJkFobi+X9Rgp7mrBmgxH0t1IA1FhQeNrq8M+GoMdpvrN0sbhwz1gpfzZXwZ/VNPJkk0IcVbTIb65ZP+tamW9oUYh0YQqOLmUd6LMDUkmO7BwyupGzIUK6tKrbCc8yJJ2BvjtzoBTp0uyKBrQlRbq0RHJqfPS9uZRAqZkLR0pDCYTzbjRehbCTMvXV8qCNtXxg72++DHn2rO9aHEWtnBVi9Sa/qTOtkt83hut/MzBpZQY1HG0pOwvyW97PZH+x2mNt0OtIS5ub9NNy8dRdN4+5HT8lj9irV2WxdnKyXoyUvGU7tNhdDQs1tL3czwUTL6D9rHzz0DDMi2p0AeVcyusI0YU6YU3rb5PtxkX6szW0F+AUMb93rmYnO35UK+7kfaiQ19j2yVEMVQg4J7K+zWSbb97rnrvloo3YyY7vtGnc7ms8O6146mx9Y3pPl9foqWP0655W/3XGgXPKb8NHs+oJCNVxqQ0hNmAVOqduLFu30Su9KF+znyBxWaVTsb7gpaN5xjtteXf52S84xVJTtyFuv9y8l3ea9LfvX+2Mmv8zlWMXEO8GccN0Ruzjrozl8NNw+tiwz8bR/IdlY1RfreZdZA90MSxOlHhlGKNTwqgW+gFjsugzCMq9na0t30s+Ib2wXKrqa3DUQ/TqCdUN7e8hl9wr90rcEssUhm49hwmxt/2Uu21uRcqOMGZ1FpE/AG7O2goSW8Mtf0uFdgM+3gQmrf3xrpXnoiZzB7a9jdp9Hu8pnteu6RO1oZxDl9RFIn1RwDaEqUZ1n2Rx2VtXbwCrxKSV8gSY1f/BSVsuJoSE0EsNdlSSN9Xjs6vsfs8+WIt8+HNB4WYVMHrNWiyXqn/JfUU+QjXbZg/3FWk2LBZizWIk297rIZMBW+4rmM+oHsO3Szvb8NmHo6RMNxO//bbfMp6Yk/zEkaTEUX+nodR41/95hpV17x/+GJfhIBTX/BAaZwn/nYmICTqpC+n6Hn7j6Yz/5gXFIsSPi5L/SlfKjX+QWBBIgfDx4or/Luo5BOpsyn/FB1+uQ7gazflvxd8lunkN5Y05BNNsxH854vEXzxy6u1lnrJslRWroQnYcuTQyFeOoQIsM/bj764U3WB31glOFKkWgafBWvXs8oKnOpfurHm5z9/3f7r0kvuUeL7ZsqH5jAqoWyaM+r9cw9VQtfoW5JsS9/+Mu2mn61KGnhh1kVcNOnU0+bx8J74jXZOZWSYPIHhWyYMKokeivUCgGs4UWNl+FLqBh0ahmuaQh1Hc+gxUAjRAV+SifE5YsDehqSVlIpBfjcVEEL40n4zttednmQ5ZqnVWBiLK97Ct7u/hEvYZFbU5eT9DQ1gvhg4W98RH8+TUMuPi2ylP4GY3yNCvnkH8ZLFAsFSzKWYaPF9PAKEjUmGbYCxjVLxRrLi2IklrGv2vnA/8+CP5dehXAsOEq4N8DfDQfF+g4bBlIiYlFZGEYxG+an96okamipmanYylf2uRQxsfkmY/lCti9htnllD3yUR4zx9ILLYw0N8EnfnBhl7oJn/rhycq0gBamCc7meGzRQ45dYorBwiOMW6eM3AKhuRSy9ovuWyZBms8y7foFRTCymaSCJxthfKTmx9JjSFdlak1KW8WxUnl+oApQoYkKpZIEp2hpRRNAtd4zRyRNnWYyhYyqcUR39UxQXu74vw3TzpOyjeXYDWpa5SzxfbcYw+Z63LMWYjESyOkB5MAqs7deaepTRvLQcmNxAOk4W/iC2yi67Q3bE6Eh0udvK5Ag9rIB4QipXAIpr6PKb9EZjt7NKa5dBh1hk66EGQ7Yc9D3Q1biLqKWnfl1PKdRDv0PVBDGu2cBurEUSmZ0dPL1DfMLvhTctz9HbzodiHPZoDkL30SPKohyZPh/ROfJvAYsdT528TViSzGmi2blchl/H51XajYVjA54coXFaNMgEwUiPktvVW1H2pOJ8I3tVNTtlV9SUz0yCQB9Y6Oj+RZik+vsvl1n7X07TYVbbae+9lhxBzbb1gOOCSKFD68eXTUMkNPpuwarLi4d843yBXKXdUa0ciWKvTWd4anRSgV67PVJo90Lj5KTgXTUYLvIPlbLwue21gUZWGtpv2WKulgcCL1X9NumsovBBdGbTN8/510ENoDemYyeJiN7aIm/ZwKSxzjTdTFC2LY6LY91JuO90vBiPZI+7mY9nxix7yrpkt835go34dxQZI01lQzhIx8J/CnQgGyzjLIgWUym9pNo37InRTGrlie5DDTXA55uK2adaPpzotbXDXW5UYd3wAa965E3HS2wuAMf4wdG7g3tDFOwBjqs8OtR0ZwLrsy0rxEu8alNE0djJriyQQMF8k8eFM2p0JAKGUhSfN7TROJOB+ewK2YFbuh1VOELHTJexC3ji3qqoQ6m4pWI8ra4NXy5JV6FNONSO04brhEFehlMSVdJHC5X6RGortOVMntSFLXrGxQeztAp2+QZaa2wjOg0zd1Z6UDp0XWW9pRtEiDv/1zMFhzfm8T3NZx59r75+uuX3/Sb6GV22ywKLv8Y0/bgHrhsOEsCuepsVJUpBPHWoVokRTbJ5jGVYJivk7qeIht7n5UZ5TmMbzvwiBo51zIRHAfrCl93CUNOs+ouytN78aXfoCERlyo0UaFUhKSpaRrRaDPJ9gS2vRuE/4sUDAY78PVGiQWDl/D5loWCwVcQ3meRoMr8NaUj5DeIhoSDwb9SEIWAwSvMI0SAKtO3EHmmJYDB9hZEvJMCwGAba3Sq5H/BNtbpUIr/BobbEf/bJhJ20rsmJfC03jWR8JPeNGkBqPWmiYSh9J5JCESt90wkHKW3TEJAar1lImGp9Y6JBKf0homFp9YbJhKZ0vslFpta75dIhEpvl6QY1Xq7RCJVeqzEYlUdZEpZ75ZIzMpvlpSs1X6zRIJXfq+kpa/2eyUSxfJbJS2P1W+VLIejaBpqbqmhNy7GTQsx7WYDLKdGa7xAX0Pd3Gdbo61EMhpvmW34IX3+w3DrOZlm2G1quA9J7tALhzh0+xdbmoO1NNxtWaHtZIi129epXYOmAshVYl/LaoI2Oy27sPl85u/eiY2rkn/8ztxdraLb+RpHilwmyqqOvHYw3C5/0nMFK6f13sa+rbBb1P7WRtb06S9tWnKap8a8joosZlVy264WyVLXN62l3pFhti+Dr3D0bIkjsojb3trC2wGyXyjBtrYY0KgSKt3DOifmidBMp6cVVv2EvuBesOOt3c63ZNvDah/37M6rJho8RsDmA/vJt7hXPnpRvtzyZ4TFHba1rwe447wa0HawvdOCYrvdLH3Dzr7gTVqvbsLvSVArrbkvhZyQ9dlw7NPFiC7ANV21dF5E63fPLS+p215Pt72Ydk0Y4XlX7GEARGpWoo2X4iGOo0s4ETl4KDTyTDx5qM2tpaRODlIOfKJqr23Er1GEVvjt2wbfO5Qt1zQVhQb9VilbrmkAiVCtp20pbR9ta+NH7UutgtrxQGnjVdstZpI6lDVl1p2OaugCdjqqoZu041Rj5Y1ey5RDjTbWRHZXL1JL7DtnBfXKkQ4OtCqj7qK6Q6JrBMbofTSjV33G5jdo5q8qrjKpUMJDM0Ub9svK5npqraLLdm7ELS/oySMP+r0KKny2DYsyvYqDmYX4aBfo+41TrYPtFWEzECjjTI3cD5pejxrDg6DCo0S00U1AsjKzyIs0KtDxKB8mI7pA3FBK9qSqJR9TkQVl8YF+HEQwmV3Jp1d8ptfZhR1YcSKlKLLzakZYio9mgjQiayVtaPVWfacocjUUHT3x0l7OQNqJUaqLAthUUZRR6GKZ1NpiVI438rNLRkE6Mv8ispDvJGRTDGB1Tl//ZL6hJBxYgnGv0jDxontLW4MziNjcd63U9V4QSJzUhS0YOe1p+OwXRH687iujp+B3BlI3LZ5ehjsam/gbEJ+EWzBpXcif8u7Ewe61seQCrI3ZnFhNtFbqk3DKmenHqVLXx2lPbQ9aB2BtzM2loYncA6NmOB6j5CLSzKmT9oIHU6wikJHl85sM1mXp/850gUgYkkJwfQqXdNXGppFbvKqJDQWzYxXhhw4iVjHCnS6xHUCXTLIgESAn2aJwJYWHx3oxHuf3oeEeO+cXRuVV1jPoavAEMG4aC6KUSKPI3Okz8Qbi0jV4oNHIe7tn+0XB1tjrQDw6C14A5VGxJ8YMz/qNwvTCu3454/CtwP5gddgzwoM2DbSPrmf9R0PZ64vA8LqFwyIutUaHojN5LEalJdrfd43c7EcB2vRin1o2D/aVy5Ogvq4WBb9kiVN83nCCGlJcwaHZcPYuQQcMRQLu0c9AA0aERDCaalPBoj1yIp+hWESDhSqXZB1FIsfzGYpMYhzl8RKLfS3DHaXiMtxS7BfBPvQYDd3NepqN8nE+EppNpCEi5wvabc7iNKjGzDQiA64UgdyefnJ7jCcmaOpLqbbZjfo0tCnTSWrIfQ6U9HgGcBqadqvIH9mvEtdZU44FP8G891tSCAQO/MWfsPRmWS6v8ISSTB5/Dezmc7qnoqdmrMCvn5+tgb1xEuu+8/K/nxvvpWO/6rm8Gsv9yZojVyE/IDPry3YA5uEn7QDM2qd+AHtEN8lrMk5P6LnXmCUAXouNxVRF+syPWzJQT9kyMUsAPFcXbpPB/hxrtXIvQiu2OMScLsg0njWDWSbmVuEXzGK3MH9T6RT0GRoqnaJgM9/TycrXSOGcFxVs6hvc+5t85Ro3DMJ3ILrqda9SSOfJz8353596XxJbVEIDmH76fAKdqPFYcXdPWH3PYV0muZ3+b8Ez9jJIjozxlSo2uWfUrO+Y9OVYFq6g9+maXKiVGYwJoMk4v+oV2a3iw+fXs6xGY31ChsdpLFkQrn9I0/m+tzXc/noQ9PDWYTNgMODf8K0jKx4Y7oEgx7+i4T3GBuDfbAEsAFvedWXJbU6GpGcp1zPRpzqVkg6ftHeo1c6hlG+oLtdQKPUqslLQCUUhcX9AIpGk3zclkAS0F2x1OH+ids2gqegdPc3ZxcXwBOhY1tfzLC+HMpKdwkEXCMxe51FY/1a3UB5nUv0VbqOkhpCMrXMUJOOYQP1OwonUpcwiEGOTyT0UH9P4+cd48dNPQ2on56KjHveZOPZZo9ogy2/29Mik1hlV+e2eZwTRY4l8LpSUenm6PYC4HSxsG//s6PEBaVhbSG10EbRRPu2BtU3pVjVaI9H3nWkhMxAeQa1tQaMd8bvNpNoRlJIFyrtAo+B/gcRvd9D5SM8pAVO2XqENb/LqePjxffTm+MPZefT2+BSKh411S6cI/48iactMOtlXeb61Es7fne+fyCyv1LMpJMGSXdoLQvNE5bOUeGrb7cKQpWV04KL2r+UATvk4XJH7KT7idF+v4yTO6abGiGCHkH2hSMV0Ik95hn9wst7do7P3tqTZNrsqhh+K35HxOxwvhy37mZPDy9MPEquNxXz8LP06tuNQNVJVEFj00zFyq9yBQrSFqu6rBAtaJIKe2Swc0y+/ptFuVRXjd7Yo3igdY7/6Wt/9mJh+EzRnRafld6s8I7eeOZ3ZjWoZmeXcWmFzXtPDyKum32771TfdXUb4vhA1ls/OPxwfnIe05liDRS5GdA1tDQEzxexXES/Wp/DNxz//+U+hcT2AHac1sd7gGAhIjvSCBSkvSLTxPGCehA9yfBylIJ5JKYBHR30NNQ6DHmHZezDqvzvcGj8OhDBq78FsgEjCwvYedP1F9DkSdu/BoO/ucGf8KO69rG3HUNgjwHDX7BZD6QxpDYma8EYakwZSOWAq+enWQLLxZerqGe3CEoxPA0o3EmD0h1lDXW9fKx43GssTbKh5EeOTrWg6q/Bxzn8uUJ91gfIPtH+sQcEqRBlsVep6FsXrrSOBZemtQ0IC7LQ8z1eCEPQ6uyesz8ojsy9OGmqbxvnMYjLISTEqckgVgZ5lJc7gDLgpfVzdfCmGshuzDA7nISjg5LcSW1FQKUQ5X/E8iVqyor91QWwnr04RmfuO1T0kjaWoYBLY76vqDGmI1AywZgOirn61Kb4FGtppHPKL/YI2hDUw+Cpn9vSnVZI8zPziijawbPyj1GOjszut2gmQQaA9dk+zGe7gzIrWju7GoXKuO8tu8+wulFw/jjr0vIT73hTNApTz+Crr8SLC75aoQXLeurNbP9R/Q6jwtXMgzA7wdsoPpRhVsQzimhzfLuZZOuQ6HlbMoy/m1SbUZ5Qhk/Jl8CO+yk+yUSzeUAc4cPAcDBW8ytLg6I/7B+dD8bo+C0ZxUaBYp6iuoGT50D3N8K19TfnJJYJoGM1FYFI2VYdoxgvAJ1g2HNOXQcJyTzgg02N8xJMK1QV6yjFbQJuobfZjett+xEidNJkYcBhBNmyFkhVlk6cT+mA5zhaZTesRK4cv3bSQIk5qKCHYDGbVokx7mAS8J4Btfe1zCIwCIw35+KX1fgbS6xFxVl/K0UJsBW+5ctyNsvoX7yXSxS2gtG+vWxb6f2pb6HUbCZcUwoV6khAgZgn7zULcDeGfOjaE1UUxZFtZ5s7xTy07x+oyEK6tBGuD+ae2DWZ1GQANSw1eCRsliTFCeawHRiiihJVm/rfiNj4b27FGEcxGkR85etiGJW7r4A4FRQnb5seOfiH1n2yNwdbwgzc6muAC+TmPLryizLekWrLqOqmUYrp2T7YFGHepT28l2TFBdnwghcQih4AXaMcGamBSWqlzegoQnhwfHJ2eodv30zcnKAdwZiyehX6tddm6mWJhZ8KyTviCJSVxdGPblhDD0IvnWi3h9YnDrulWG6VhZ0Tc4sCyhGLyGnZ5FzmJ+7a1nHDnyZiTan5NqGv7Rh/RUfYujDY/FqTM5fxlAbNhTGwQ8hLEAM4yYArSsKEgZV9Rh28KZnzyMjjE4ayYo8NZfpvNntXOzaLgZlLJesEytZgHtnrx8yD8tyCUN1RQ/b6TPA6HrCikp+RuIIQ5YmpePDMOuc8uSUrzpRbsKCjzlKvBWMijgPQxV4MIgY+CMc65CIQcylA3S08he601TIdgPOm0wPxSfL5LLuFh2zJzbOa0e2ccWjhwLcuLpZJ7fRZShU6Jn4d0/QbpeCuVBDJt+9gEsk+k4sijwdOcJx5lW3W6URuzKHXNMS07CfHyVC94gSpogSp2nl6CU20TPyzVn4Ixr73LlWd8fRpGY5lqbB/nHz6S3Bm2jv3zo9BzCmhdVh7UaH8cBA96UDzy8enBrOzjMHRLRuHyLy+Zi7IKx1XMW/jnbxJxtjJSsbYrbnuYiWjfXZmDEPvcr8X+fEa+Rqh65TUf5+VO7eUHnBw4QiMcmWSOlTdk2uw5LC9ZcQBHVqZAggjT6BJ+w/+WQS1Espa7bQoZLD4R0+m0osGkGRWfgp3Yl42ncWbDnlQQa1OSJnA7/yIYI5ciaxeEdpVaSrL4mQ0PLi1Zfr8296JuQb0XPh0b33r7nrntrd7y1tnx1EXRipsiicKSsl82L42aLLrv8kjzWVLvx7J5II8nq/DZtwuqimb0ZftFg4K34i/b7hwUuBF72Xr90EUveT01Akw1jCDj7nweX0UwxS21TG5OLJY3NHlTlfgpqh3zMoe/fFSBkJEvsfMpJQYMUb5E5UsGxoVrfR3j0f703Ye3+yeC6DWcE6uZ1MIynmLl5ahYpMKEKYsQ5FMrlpWRL4IuY5l0mL4Xw9AuB9dkOwLW5YdH4YGbjnyIXpBJeOZOnHihMTEeuRkEOWVy4k0WuSc6syS+SEjchERe06kE7iQRndjRArrW0LIrRULiJiS+1d+8XpxUtNGIq326cX/77vDIBCGjSCJMW62hmDbBAQZDA6/8BsErlLICwj0DufkeVQ0YYnbOvt9/b7A4MKYDEfv+6DBsXbr1Gw3a5dTDgQAB4llQ35G9M7Uz1Pa9z5i1tcasqwVfsLPKjxHpbUHvOnqbXLPT44PfI4Hcx6at1ZN89/x6Vi2uroMyH90gAV+wqVs5qlFhpnQP4mOuSiLUEfCj4I+aPurEVGEwzUiZtGQtiM76HpGZNZKw02lMvsMhTt6slOu6RpRh6EO0FXHgRQ6EoW2XeDjjqGy+YDbHKJR1fFXSstk6NqxXOgScmhwCH1GkWtMkWQ/RuMjucbCLVzxk5dPG6+mFLoSi+d4R2plPMGoGhawVtR3La/c8aqCwF0027Ks1vcN+O1bngCVWXXPktKCX+tVdyE29axe1pVIbXw08jILQ7OCrOvaKQTa+5Is28SXvh6TtMP2QUAL8Jvhmq+0eSOZ5RGj3IsgAePHim63+4yQw477EqFodl2B1x0stYAGus2xe92CBvd7Fdov63pOd6Wk6xDuv4g0ebxDEulkAmCFlZ6MA0m74LKur4jZjzHh7ySVRfmi3yrBOSVw7afHaKdAYUyqSLQuQjTUd12pIXlSnvti6VGhqISQJ/RZI8hofV5FysC5hELAnLro/ZEOpvRZAJDGvq3CgM4gxzGty7d7ru/aHcmkSSAO71g7IjTsA/oZt8lAz2s0e1Llqd51fbrRBYP3yVfefpFgvBsVfFlmN9tH3Gm2zBOf4joHsKXAdLOpSRv0gYW9Po5WRfgc+UN8Nh7Grm4ydtmr0Qx39sPnDUBk2sspVM/4v0nEWY1R1WaMlsg6iDYyI+n1FA7qQKkSQrAoAeBXGJ7UCZkUx1ri2zaw60VgNaZTHfIvirSRfgwu9bdTPRTFNjq97SU+30ZtN0jdMlpKJyadYLUX7o34Q0ZTaujTRmX+LCtraxiWBe8mokwdGeP2pEuewwvwhLhbZEfp76I3D76vZDVePVqxnD4pSj8+IexijDsAw2L+Fs754xkuUDx448KisSOZ1RM4rInQ9gJZHF5kyyeauZm0m29y3GitaZZto05/KJUb4//6fONOCUCequfhD/ZzT0L6YJdWSg9ZY1M23Sg9hie5s0ExhmYXy+ArnIjziRnVWZCOgXy8rodXRXZ5eYXfxskQucuXHrbIjRhlx98HRC+EURm51NYzrmwq4BIyhA6GeBhCDEoF678IxUE+bWYC7WY01/Jch7F3/Rn/v+WcS9gdOnoOzP5g5RvUtgM7v501IfMCuIc30S+PaVnCgqlHucsg9a1BnmALN5hmyJ/Ob4dHpYb8JAptaNpsjiMIrVsMIdZFwelX1EKPZZh5eKRuQwh4CQMFfy4afOzQRCjq5FyIhsJ1MCmep0b15EdJQr0N0OHcR9iBXP7z0QEN3D2tkCxmk34ZwOFrMUOLY2+o6llgcis1TOc1eVWfO1Vj/PAtssy2Kjem3ALa2qN1PqEOt0M8KoJzRtLGS1XV8lSXVPXBq2Lu4vIU4VgNa6qAjx+FBXJJfWHyiLuhXsWuC4Q/lD+VDphezW2CjEhQEVOV1VkyzmfJEg64GxGgQKbjrsd2JulrMRtmsuisXwnhsCB+LnIxmc6IO7ZB9bWT1ybIWflGgmqWcldJmFRuyoA+thUdWuKWcDCKsbyFYhCh5LyFiWEpkmbZXBs2J0yU/LxBzYT6bLK1HOU23CiUsnmjlma32R1EN5yC28OGkuLOo8Y4TkWFdTMKuyNK0i2/axMe+JBdWovui0TVa/mBb7OivCeNrWLMNM8zasmCDj0vH5gJn5W+DJXdCfGtxMdLUtvKy9X/Lht6lWZDE0VZGNQMOvowLckJpdDGgw3J6AueexKPdPbHwIPSshIhTWIBh30pqAtBAkUUqV0B9s0e6MkJ/NfNcx6kyDU1XSLJ4PkiT8DHk84udgvLC0EBUVPVcYdJYpaJHT1TOg9ZMcLGWVZSgNyZ+aFtIq1gdq4TVTHoYauk961qSkMbG7mWJJMHXZYvcWRDVy3IUCZcCZIcN1sqe+FVciqyy9aqWEyMaPbWTpe/d3K2RZGdo2+tXmBMUwgx6iYvbgpQQjzNJclrSaCaJb/lMlo7VcnVjt1XWwmci0cXeZEvBWNqLHcZjJqjKipUJQYZxmtKb9zZqCLaGScGvzLm/JjEw0UlRjW7qKFniY3jEwz1HD5Cv5td7XwnyMxwba1J2jgayrhnpXKPDO4XB6DaUemMCXe9G8C/sX+wyfs3FcAG4IQtLTL2bAW4ScunNraWXoeUDk3GEohg8GcCqu9YyC6eDw3gev5kh0+tZm9Tqp1biMfvN4772ZF9prdLJw1Vn+4X5hAR7xFXFSTLLbjWHNVCu3vbCD+++D33HnSQmE3gN7lRGYDqx9zjY9EnAvm3vK+PP7U0RBYWhOXkxsgFCJUkfdVgQ1d3wEIGTaUrDaMpne2GIGu0y72/+Od78Ce1N4/kJMdL73ClXb3S9KG8cF02ksgJoCPXF7leX9lkQYrUUyplUhE2OMoTs4omb0Be7O5dSnZNzcg8SmViSypm0TI/eGxCQMUjIhq1hm1rTQAhyiAyW4qhVGCC42H1FyxPiomHq7wSFw5xQjOdid3vnUko7s7JezLKI2UvpoZqfY4TRh+ru4/FhiI/VUZ15nN87xhJs68ly+nTwOxL5CrdsIgOXifKr1imkataYQvxYX6QKoSgmsG3OPBXOAu97X7EhBtKYzdBRRV/ay2BX3FRhOpWNwwdG97j5kAfPg+3drQfC9Jg+hq5VOInrcsNuGNMdWdmIGXiiOgct+3fr01fJtlLm1kbVdCmGD9lmdjpZHa/D6IwPE9JlVNrskdQ4Mit4p76ei5n0QgHT4dAAt+ghhe7VHewdyMS2LO6d7dc7ubdR2mo27U1iVDeTO/AaET/BOgoDQzfP9NM6CChJFGElydf9eMimSyiy0oJ9j4FVbZa3cLY1m8b8+oyNF0ixQJY6stlFwfbMbuStxi+gB9/+14jq4kZWSRRI5oJoNt04M2iIBnkw2+UQn1jMo7xMs/seHlHEU96NL4BV7vwHEOf7r0+Ogvcfjv5wfPT9Gjk2pM83fNyG7EiRT3IU0GEXUnhPIItgmYlOjt8en38CmzLAax2Pg2DSeBEHBpgiK/ZwCb2lVWKhgubA4givqRm1ggwYj/jxseSAA82+S8tA+L2qamYmu3oU+RvhroBb8Js9nfTp9dTweHDpKYx9N79KWY/ba6vAegPw9cfjk8Pg+DA4OT47P4Pt5GT/4PfBh48nR+uMxnVKeAscdr5JbLR2NCekIvUaGPjAh+ca3vZ6sMuW8t4T+f17nzKnvF692C2JU7nXN5LMDiurgREaq+LLn0QeQRBt4kMrPHYk5LDjK77mQCbQx/PJAwSm2w67TIviEyQPH454MWblJD4aCSUFcU6KVbzQFOIPbgqGhUEpU71JRUubNkUhTctKrSIPy6DrYN6kjkmiL1Mu8kvD3TCrWGl7A7p0hFPWTXS0fFmqj7XCku5Yy0G1DpO/QFqK4tb60PWJppQG0DWeNkxHjY0j8qSZPDFOv83UwnCukJCncd05gg46opUCpu9Gci7gH6ypUZjoUuYVLOnM1D5WICAd5MfhyVe7D9bEKqaD4KuGYTgzx0tPjpeUw1LXai/zjVvmeFWZb156cvjKpP2ko7Vv/HX/uQ17K54dD6IdL6KdBqJ0Ub0kTnSGm/aFp0KDwFOfywaWnS4sOz4sOy4Wo1EfkTpYNziYQ0MotH3ZQYWPRAash8qxI3JYvUKd0tonxy4lB9tNOmJcS2eTQ4yuLn/b7CpvTzUjJ1NPuUnaWtbr3YckdQnmmX9uFV83a0iD0ots3I3szevmVOlC5psxDRK+fgoJmwXJ4Zr88vHaHIKveQgmegwmOAhlPayua7paW9X0P7V1zfITuuZPbV2z/PSu+dOTumbZ1jXLX6Nr/sRds9RdsxRds3S7hm80upbus5aeqC1UkrWRMjM+J1qiJwFh8l/MedHpjTdXYsIk5EqZdG0LpZ18yuSsPFM6nus/XRKtOGl62l9z1Rmixz8DMrjJzSMHri+3tuRjLMxiXT4s6kyQoqYTkWplhHK3e2qlqC05/O7ZTtno1IxgOGyBf8QwWuTYWX1HqsF/61S4Ja+KNev8fC/QLtaoJmpMYyOpAGOoYMtim+VlEGeRwCMlxF/Etg4EZk909hilcIOgDQv++1FiSi59hn5R6/BHv21f6iuaCD0YcT96/Fr5NQ+cvD9CP/bt2SDoQ0Awyw1ywtmRHrAE0eHZR+OtB0wTOJlGQNJirA9hfD9UjFFNLaMTNYlkmDalUWECmqE+Nhw+ttAZSKn1Dcc5mmchxOZrkrtrVHsw0F/cX6L6/72jQWID7FkRdqLdAfcO7P2le36713VclHAAF5Ukq7amVF0iooYYBmpniZWSWDdwM3oYMEt816TWGilJdzGLL2GwGN/JZTspEBjR2+8UbGS/XRtZQsjiLkWeLnBrCFDRNGuNAz6p/whlO9LpocVHOWbR1kWgEQzQrt3AWVrFMKKAC/TdRyJOhl/zhk1iX08C8v3x+XfHp5tvjk+OgsOjw4/vIU49GA7IvueA5S5nRwfv8Or04ORo/zTgpyCfSw5z8OHd2RmXH+wHt3Xw+u9TDelCg7bfJ4iAhDWtaFRkcRnFSZSO5S50VVRJXEDn1OhXRwIMREQd1WU8ra8rMtaTLjeUIrAv1aPvbGG17xV5tUjpcZd5m2JnGgThvmC+0qQTFCodvtYPMdLYEhJIzY80du8gqEKSbUhjjSDxI0i6EdhGiCnNpItFEFet2b0OaoJP8YanHMVzutbESU6aO8yEkfR8QJuTNGrZnMT4ijRFcZQaFRgjxoTh+Qk4CRuMoiScucD23Gr6yObAMA0bOy9C6yuQvDGCLPep+jKkHdA1FYPqVk0WzmiMII+35ylJXOc0aipSWytopzeZDqNqP+/piyrK1X45I5M7LmhMhlH0qq91Iq2teTK5tX0OwLoNFNlaW6jT2+7k3A5UxfjcMHDPq3svZ9S01qIJtwbB/TPYHah4GXYxVfd5mkhK64I+Ly/NW2E/C4GjQUgzrQtZGiR23YisHlhBbsfnUaMR9h1oY/NQFflZF/SUlzyN4pjfRv0606L62C3Z8ZJhbz72Vmgser6NkJM/7zbIOI1nQnAkn1IsruUNLQJxI1FDt8uY8N33p6F9fFx1w+41Hi53rYa2gZHgKB2wzo3FVxt1NVeUhrCocadNmj6t6wTrAa2ns9B9jiPNn241Bn/mz1np0BHZhNEh9PgZ9uoauEUTdJZLfmI0Xd2IlVmtWr2l1+q/Iw30datl5iH9ml5OLmDQGPa2726MBlHfcqUYUvboHSm7r1uwmUcW/OUTy+XKf0C1+qe1l7LwI4/904Pv3n0I+ZIULeeLuyht0ohqcy+ROzW3K6Ts70bn8ZW3Trzpci3O0RDIOpW20WK9zcYQGqceaGiE0a9JF50D0bv2pFoa+YGfIqxdhpuRitI2i00LYqZuBmTecM8dxsHDXoLNhdpZjOEU4luLn3zaoGIh2ll00xim6kk1igtjPmGkNcdscB5LvNBEmvxQr2g/bD/eGGep1tYnbutfd7d+3aMSFZsmjdYnvtYnra1Pulv/Omw/m7Fy6qyq61YCUKriAFZtySQb0KRg1D5qUIpLEFkTTnTIQpEeyoh4P3Fkphb6cHX99KGstgas//zKtOk6vTKZpbziE8+t9AIK2D56DuWqXhAHxRU3AICLor50mSQ2l2dxShTVzi5Z+vuN06KQ789YB49R5XN8TnNXu8rLV/yGgBRUZ/ywxNM7xCmI98gSyux4Tm8c6Qi5WEObjxGs9wdX1gME/cDHJbIcDgzPbys9hBaSEl9S0ppkDrExvmMRicrHtCRuOeCmWVdEzboivYD2e9vmrZGN9wLx4Ng3JOaRCQPo0dmegcHqsepuvT5TcF29Ju1l2FWkjFfkX/Bb+te3JxBM5zeUw1gDRNKQnrX13AYNgvg+r4EsptNuYFwiFCBbz61v5btpn9VjKTCmF9CC2/GeLbab4D67jkaOnWYOx0KSAfzSJeC3DQIhWxZ9gNZZ5LGZNkksRQmtdGxwPTDkFHUbHIlRImWhgeAr12CKRLFFPElg+7/dDTYjbc1N9ohPCXtFYVsbLbzj2i1AlPGVt/o22ygpZ9tk0yRsr7dbxLcemaSGdzl72DfmVSScRfRa4PQ7xlGFL2NhEI/zoijjHs0lKaJBhdEVQoj2iiAmYYatIhv29PrPmZiDwByH+OV22yBwSDKwT0CXTUXwzs5UlVGGhGQ+Q5/ekNI9CZPI12/s5goLZeLroF6y3FM4BsEN7Ox7ISzOSYH90abovU4dm8LI1YPbrCTJ1NTArW2jzPUigZrtXZgndiXcGzhqFdl0TzwW1faQO5tGGpbjK5NTEm/gealWk9hYvl39UmQzqjur2T5Og0qBSHevsJS/rsjiMVfK0RGwqmXpBNgNGV/5NlrNQ5p4GtuYKzEwMWx0T0DVeIJOlj0LWX84WkxG1aIEJg2lAN5y1cm9pRR7nb7fDdY44PctmaVijI0RSiy9R/d+xQ2LecHSFN2uI8Tf6LxZeapAf2Odexpu/8pbGhtMz6XdZkPNy5rPcl/guxJ54nUBV3/VZYGCetpVgRpCNq3sawKDNF3XBE+QtqtiJdOOdiVR+FtHsAEbQve8nhbx0iaF5DJxodUP4Gkh05Zgbb7YiPfw1Uaqtb8a8cY+a8Q29tsNbZWmm8BkIcHYR9hUQsNWg9k8aa7BO6qNzcOxsuAxLezcEFg3Feaq4LuqEOmf965CIH3SjX1ZCbPn693aa/D/QDf3gi5rXtxr6E+Sf7iMkY21nT3yw3W0xcngMkzdPJLFF3WzQ42CV13sCUjfvZ4coUIHClcrstbLZncFTypN4juW8E3fWVJDFK2NMLhyicshzqQid3Qk+fLWcZxlTzz02VHWA9CMal5fc/Vse2h10HCddOvYHiNDY8rKGP4uigJ/F+VNWd2ReR8qgqwDsSF3BkfT7uGjfveIymVMKPECsnnliLP2qY93PWYWJjCl2NaNNMYpYsgcp6/g5tWeePRUFfTKm3Jb1lpMk80eKE4JDbewHiC272L4IwCYQZAsxe9E/KYibzS6rqpaWJ9VJl6wLQKrY/dMNkA8/VW1IGNTMtH7zleBrn7m61qsIM3Pip732ne7ItK518VYnb7itlSA8CDifkA4ilYtkmy0Pa77TQREfZVfNrkru7Y9jLQxDELjJ5AJETUJZRUqxoUqViPoKtgWb7QhDEPHC6uspf5ar5IEr+uosjerqFzpmI4hlrrgiQ6mOrj01cXM4E1KP6kFF57K9sbhw+TxxUMK/+GbBuKxloPJgPyxIKfN7aZ6wg7DXxPri2pzKS51O/pH1yS0TBYjkGXfhaCNFwFixRRTRK2KNI2MdZGIsdbKyKUDR615KZTNe/TRWIyboxRje0O/omhKyXPVHbThGjUQY4fODYaRXTxvytsFcxqznoL5ONSGorlqA+Ute9uYTcL6kopVht+QFBEL38OfhWEW+b4GX+lfGucmMXMZ2q6tmJMD36tQmku+TDxQvHnkC2OUWIQPspKPPz/oOuDzHIn60c2KIn7jpYrC53uxYuwgGrvn7YuL1KqYVZkVpSjIpxZitT5cuy3a296nFtwsTE8I88ENIRGcoH0h+JSHNosy/0vU+i7GvBCcVnWUD3hS6ksrld+1cbvdMkcJHSbOtlfOVOVU0oH1zFeaMx5Qc7LYHeEBNiZJ41nPj9hu1doLogYZ2bn0mALfMRv/Y/NlzxgBZjtrEUA5zXTgW4hgjMrWcwOeGfDAgIeFnb7/iVDzWZUis6cyrWQ2qmMAcqWE84XeBN16Pb0eiacebT34BB+eqwvmtXIOJ7pMrd9eFHzFKseYPF4at6oySZqyouvUQQeunXZcOz5cDVSXfS9hDIN/bfJC7/JhXsNH+lWZvlu3pkV1JyYdiaOI87aFUJBEW720Vj/qu1ZXsZPD/mXTuKerXmIUKsZJa6E7v0ah2xfhGebT1wGS7+mCVkJ8KZzvBDZ0xchTjeVHhz9mmfKh4yXNk2q506hlw8GPD/qXV1OJsbD1/VUQO00IxXM+V0ynWJUw7zqnfl8K5u07Ak0aLB0CTZFuSoxElCl2s7SNTGWgFYJroQ1vCX84znxOh9aQtF126+IGFSbQKvWVpRBESgzGsYA1GjZDHfZrNmy5hKWS13xLtyWqvJjSc+Z6MZnEs2VERowUyxPjN8naZMPgYKPubaOYwYUUpBs2MWFjJG4X3hQvABlrJ2SiIWNW1+/CSgAS7wroxISOSWTYhRrTJeZu2ETAbihbXUh8gGdiP2dCbshHGTINyfCcycEXlPQ2QaZyY57LZm2oOzAJQHV6LionrGxmKl0W9Fyj3dDWcZFltZqgas2bo/rs9zsyiTI4i/jozKCqwlnUZ2cm2WjOI7+6y5GEEOXIT2m0D2XSIkPMbRW6nNAIEaK6iTCW6ZjKZHtm3ncO/iqJiWWQt69vKGAgPAkVZFBE12h4xDwJEWUxOkMjo9H1JFyYg1GR8TicEX2Pmd7WqmAn6fzNFvMSi/f5beuAgOivWi9XVAHWaaGx2LKSWoloqIJUDVgyWveaYyz5G42x5PONseRzjbHkM46x5JeNsaRzjCUrxljyOcZY8vnG2HpPxFEqG0ziMr7KZkE8wjqu+068qOI0Io0tWCRt1iwd42yJ0ZfegOCyNJpk85j8rPInVzcetL0ld97UWVfMVl8aDnAi9v0Scx9Ic4JsE0+nKvuEznHadjrjcQ9yF8/KvLzqhd9zAI/FZ+S3h9zbBPt8TzMM+z4Fce21IJ0JA5psVx+5TyCXYhKpwm4SV1qzf8o5lPAbJPzXeZruQtgO49juMV29QW+yQxi6fecKmG7wBlhz3TTR1fEizYHjvsnxrBFpGxROvBZiqxLX8LARk48NJ9sCGHW0NCi8EvTM96esIdW0cq0tWhsugDiDGqnaDwa7hDBGLblpwZzolJgoEhJJwl3tPM4mOnWPx+wg5OR6Q1Yg56NbnpgWjuVXlgg0eBI7T39j5TNiV5nF0U9QCPxuH+R3LGMGxk2j8SK0mi7IhSUsKwvgEuVlLvsuqHv0NoqUpIzcRpc4sy4vx1UvPKGmkj8eMde48eSC50BcLgUPTDSF9fGH8gMeR2WCZSWwY/Fcw3UQVi1ih0GZ2NGNBTFpLoiJd0FMnAUx+TUWxKRzQUx+tQXx9S9YEJP2BTF5woKYrFwQk191QUxaFsTk0xbE5G+xICYrF8Tkb70gJt0LYuJbEJO/2YKY/MIF8fVnWBBf/wMuiKQ45HCI/L5O7bkNq1JrLDGnVUBIRYtJ8MybwjA4KHI4MzxD6gTnBLT/jFchFK19WJTBezbpbi5KprMgbAVXWdh+71nM7DyfF9leyNbk98NmQxO7ocmv0tDXvoa+/gwNTfwNfS0bym6YlLdKq1ubTp8aDEP/Sc3/jrPT1DA6Wa4isXIuuszmba0UFcXVtlEb2hryvlaNaEKMFjNGgebo+pZ7NYl6t3ueHhLFjJ2RpyK6v0W9Edmma6OtnrZohWnNFw95uvZkRYTymnUq21ijXhxKg54YB1bFdn8oQzjthvAj9CtkaS0jIllrRCSfc0S8/qUjIlk5IpK/9Yh4/cQRkaw1IpJfPiJerxwRuMvOIjEuJvHspnaZYXscN/gb32C3FJ4ceCvN27QDrJJuWUC14opiM/eDfw5e95HIa0lNju6R6ajXs9M3yybVLRRSoQm+QF0w4RvbXdhSiX1hEwLZ/XwW11FVFsunZVR2+p6SVV91MRIj+xqeJmJ8QTKrRkDsda0KEps6qdKM9eXwACKczsbENAoXvQl9qMuo6kb6dhSwroczcQHjwiVeOJiohJKUFSDQuCILkwp42gYs6R954WOimsqRmDlifwlGjsatJKvDM1crZw296pNetPSbk9B3fRAOwqZgvBlJI8YTjbLHZjQyxeazGZ9M2c2V+ApO/AUn/oKT7oLVFZSbT1w0udHqMslNkDdGjQzyVkhUwdiq6AC6J9YzWN5o3zAUgBtncqkBifncxybXcR3P4dCEiXiOyhp+o2WR7Np409gM2uXL+tkWjatFKR6R55Np0RNrAM3BXZo2AwLJy+liXpvXDGLJHsUlHoSVm/dWYW+roCPulnu4guGk3dyoLRBRX8Z7ota8tjAl8eRN1jBz2iWJcXJatmGsSLcxZmwLrI42dR6sEhvPufje1+k9GLr0cMjQmXAzKhCPjT3n+Kw/jS6wdS/MLrBzJ77cSWvuJ5zdzbxGN7Qm68Z1pa/M7wNw9FG8DuT19GseFQmVTqc3wAUyXZGONa2sip1yL+jpdPkaBlOkyrLtcUnlTFpzJu055e0Cmwdxc4vUVdmTzuwdpV+nMyqbbuWczJQGzN8W5d6yMyUdmRJfJiwZCF9kt1nRktcEARSvGMcrA0lesj1M0q6UjIuLxgICPPQyzhQLAS9FL3DbGSvJU1mKR5TPO8aiRR4J/rkX3uljl3ECew+jvM7EUSeI5wF+q/svLZ9YIfLFkpDTAR7yCiZJ3dvqN9LqeTxf1L3wuMznOT4JwYdYw6GjjIrQJteE0itzLpA1+VwVFI2S3hStfpRw7hAcrLLGLxy6qa96muE2Buw08Lv4YKtJLqsRAnHfCyWa01RdDs9gfUb0rIYePLjVegyqcfBgVu0xHHjwPIjyd4fb48cvg5+DB2rA7nBn/Ch8SryosxEkHJ3v7wYPY3xqM4/mOWreqTa6uB1aRnScFG9zDY/GPEWRBSKxuDjU0fGO2IGInM1SCBkA9aBvHrtkLejhNwmurVezOo1YBF8y479GT4dBk9KyHiYH5rUpravJM5EE3X1+qyka7M1HKgxOXhaLs2EnJRDf29NicmYbGyuaiRcXGRevEJoPgk00WkMo6QbFvJHg7nN4UEWj/5+9d21v48gRRr/rV/QyZybsMUVR8iW2JszzSpac6B1b9pGUZLKKXj7dZEvqiLdlk5YUWfucv3H+3vklpwDUBXXpZlO2k9ldz25kdncVCoVCoVAoFODrmIzyilZlDrn2QIAxWVWuE2qVNVRyzFLOWXEVtOARivpueLIahnueEu6ZU8s5YzFtmpMW3aeHk1NuGA10VrFlhtY+33ZnCujo3gmHXhjEHlPtaltyP1wi95SABhs0iC9pg/dEtMIDkk1LXV9jitj5MiVRw560aG1vBaxGFU4mOp6Ac6TJ2g2EHHd8AXpJOfeXOAco5M1pmOqFdcrVS+7XAvDM4b/85WT08BwAvJN/syKeixXpkiAWTgR+M87pw8d5t3ScUzXOKR/nNDjOqRrnlMY59YyBrcqzcx0IonSc0+A4W0ecvbTOOKfeOKfeOKfuOKcV45zqcU7Lxjm1xjl98Di78xlXFRQNpZG4qzS0EHNgKJVoXWmCYe7APRIcG1HQ90QG4MEgt2idgCUgH9OdfxQCTJVu2Wpzy1HrArEJKsLSLzciVMWb0P2IdUdKy1IP4/KpDP6eeBoLgGLl9QlveDPo/4kLA8H7ZHNe8UL6mXhhdwkvyAQAqeSFNMwL6Ufxwm4VL9QwCtXhhTTWHanmhTQun+6cF9IgL6QeL6QP4gW4Y4IMEJQFn4EpXmLsYgytJrMiMcawb5KKHaZYiSH4hwp47BigdIAi88b1LIEe+VDTcqipBzUNQk1tZG0Tn2Pbcy1HxNjqwMXfobg9b3lYrzQNKjYCYcsYfywvbixl1vOyCqlTwSGkZy831E0ELw3COXhMobS6EBtc3ueq5EOsXFmem1DRpUmHAoVLbof6vQ/BgDvRgfclEecoOK8VlMkKNVeBSfoZMNmtwMSPPn09JpABRPB9Wfw8nniKjNYVA48FrBEviUmo3apIOIQZT+GcluCc1sU5XYZzWgPntBbOpGMEEGb6UR20uTpVgTwrtqQLvGTtjqSVHUlX60haryNp7Y6ER8Q372FgQ9wcgeR1ZsIHW1Z84CNYBSploFIbVGqBoicfLRkefTjPxGbntkfupb1i1pcbOahVEoiAHOREUSfQvngjiScjYWlAaP3XxNTBrrBGpSDll3ex+NrSEijL/pN+l3BIOy/ycdN088wK31q+96jSJn1aohHfjP2yrGHpatBTC3paAd1Z+sv1YV6wFri0LjhHc64+/F6N7jWxXhlfDemr6NUsy37PTMxpdeAq2Hrcz9DNgwJgR9LtaoOO6aM0u0ze55NZhTtzOGdnaSKOymSOftjFyojDpefHcF61Vm9T4m0ghTRoYxxl7qkgXcalX2tEhfzTIGtTstnp2Jsfy/XBGufJQlAoB7+OSdEGO09bPGGQRbLqmXsT8g5VsFgatEdJ2D6HWo1eZPP+9aDpZym0d2zbJfsGhnkhOHKe3cyb6gWU4H2JTztnYTBpfTBpGAyZuyn6FWF239s53OvRQ3rvR6SYJxD9sbGrna+sLMq699KMtV3e4sf1X6KxYxy0lmSMW7HZtLLZXWrWNg4k74WQAVCgv0jeFTyDJ/6Sd1oSjxYAalmugSo+rC01p+SpC3gB1LvgUSKCvO/dCZj3PUxP3L4ZFjclB49W6eMf37zZOfqlt7dz/MPu252jvVBN2zB6ngyHKUWppq6eTk2oak4EwQ5TVKo8Ck9jpRHYPbTpPSqQ1uTCCbJ4Cg6heGniGJrZlqmYYbywHdmZSHcGP4kKJ5dZ4CMEqUonkyu1MS+iY7q424pe0j1nNI61ZLiW/hVZjot2w53zNkl85oOePOqCH65AfTLKxOjPhACJSFReZzMhoKbZGOODTvpXYDwvJtH8Mrulj0jVCIxsURKNhTyFc2AhgEdT8Root/3reD2Srr7iBzn7hsiuB8pGOS73LQCvXJDcLyX94R6L6E9VttJgfbUUHKH9jL/RIxsdZ5n8HqWTm6g5nnASIRXI9bfGHZclvlI83eUcrsGg1HqJ3ldD0cnq4xVTzl7TwrTgNffh1k3AxIruQUJBEcXnl4CqTFawbV/WaUXXSW5iSWNwAojXJCv1SKIod9Qajkgo1LAuuGycUNPEnl2FQYVHogxwzRwSzU8wsibZaDKmQNEUWEqFVIBiUncirxhMvfYHIo4kfAjm0kNPen40azhfBnz6UOGq5b/9k6BQ/YTwSE9SLJXyqjC0M1L4ymeNuNBBh/3DSQRRoeQlKJScFAu1zhWPJffYgjef9IUnEv/NnUe7sXW7Cw23lb13delP03eTs/KP6b/OVXIcIAJuAKqp4O0RPg0ZYKlcn5yv76HR5A+ixauDw72Dw+8DhECPTpcQ4Vs1eQHgUahD9DaxGF/kaPIYyC+wcs4vxYvzxXC4Tp/fi9kplh9cl9w5mITjoHErvPSs8d5/8unIbuV95vnoz8akIhe3RwXz9rNMy89PB2teBmZlCTVcr3hOE/fbp6HMq3yMXiJ/EF3MHLWJIi0wVXMlLZkr6WeeK7ufea4EVq60fK6kwbmS/gFzZfePmCuhFSxdNlfSirmSfva5svvZ54pDFOsGILu/W3Jgqrpcek662gyBFMNqZ/VJex68hKOmCrb6Zufk5Q+RZhda6KN16Z8QoBFBUSQKEc+r4izaLlmZA0OQrvL7gwlLMmHn8xNYOz1YFFbNN3corju4fwiVZz6JgjzoSO0wsdIlxEo/kli7fxSx0iCxdsX8dIm1U06spGyNs8qsrZSbyxnR5dm5nF79i+TnqkjQter8eVTBFEsugdVgF+9e//kDkoBVISPj+HycduHQo/lSnWT9OJasGte9Qo8X8snIer4YrxR2kMILkBm/RthB93op+Wx/ltul/4o3RP27hlagRIrS0OyIQb5q7x/u1UoE6YfTdIMQ4j3lxgOBIQzYW0/AyYRSLFBuoLPaaSptV349FT0/e/eL79GPEaqUl7cdrgoq33/q+7Gf5IZryZl9ydn7x1xjhRRHtWP+eC5+fsCdMEPWvepeO5iwjINr3aWvDKZcq6SMZFuvrBPndXnkZKtoZVjYBwTzcieRI2nrxDPUlz6sWxP/Y+7iV0ra9PNI2vSTSdr0IyRtukzSpv/K0QD+BWRluoqsTP9YWZmuICvT2rIyXUFWpivIyvTPkJX1Im3D1U4CG1Fzs7qRk0qxI0T6uFsY4N+hfW3aGt/+ssO68n72K8AOqk/ByoEOKoAOlxwqlUNlFFgjv7ujnZ8jPIKl23ZFNFoU8ygZXie3hZiu5xjJYH6poyzSzpYiLxYtCQS2kFCGHChFKZrnslQ7OoEDHEhtXEQyNmS0IX/twpZV4TK5LpoHgEwssJnNoOnsfTaOri/BVWOc0Z5IbGCBS8S/83yWDTGVFDqLSCfAr2BHrFGZiK5Nzo1joUQN9m9FO8AHeDcy6Y11TFWtK8eB8Jz2BZjOMk5hwDtlbae87dRpO/3ItlPZtt84YYWJAezA2ZhrwN4aeI3rji1BghqxS1tYpGEsUh+LtAyLdCkWKcMCSq9Zl/8BRzCjwL//xpCt7oWunsrqKa+e9iqatwQ2dU6GqKSgV+AtI5pxc7bwkqkpmZaVxFkuA4Pp0kBp6A5kwhnH6laJTDEFtaQlC5NMz/pVS3ANo39pviadoV3dHqrI1VwJhCV+EjhbKZi72IG2WOeacV1lwOSGoilEYpfRpQ+E0RbAflpSQFn8Bj6EgQVhkJYU2NXpt70CQwvCMC0psKvcuMJ6DdWgdU+zB96tHJzTdY++HRxEOhx3QmoHAYPVzgI1UKAGq4LCNc6CNVSwhqWwqjvJZlefXxwIFmYTrJ/GlT1mcAdJdVEGdVABVV5XMd1PlpRlcIdpXE6Ii5mQVBwyxVIR0mAAf6CduLpu6tRNoS78gXaX1C1tFdJuBQDV2XBZvH1DepJYNiezHoqi0KmXFf7pxhZhN2EZdgOT8AYm2s1cL+Vhb0RLqgmJdlMl0m5UlK7mjSW8AjcUPUGmYaQ1YeyGYYSduU2PbXmL/cflWkzJG5iSpfPIkNyM/E1SpwJjspu0Vgu8xvyh7FP47BPMu2gxUGEzUBFmoAJDwVFAtxUZqKhioEIzUPFgBirSmjBWYiDTY5uBCs5ARRUDEcnNwBbJkrKMCYp0SVlTsi67gI4k3Vvt5BXgrtAMJbsM5wRqeTkx1niiUTsTBobeRs2o6aYQEl9Oz9y02CxTaT6Pl4NJa4FxU6HYSU2ropUfTqi3QtLrk0TzQuC1EJozHCGC3z5G75f+UWK92DAeILUSp6hjQi+BKr+gks1mFO28oXIcR41jFZQMxn4oNqXwW6aVNU70DTMqEI8HYLAsvwY2KE62KnsmVjVe1nyhwlL31SCcRChiUkKDZ3xVf0+hs2Rim0GeDCcXYr5ewZekgHd4HWHNuZwK6eCzm3k2BnfKbiN0EwXqwrQvuqdNIggOQgGU+BtViM/sKjmFJISa3cYO8tex5C/Y4YdakYeocMkkwhpRU1ZBo0Bs7n/ELvPpzm9XcgNe5RDcgF34eZbPM5Ag71UYtWx8kY8FBnApZHp7M4TIL0V0jeVswPOJDMhGiqFJK9OSxSHsBJxBS3d3E6AtRIzGskwmdBdH5jIh0mTyVVQs+uAQD56xt9u/ju90f+4bD0lKUsdS9+NBjWKzCSa+nV+1T66aMT62cYSbjVcLQbnjRdrby/YwCs5hdiGWbTHo88mr/CZ6k80usoasc5FNRpkQoc3GZudp5+bFk476kpzD+D3ttKLedDGa4h0OYHfRCy1yJX/3zidjwgZ+tGEg5hP42WycXO1RmVfiUW3AWC3wMDjPLxazrHmejPLhreDQ7GKSCSqI4Sjy37PuZideC68UtEr0prMczapiyk7gImPjq8HzQTLoN9ZkzqZzEE1ELSWlJvNWlF507bqt6DLLLy7n3WdPYqtueypUfrhCN+w2boCfhtn5XPIlh8vrBODHvB7BLPKBYFp4Lbo7TQY3orv447b7jFMa8x+8O3gd5SO8SXQwEmzcon9OruSJx8VEiSihg1P8IXjWd6bwAlVDEC2bo3AZFNeDHlRrT8cXDROiHi+MyjrypEADZ+tWeYMuZAW4BtB8BPfjsF9tkBSskFNG/G2LZgWLNJtPOoJqcNU2trFj0E6u2u8uJ/MJPjTFB6co2rRpMF/DzyYbKyFvoFZXwQyOrQ+ND/EMGEuNcROxjcv4mqFOdmSc2b1hauFHRRmSpA6JxaarZYBo+7ZaDlAtfyoQ1110G3vJ7CraHS5UUZi13SafpZtPMJbXcNCIW2uxQdbjcMHQcJs2myfTS9Gv3igvCqGzUKdBgQl20Zw9yYll7GPY1yYCm+d98lcDd1Z9s2FdLBTFPBkOo691s1/HbGUs6bjq/FffJE8HnQ6r4Hf/heh9LtrI+9B/s4CqDpVyAQg21HFHyRXE7ZwIJVFssIZ4pQRlVEedsCTj90lBlHmJv+Vn8gKlWvBZfD/GhzSZUZGWUDfzTODceJ/NBJEgUwJY6kdC3esS3PYt5lhhLTGpfEvAVQ3ZVhtjmPLWQ70kgXnbsED7Yo+K4c1yvD0M7eAVNWkSFBTP5Poi5SyBkmTOydRLeM8yOKkS7waT6yZMsw7cLsTHrgQEV1/7l5NZtzG+bnBr7ASul6l+9+BUZB5z475DGer3LLsAxU5+TSErW0PwWyO2sW+nAodm49uXqvp3oq9Wiy4mCJAhVIKPUIdGVKgpugl9Hcwvu1i4jb/jNU79UjSc5hxsRpNFkV0LNWtYhsctXUtCmqC1aX0z+ltEpeGgWGw9NqLNLSH0xHRZCPVV62UMM4yd3fj2DbT2M7Sm0DPtxzxviSTu2hrEjhZcUDqRRFNeGlrkp+P5TAiMn8RswfPbLrgyePkZS0uuzSdTOe2IMQGPGN4Sm+M6jtN8rmWa+NiSksvo+pDJrX0xywewgep2WjIlE/wqxKS9EmrRtRIbT6V+8DRes7IKEx778ESNECs869jlQs1s+qAtLzgJG/wtJ8DhDPzmc0BRTDox7YWQlqEdoX9iv5/DAHRdyjvAQwht+QiJ9ncX8/lkzCm4C/vNzNrpKEmlrssK7sHuKyyanBgt29uv5XkSxsFheVwTPZX2Sctb5joaBPzEB7yUdXaj5gSVh2QYcy7afAAXpTW5KA01s4SL0k/KRakDPITQH8RFqc1FqcdFaRwclk/ARWkQcIiLpAYFEQhoGJCljNRSjVGGOzAJRE2xSk6KLLq+zPuXYIFCeSvKFrIhjNJQgM+1gc4E3zO2o2F7Jy+zOIrXg/HcyNZOvOal2w2WYpPDoKC6YmU2RPzRarAtJknJ3qsJqD6JcRyOp/nY5FgluJjfAqbTfNKFnTEx71OHU90OljcH2wBobnk3dlfqxifuRbq8F51lvWh2upQcVAAU5cV0Fsw0oKds3m/HjRoNCLYbXcyqGZhSz0ZvkrHYsc22IzRxShPSBoW3ehTJVIOxhFjJtBSleG0txYlZAEWsNZ9AxPy7bS8IbOzNPGe1rPkuURT9gEyBzZ2/7saeCOApDuPSbQbYEcBjrxxz/ZXj7evjIQNFEpYoGqIzF1X+zOZoMZzn6yQnQYAkRnbU2huYvqWrYLBbjkH6YAzWPC9yuYeVWZoTmgfUFMSS6jbQIkzpk6XJaVPLPO6NvspmKemV7ANl+xX7QL9huSVUMKt3d4G0tWX7Rw2QNpACtfm4sJgSsY3ll8BEQr6DFAzTrI9HApQXQCVbbrSsDLRit9FsQLQR9p1nhcbvcqodq8yscJW75acajs/W8FRHVBHbekG9EfiMjaNsvBhlM8h7qtBSR1N6llNvJCtSZTWTBZCQJpgr/pKdfmJ0uAy2rfpoYwY5RvBoT7euk1ciCQG4zEhihqUvxL5kPMbAaZiB01UZOH0QA6dlDJzWZuDUZeB0JQZOlzNwajNw6jNwSgycVjNwWsLAuzYDpy4D79oMnIYZeDfEwGkNBk5LGDj9Qxg4tRk4rc3AX0XHGJV8nBWFFPGi1XE21KloMbolBEwSX348aGMO2fEiGUbi60w0kEWDHK8BCv0nK2TKLRbhJR/Ps5nYYg1v/06+suOJdKYvogTieF1OrsctARacdPsQMBACQI0S5NThLYYmKyIITgbXaaGQPAaJIIJSLtoDGxdgmSVC2z7Ps+GgvbZWDPNaavsbZfo0dABLhq5douEoSxhk7cpmUENXlw7tWvX27B/PO39Rd4K1g7upjQZwKBo6pg+1gjdJzht3YDs6FzNgjpXj+78svVxSBU28QAiKFPl4nM2s6WpoFFuFiGIOgSQvP7eUXV5LjQeE7hIsMRxEG5HMiiam9eRq2ebjeWyNAYGWAyAYKXMaI1X+Genymx1m7RVY57/Ddg3EpJhcF7CJf9ox87ZszALtIzEhfaD/pUJlh65UEcns7APj1wqdM3TMOQPMFDjOWIieNL7afPrs6ctOgLRw4sATFGCPr2dJyGoXLugd/a2p0K4SxDv5CAsVzUhJ7RfPBb60Wg6EHJ6N8nGCew4FwLERUjA9cw6lScfn+cFgmDXssgQmXoNz3Ly/DEDDKahqr+m4dSgFA/SxC1RvcZpP5JbNqeRK8Y4lxVVhIfWJ5eXisyZdSRgkfth1tBhHKuxiE063omMIu6eOeRTDh4Pz6fMvzUXm6Ov6Mp8rsxA0+z4D+a25rrP35JudTes750pee5YN8+xcqB5JXoDSRI0OultrMe91iQmYSPoMDjMEleR9I32oXb4oyKiTTUWcjegIfTNhu8OhaEZkI+gYbTyf+ZJFoYN7asdrfmlZx2++ovya5168FDjzLq5ZthYK2s1yKVDtX1mrZK3GjbvyUpjGW7le0VrtM6/mpVCZF3PNsvVYwHZ2XT60tq/rCuWTehWM82S9oitAXUIQSPQmXXC8k2laL9fWQLJYQp0LgBg/LxHpSj3BLYRR3J+oi9kAIaiq55aQf31wuN/b2zn6B3rqfNOB/2vQ29cH3/9wgq9fJPB/DaldXhbZtDlDiTiZma0J9QOaNJtQ9MygckaYzrgwlX6AYnfgbUmoNUkYCIWLjUppCp4YjuJCr1BRkT8v5A9XbfTTPaRmfbH3SY8NUuMC0KL+XoHpDrYoQsIXAmaX+o1V4Y+sDX908/RPK2KrVnqhnXGgK+HsWlfXpw34ilex4QevonpfXVWV0iDUCwUK+bUUBHzFqvBDV7moqHBBxS/kBR3LjeRvf7u6Vjd3hiG26Gua0z+Sror1O9ahs4CxtkY8KQpodjaHIBjg+6vzb+D/GmsWPwn+7GhNTLKOmr2Ga+CQTvCHhhd7QDYVkB0fSnXNLX3otWrNx7rNR7sQ7HdCe4tqGEimLYtMNuDHjCLgrRnJ26lV1Ik9EJvOzPTUFb/KVkWVNFzlcbCKd+nvnCuTdi8URZ5IiqDEc0ny1CMJBnFchSRPy0jCFCe/ylZFlTRc5XFFlbrEeFZFjG88YuhojtvR22u4VomJ/Vahzjdl1NHahF9hq7RCGq7wuLRCXbo8r6LLC48uMpbjKoR4UUYIo1/6NbbKa6ThGo/La9SlxWanihibXK5+DzpkmXgqJcXmZhktrBt81ai6ALcqAKYPAfi4AmBtUlaJ401fHvO8mSaSKkakXoG8pULa0eMD9baW1EtL6j1e1l5an2ZPbJrRhl0bZUMatSpS01CyNliMRre9QVJclqj661pDTedjONmEsHzoz4iGnB70Vay/4hv96o+EfijZVvZ0jpSS2Qicw2ACFqtvwYOLx+p6s1atzAErR8I49TCGCBjnFCxp5zHAWC/4McNSB4Ey5Bo725UIOOU1y6hZHzA2upxSitzWUuR2H4Zc+ocgh4L0YQjOPw2CiJ9mejWv4BaZvk9GB2QU98Q54VXJ4K14D2ayBT7H9RtzT+PKIcvBCmmZgcZQ4dNtYchL1ZKdB0NBtO7bB27VO3pZuFUt3XXLLDKx0LUeaU3Ly0ihWmCXj707+JYGVBcDCvop8JBJEFp+NGTdO/f6c8kl5/A6EEYIg95QzgeNkoxirVCxclIo6DyegB81wFaA4FDhq+gfWTbFc8EC7Lp4+vjjgRDttxBdJymi94/bT9uP4eDyiA4MBxvK7ZNOC3WaPtgli33scBgpGwzu1y8h3VA+xsNJsSIIqOMJ5l2CttbJRx49/GZw1xLONJPBIBu013ryhHLQw3bMwTUKgGYD/bqAOq/QwetQ4K9uQjQbo3wARxbi6xv85XweJlT1deLVJM96uhiKv/SX6WzyPh/TZHynfpsWF+O8n0+TYT7HLPMvxb/RRvSGv9aF0wSsWAkW3FW/9Vf0d+xDLDhKFoVP0eGE1YfDot5IfLlEEPAYvcFHp8xANoIl9ngr9P2WnDxlgV/gyVAp72fjIpN47M3y99ns6yJ6Ta+jw8UozXjxiwXS9PsfzbticX6e39CVW3kbVdP7bI2NsTyWNonZ5PhuY3ofAqdGdZvyTsmXOJZWMTmE1js9eDZAPjrWFz1C1ls2MnZpNh6BDwMXDqO99Z5R3H4vSGu9kHS1KVFk7M39GppMae5cZbetSM4jOqgTMzI4wXQAcl1R3Vp2SgsJUvAMYd7HUw3hTB6rTSawEBjd0h99SAXZZCijhw74uwgRyU+e+AGh/sYPNp+GnYDJMCoP8xz/GQUmdkoFT+ThJF4rkHb5Fjsf/GskM+5RPoxHKise8yEKpBzjdls8Itza/2bv8VajxY8HQ0eDm7tP97c6+lvZsaAaMvtwEA8GSzWiZzV6rPqKOzdrz8b6q1zbZYQFk0TWSSjgEcGocNVEUOejfxIRKJ2dfduN5TirhF3rovbu/vFJdLj/z5Pop/2j44O3h9Fb8ePoYG//uEb9Vwf7r/d6R/v/948HR/tv9g9Penv7r3Z+fH1y7EvehlryG678bbyVOkDDFsJeDS2nvBpaSHt1mKj2MbAFtvediW3vmyW8/ZrK/6L0o5LXJV+V9C/5PAgjRTLbpycKfK+0tT74hBvCPKJvDOC9zKUFEVjVJXjm/SV2QjMS20byqs06HT2U84wrr02zKoFXMhxOriF2CmuQIu6WtXbK+sUAnrFecGDg6tKUt8BxaWtFLmBrNVu2knnR7yRaBiuAl4q1rFm96CFtYpnd3vRjlWC1cmVFyHjPEFwHe+oNBEqkjJu8+3F7MZ1ms6aK5EPtu1Upmj8d8KqX9mDaQRNwyETZVXhI0R5akYGTttfcpPMcvu+xp5ptgm8d0hPC2DQgPNF8lk+XRkVV9RsNg4x40waeBBIJZrhrkCMYMJsa0Ma9h6gZ/gpIEz3posZ4Ml43kMWTfAiAtnkjSBlpgssLR2lbwvlfRWz7g8ErtRq+cYyiGXdcYi0WkgUDkEgSFqp+OkzGV2Ifc7gBkV0PhSq4M8UVG/Z2xeVkMaS0JEUGgTXRvDefyKCt53IL215j2WENimoMbRqanZvZO4CQ9Olmks/KFwB+JfFQgsKMyR81amoEBrPJlEKG9YY5Xvu9hhAxRtIsi5Sphve8pcI/Qe7kXiDSkB09CIkMQQplCtUC4t40G71GfMYDaGhgpkUOHxzNmwYwzZ9RUlxR+CFd9AyiQgk1jLQ1GNZtEOVN4EDWe2TO5nvKhfweZevkOhaq101egNsFdRhdoXvTSZErtVtggLfJTqHps/Z8gnjJC/ugKmPoJlFqOOmf/icV6k+mQi2E2BvZnJLBNGEwpBc/v55NAFp+y2oQk7lQBFQuH7xrNsa9LAX7MVfQQqMaGElfbl3OZOwzBmuJqMIqFEHtMhtO0f220ZMx2o4m13K3LaEIcuXjC7HWz0agt8G/Qif6PYPBEyximAgJKyaF6PIMe9eQQ98A/ovje9VD2ajNdyBsiUqiB0SjBqUaZwiUUQTZVGAnevYo2lpjYa7kWMpXghWKbDYHczoh0SKsydEHobQksEcyenKsIkeaFuXQnudjIkRwcFcZ0KWjUEK0bTeoF++t7PGMonbJGt07grOtzL14D1Y2dQ9Rp4Rm1898ToduS+chsXnsF+9750MxNBB+gIJfeazcaEC8s0H08vinjROxhaCU4AuIhyKk9zllEBeabJSN+xM41y2iFLdQ0UX+Hgotpm2y2mPkbyGfRfUfhU4uxOxeBn8xKXdEpgeATfnWkyLaOTw+EEvIz2jxK9Y3t55uicfXieCj9U0iqFgjATmxcPx48mr9eVuhLOPjFnOKacXDpxs8IdTbYn6+/ny9yC9AdOMD/OhPoTH4NYTWNmWANyTa1XUyw8ro5OREfuk6U9jsHAcQOQ1cyM0riIKudaNx0mWWGfjfOOlR5HL+QZ6A4P2LcZ8uxcj+LNdIp4O2GveminVGlbviRyv6299YD724yIFBs2OIeTTPSlXX5TWlNvEqR7txAhes+lfboifI2VEqRj29nQt2YTzQnyUFXC4JxHJfTgE9+uoNIVPATX9ssxEkkOzZiRhdpIlp8qvo7VDe7xmI7qo+tO04dBg7ilBpzBocH8l64pWDB8ajO79cOsTnl1UYl41FaBwgtLJ5HXRuQ3sIK7VmXvqM09TETsG9rINn7o0fx6gjCm1wgEVB5qBMwL0hyS2Kt4e3HuaT3iAvBElu8QiWTIStiGQ37C9GCd4j2Lb071IhbjT+vMBgTON+pmBCnFOvQskORAAQg5AX44Rqx+HNYeNhgWiD6In2TvJRJl6Opk5zaj9Dew9BsuktEG8uijfrx8INtgrRc4RuDFegxKZVZRv0y8HOtypYvBgw8IhUV6nUisV6vBl928Vi4p9nHfE/P6as3vcN5m3dwc3nL8BzaKsVPe7EEEQaYg+OMozy0xwktwWKawE4XjUKPOjVEtcKAjFcWpF8iBkpBnpg4EszDoNhUKRtwHAergWjOQYPnjUt1lf7XrlRWWNy6R2oxLP3mcrgUWTTZJbMJ7ONyQwkFt4bnIk5eDHOf8c5iWH9KJaL1UibM2ZjA1U9gY8T4TiDa0ziKwttDMXXK4uvNyoiIRuQhq8mV9m4N03yWSBQ7oisOO1z8D2DeJGzRvP20YfRow+DR+geNJrHbmxPDU7Fu22O2mAXnjY349POGd6fYq/EJmZEG61mbMeBZqDK24AElU2xxZTm7uhmO7o53TqLHRuPlZaU9xBhSY8WCIU5gZWRt3C6/fjMnzcKPbT93Db8Al6+0PPG3aANhs3tzpPBfUMmaxIr2Xfd6Anxpy4S/QWCHW53tgb3DT+6NHKBaX1Uu3U0nBJUq/ktah5j6FOheGmrg9qtCoGxrE1RJPak1xAXeAENDWyPS7NXCJ6myJdUmE/YV1J3oJlJ81DOcrkY89nadlcqe8g2LApu8J6trZXIOFx85YETLMCY2Zb0XKm8mNi2gU1ED27Qg5RSgTXFsp3jom0CZlIB8deYVs7xJZlDG1rQ4nWAcdRstIWSAzp6e34zb8TOLqrODsepUXuTz+qJqYaB4UuMPO5ILNttUidv0PKiIhRDQBBGJyS0tsyIom1D+oKHw0MCDM6bQiSBjQtLuVRS6iKOatORNYJattA14Ys5zFYg+1r5PqhkL1RjP/QvO1zK1qbJg8oQe/y3btTx1Ec9Pp14qZ7HJ5etdgnJ3ywpiMH5bUyEWjrIL3JYoYJqjcYJhs9UWyFrAlhChSKK1jg8aDBAtAVe72BxiSJ+tnpeaBHQ7RqI6mU15tYUAGNTs/BVd6EJFItUKAGnvxa9X9d/bZ+BHGnIsP9aZ4p1n5TdjCBqlOKlXVEoyJ4QHByUet2ogq3hgWuGakeU17/p5MoxuVW1ZuffEl2Eu8ezzMrUQK8BGzd3dymuKDwkoPPF77+3YYXL4XRJbHEnTW+E/XFwllQC9h1D0V9VLfTx35IisisFJy2rLNb45x1K46WLV1PTlFth3uBO+SdYbmmHfN74eTK7IuRQiHx9p+l0/zUKHEyR0I52dOwPSffojn7cN6pXbkpW8ceu34tCRxvEe56mWffzmJx3HCGCo2+++8meaNdAiSRQUa6tNviJRMts8su3icw2r2bGzFxFhXgLyU3zcoY+w3ZNxFzpnXcyAYVMqK2TUGxHYCHfNLkotsEJIls3sfd4dgqofW8wr9nCpd9ELfB/rPYVOitavrzDeMgdkgfgQcNBJxbiP2AXASq26IaorENYFQz6DelN4NhTNl1By5pqiMwb03KnT2DKtQJTrAdB8YfvM/kWTuZJH3QlQ2iVmCXXnh7p1us6KEgdka4gB7W+uPJgjA5S1GNZ/kt14jKLvo06wem5xvKfRDrmBnhWFxlEV3JNArPboOEJd+TvzYkmVdZnlEo/LDkJPQvtyaE8QA7vTW0borXKdqTIhMqnnbM4XELrP78Wj4D9IlB9OraPRGiH2jndft6xEa6Wgza2VfJQ7HtBHsJmWbBU7CuoMJnA5uPIOVc1qHfmnFy3czgpnpmj5DiEuWqUnBqqmi4RFM5sWyLEqUFLHlhsyfCO7+3kS8jminpltPtkdLtcjXCfkmiXD6faJZGt3Kz4afBhZI6aMucwMnpcIurrpY9/qJh1t+Js+10lfB+yKj9sRX64iK+3jj94VD/tgv7RizmdgMG1G8EA6JcKLgHoe7sJEhTcoDZbeAYn/iFnqM0WZUjbUgW2qMCWKrBl+1IaG5DxEGHN4D429GWL6Zd4EqSKUKtNiZ1T3/pYCgJbwV6FmsfelFWFLYokhFPXfJGVY+2wpR3M58mF2KKOr6yDvPdmhV1k6pBFelDqk0Y0Le//c+flScMzPnTsUq8PXu4fHu/3Xr49fCV+Byps2hWOT46CxbbsYocHL/9xuPNm3y/42C746sd///df/FJPTCncNh3/vPMOU+KJf9/t76mfvZ8O9n9u+CaWp5yrXmjaZnCRDeRPTcIqd7sggU+Oftzv6at/tTpaVXzrYQNjEkNTJ6mPsM/Ox2AOwHzZ6BTYG2bvwZ/QiuAnJnreU6H2ev1UqsXcAxHCp6hwQipeo9iRQwI5sQi3IpVYSMynsXS5ZbEKZdYVjgLR0cx1nMIYU1aKC8izMhnLNzQ95QOe8NNPmj70YFSRQTae5yjnoYx+nF46Bakd9Vo1op+xHf0km/KAkKc8lRlPBHuB92gP3P8AQv+qR7qBU0s5xVM9coJX7UzxkgkRS7rcCJKmi3w46Fl9oxXQGUrLr5oWboMWnD32swLyjYILXh7JILfv7di2dj9wJXx/vyb9Lm96qdBkrnqQTkwAedrpdFRymkGObhF42NeKerh8YGF40lk98QWmBbdr9DBEp6IM0ZV4uFlCN057iwxdhyg21l37kSWicjDCkMOCPL+RAip/wsVQpxzYytWdszC1YVH4rbKItBPJu1U53RsT/7SZp8SgWADh9o5/bCqnwjVtRoRVXFkb7+7t9yDvAq+1lux+MFIy8BGns2DEftnHIUxzDpsEB8b9UYTFZO42GY0tewyHDANQHcTfLSfyssdoSLLuZmyZpuUVJm2xdDYkaOw8WoyBuNIhiG5BDbNBg5s/kFOhoR7mTFMD5zImCR6FsZu19JPJtpXk2yeTca6cW016mUys1pxcuo6Es6zCiATdv+T8aC/GQvfm3OM5Koi3qFCwMv4GUrsLy0kQmBp4uyOHgHKNgG8BXTSQsYPBLyhoSAFcpFSwJ3GwNAghT4MRFEFMGuZDI46j73xdx+pTwDFBNRH8gFb4mm13u8saL20CBKWAetpAKjTO1BmLJsypINlZsLoPNA5ZtqyBCZu3Ai1S5k6NVnU16KJVCwmzpBIJTFaNKDvTuZ4bcTUAQ3EPCBuelqe1LgFrxL0H1nzC4AP7pIkvB8cWiCBM9t3puBg/W1lFhuFLy3cRnZU1x9Ffoq1OB5ixA2yN3jaspDPpVezxLtTcsApGfwMnIievfDIt8IyZrdHROlu+7bPHaYaFAbKqCidC8ifg7NiN5TWUJIdc5YAU7+N6NI4FJIIKLI0/SqDY9GrKfrZgkWUwWwSkZRqV2iN6lxVmpddBTe1J0r7KboumvQbnpcqOTXq45ZqPmUMhJibE0zmQ6OAy18wNDxBGEGpd2o2asjxeD9G54dUxvN5My/Zhk5epmzvobgi3+cX0sDPc03KPbfXGE32YVmC2656mCj03T+W/3FYp8ZRXR2PliSWKgNfWGdi+tLPdxTYaoC9iowRdzKZoy+cNGro56D3qyr0gyVC4+A+3LwHkbBqzrN5o5YFe3mAjN4ToFJC7ATcRqnvm1ugRGqbD1gDKQq4zDXcl9BeVdW+JKFtjaYWN/diwIUuJJYSrq687LCwLd9qdQOEb+1UcUE/QrtYjD1ZBYUnK6JFDROsQAkuLEWwptcQougxcSMkNMnc7GQxAJ7GFMEUwg2tLYIQGiete8JGlTik9Ru974CuME2vzmV8eS/beApJYXvWnDPLRRIa6bewcvvzh7RH6MqpKIKE35TG6WZ98UMf5KB8mM9B6j6WaIDQO8keUeroUc5DBeq16eTdqnBDy7U4JPmGmMFbQboDxlG7IbHEBGKRZg1EVb0utlSswjIwnqFBwvdRp9fujtz++awTq78Hc8qsH1dpKRYGgaQY8SSiwsIdkVSVp0LYwoXWfdwVYTMgezezElCF0XhqFxIJpFBUGN6ivaDCgg4TQc/QTA6+kazZ4d71RqxUeYNjrVdFTJy57yTyhiC9u9bK7pZqJ+dKGfKyEQGe7YwkBb5kGCbaKgQiVLm0F8QTTmdOKumgIuHhtK8zoINpFrLIfoRrb7EjBtH+KIcs4C566BrvT/MzoOz6Wa+YmAQflm3lCNy2rUZEn63DDaIaZk3015gOqMR5W1o1kTWaA0kQu6coFHI9OJXy4OqDueOUXY5CoMSeqfUu5B2ZHi01bVm8sK7QKkdMcnEOEQfE3Xd0Y/QeYnhPbNpNoW0kirSSJto8k3NYM7/PB9JK/RUD4SgKh3wCGfhEguxItAwmQUlIysS3IiW8+TpT1JXFNLzJaRoXlGBG3KQ9G/iBxUps4qSZOKomTauKkFnFSSZzUJ07KiJNq4qSMOKlLnJQRJ7WJk/rESRVxUo846VLipD5xdhvanq4GaLkhXZVURnSrflq7fmoZ4SttxJ/FDo+zuBniP822jOx6tFYxyuuDhApTPLBsqozx8qHEHJ8YKWyNFq5SqfcxZSZ4XOrBBw26p36nEJwLvCPghF9idHffwv/kjnFVA7fpwL+YcdsZbWwUkQ3Zlh8mOMuF54MFqCVE60jIUG8eJunKpd2DJZ4l9eqItD/c2G5PB31F0Te5EutY5tMWN4u2wsbOVsCpq5Y5s7XMQFkBudr6iARyFHqxfCRqf4BCg3YESaCYOktoUiUhB+7u41CPGt59A6rBovR8zDmAQUbe31n7lIZ+Bj1kzcfPYTt6XHKHAsl6midnOt6f+79Gnja2kcsk4G2H3YjVtm22U+y2XZcBsSlGge3aLBkGxdhze3Wm1SA4w26vwsvwv3uPT1POzimxc+qzc2qxc7oyO6efkp3Tz8rOaTU7p6uzc3qap1XsnAA7J1/Y+SPY+Y84Knr8tN2JHkXN0JHRM/Et/h96aITqZIb6nfopZPNkMYcEUPKcBQ9hWuof/9RFH7mMFvMFazHLxj16hYc1LXaag7so3NeeT2C5HPNlOZ9nI/uEKkWvXyh6CguImb34cokYhE0GFcMtBf48hYl7BsyTJ9vuYQiMh9T5XXHIQlaz7vnig31EQ78OWWkdxjBylWtlXDNTNNDiRb1QkoWe64oUXnqZpsbLriIo3HrLBYPU2zh1zGmWzWLOydUNxtq4Od08i/nVXiItbZ/Qm1yRE8koGB963dJ3AwwlWpHpKf1m2MfcQjgW22PB5mxShM7/6PCPHZ5eo0YKJh554JN45z1Y6LQho8TJA5mG+7niMMgt6p4DbYZhecc/4WL6fAKJ6JaxjzAMYb1y3uHEqLoYM8BK9zVwAsbg4OsQBso7gnBBcQBlwN06qx1mSebiB08j0sPYGZN7puSNu3NOYvgxVNA5CXEY1ma8lBgvlYyXBhkvdRhvt+F+rst4qc94W2FYhvECB4tWyQreS2vyXlqP99JPx3upx3vpUt5L/wTeS+vyXroa7zHVop1BvOBB8xRZuEXQzlg0g08ip2Gt5quG7fiCuodWgmAkhbLBXqVLHGB0VTrPZ8YEDYG+MF3icy1IKgSF1PDcpVO3zz1KHFMpOJeIRVTpOIYwoDrBwup+Sc94kEprNV6HoFStiC3KzGr8SYfX6fJ/9xG20jwFjp7NDKs8dWY6vjyoVWqIe1QrMz+B38sKZyj2ObMm470BmdYGmVaBTO/NSapAczGWH5RqVXKkanjiA+vhshNWlYLMaSVd1kqqWknrtKLOce2Rbvl9bPkIyVPd4nKicrjJG654C2Wez4dZt3GCUUDU1UM64wFWsE5xnShDy6KPk4NZUYgVCJJEAgKg/TcbppnG4YTuQELkTorI2Xb9OrDrLFbflDL8nEymeCqMPoQx/97GPjXxr/3hIpuMMtGNZmNzs9O5efqswxqTUct77/MCok70JrP8Ah25pAdA041nbm0pgQftUOgMJo+07kKpE2+9so0aJNah3cfRXAdKrUVuwcpQkCkXijziC+yvJRjRo1OK5U5cG7N7yBeZunMEoUuedlp4M1ZxGPDOO9oz9t7tfL/fOz749314+bTTifnwgfFCWiDoDg227ZaA5grZ1CYaMS7b/SwfNhmADYMUa6G/mEF6UIRALHYwnpsEUpuxbWWgfBahLKY2T4npjSxw3ujf5fcN45ZCAQzwtFEQLmZDrWLTYlw2qCuzGspbzRTOm8VdR8KbI/fJ4NbKNiU4P7a+8sSn6QTTumU3EGvY0fuAgySkk1lGMgPqq6TyRVf2r4XSpYu31yGeNKPA+yKVII6F9IJYhDMJQ8wvQe5u4302mwtll6cxgpbat/CXRfNYCuhS/Pgd8u54oG5sUPhOJQ7MmrcFAtQ1ihRclIWOYr+/pPcOmItZjk5g3Y6iCvwqRI+ubruNcZFd28QIVdjkFRpWl03xzRB8CzoOLZQHLjfdE8Wvs/zics5Z2JQloEuKI6uBn/aAYieNo9/zaVOPvsWwbhhhoJJkjCbCwNxVA4oK45UkdKigzLH7BDssBFP/0mHQcyH60WW3hNnpe3V64mdMgph8W1RT5dn69ltcGgOZxXoXE4iMDJus9c24LINsJfBDCC/13XfVwEOwVf7YZ5Q/ljdE6Wp5OzpjrRJepfA22c2r3xajqSTv/hgWTgWShuaZXZJDnAELNZZh1W18P4HFCProJOAlACz98TPWRxmlcgziEOnjMB0K2XmG8e1omgqtt3+ZDweiUjNwvoOFBtkwE1omVHSD5+KSwFcIyhBkm/1lDoomFl+PNuEwQa809oHCeCC9/FWuCV2uxZY6G37/cjGm3Cm48pFujvW3BTw/rm+vRZFysB5Y0We49YjLIxE1GjwG+ey0fxabGLX0WLLoWHSUOTYw0pXADLLI4lWKLrTj2NAlN4JkbZ433gHl7oAW9xD65I6t6fdCYT5G/eYiwlX8TpFu8379TjTDa0ABHjcY+cXMJgzj7dBhnF2rpV/pDmJ4GAKtwPhDZHCE5YwUL4jOnhK4XczmYBtZmE89ha/Qc8UCx1TxyrhSshegZDVxWppkVpsmLkK8YiwmCXWzuqNlpMPOx7W6jyincHWo8e0RaqMgGy2CsJ21D2NJkgLcB4k/M5ltoJnpTZBUmiN0vWpQjkQZmAWOi8VKY0x5YpCEgCYvxIISbWJK7HSSzAbkS7ussspKqWtBRqR0MrlaXldWpT07pBqS6RzpJkuEPlCgU9FucDGFkLz9SygRyoPbpiuVzY9OhvupUtOGkup+VPpcjjZL+Wj1hqc7tD5gekQr9+3yhLiCeSoyGYbp/QlSYH5MdsjqvJPlKTQ/XZpMt5c8OadHASs7ZUlKS/d1eYZQPzUmDCGuoldC+Ly3cxjriWMdTbv2kd7V0gSQKu6xm9mxd3WG4hRTP/be14yi3Nt5/frtz/t7vd2Do5Mfeif7/zwR/EdB+4RuOYVo4LPG/zndWf/3ZP33zvqLjV/Xfy3OHv1fOtqrCXJk6NsERc0O9QQB+bB8XkCoN7GLqCxOyfOohinYy8dUN13MwSRVYAq9pLiyQyiCwwM+l6ZfEfrKMZZoUngyzBqyFixwiqicCb0M9t2yHbBHoJmOgi/iQ2yHHTJGWArQ0uQWWnasQtOfrLrdTrtjPEAL/Vn2azBXARdZzI5wTCgsWiNoUeOHg+9/aFDEWoUSgKQEKRRiRDx28B7rixftp/LG3pv9vYMf3zTc5pYFPVL1MOdpsAmZ4aAhuNLKN6l8BHSqSckJkBUcgjv3L8UkUElMIKtit5EMp5eJKIBZjGrm5VHu3avn2JE1l0w4VCd46CkZ1CUZmNvIqC9fooOI7Ql+KZDA7rgKuXfiMAGl3ySglACALDhMnDI2JKyIDWGZZmxnUdEQgE8wEZJT/6tohzLqAvqvYJWi3J6QSU3+diuIDdQc7ik3yY8ZU731MXn1QH0D1yFAVmiFl7fTywzSWSfTiSDiRDw5d5G/Ap+tfAJGBrD14w65aANe0WR+mQFtxfLSxzYhu6jkpHY1HSI9KI31r9ut/+//+X8bJaTBtFkh2nCY48XIhtne+PqvrWZcBpRW2hWBbqyXgcO4/AWmMi+FJ2P3l0BAhSHMAbuAa/QGCsjDwSLC85i8H2G1AgcGf2K08WIjSVNQpBNKg2mDo8RlM/wX4lsmYnYm+RD4hAZaiNP2RTv638m4BX/arehYtEl/a42q6awkHnSw3XDt6JMru6di0iqfq/6ltXrI0xvI2deGGzYYL0EUj/UawcUWX8GqZdfDVzWMmyxXKzBrCVkGq4aYk6IgXJmjcnJNhfU/uYZtnw6JN86aykqtWpNl/4plrXys77dxsSwT0tRJdPDthClCh5UyGa/Mi7tMtpMMXb4wxDyNhqhUlptNfEL+Bj8SwQ6YLGeWTWcw2ExGw7CuWelvDgiD6I6hfx81UXHNx0UkMTNyCJJGRHeqvftYL3I06zFDD93UwlwL8Ic+yZ6rBcV8CKWhliEpihhCPzxH11c4QzLM7xGCYgOav/7SeHsrrQbF6fYTdv15NNLvn2w/Yx8GA/3h2fbzM34UI/YYYKmi/GNtfG5a52WisW+jNweHvZ92Xh8ojfWX/Z0j6Iv4+B0BodxDoSU62B9SrCgn2u2tYJFRS2DpRUU3n5Ys82Vt8SHFS1/8GhYbut5mK2JPW+7NOjRnb1upR3dQyEZjyIGl4ReU5HTx+++3zBQwnFzkfZki9WgxVPel10n9i0DlNCCi9e+0d4cshgVQcm/AeKFsHuTn5xnYdCLMwsUqRaAPEjxeH4s9YrDAJiagKUD9DHLvfVPWPK8O1R5JIHb1zSXVzxdDsHqIKTEQz3bdx5sAuLAgEMlwSyu0mZdi5yr+2ZUbVRgjsXLlxVyAuAUBLXgc7pnCznAoXuAB/SAS6kdmp42VcsgaY4+Z7JsDDYWTZb3QX3Ekw5/UBQMMfuF80074jV09/i8vs/5VtCeWy3Sod+DkeU8ictOTPL3NkOxJt/yCWyVCCsXyppJP6dYK1HByOFnkCH1T9KhHjTdC+wat4+1Y/H8WHYvdnU+SWzBliv8GQJoSAZ7KY7JbiJMs/htsVZTd0oRBtwnIAG/aaDEY9/8yhDoU4/ZyMoI0j4JvfBqBLN8EJet2izRBfBjRwwAfBlt/8iQ4BlGhuxQYZyFFehcJHHUladEU/VkX3TE5uoJ9wgdR9N+g68y+LCN5vnhMIWk0bCGK5Ib7xZb/aVPe5ngh+vFJiFU18uR1VkKtc0YuqfjviJ7uJbd/ly9/AaG7h0K2iHZvBQ/Ljtzjp+Zx7JP4q+hYbPDEhJMK03o/KbJtJsRx9TGLUXEN2rjcyH0l1uZkNB2KCp3nG52tjc0XTx9H74tI/BQv4KldwYwDPnAjPlaiETbuIxj3wWb82en/4mk5s1rkB3L+PTQU0Y6gjqDaMZEpat6pvtxTwQ0oJIQ9JncLjIel+Q3mm0ZbaxqJxFS9+RYvoeUUK5HcMlI2AeQ6VIshR2SxzIBpapN+tVSyQCMDbGIrRgXDwDJYqAKff0CfrTCgOIY4SGpoaazMfMJhFC+DM8kliyWPNBuPiCybIbKoAp+fLE9WJQuTMsTvHlnwdT3CSJ7iGSD4e9QNtz8/ETqrT3ZceXnXJc7lXOFhX6EhlKEd1gyCWsGeUbEPiuhkMhGwZhKre7U9qh8nvG54cByCUAASuYlyY17iXT8GgJLxafDhGAky1LDCA+74WV7Ipn4cuisImrGpGVCLHUuUKOWftntG4HD8zGaDNs50hwMvwMV2ME2IKGgBMXvfcj3Vu56oFFbpaX1ngNz7uJej+csbSJjdENUxnfDdaCQTPPs4rwB0zwY6GJQCJbOYugahHXgwzo5mHRkuVEbbcV5L3igCYUKZeR8K6HsWBcWZc/yFwXqDb6Nvo60lBwBW8e8c9ndyZdpoP/IcNlzYvO9WaehpYvupUi98T6vUFEsoo1FZaexNK/oN7Ebi82ly1qIf6VnA4QTvwQ4GTRW/HwLTfisqo6Lc/E1UjUOWWrrI3eJdazmEWUFCUUydhwVOqieu8NKnFFjyTVpLhCUfI7+SP0V49ZL/IuIrjOhHC7ASsPVF2BKWSD+GJdI/hyXS/yoskX4elkj/JVY1yRNs0dIcSxHC4+ivEXudqtfOwofSLAQFvBBFFcSfr4VQJw3WSUvrSDsrNSeNiwSo/lIK99LcxRRg6e/pH7vY0s1EwizgzZzqr+l25VpJkSo+cmHUl58ekr6IJZ0aCUalkCYqU6x8Q5eHZIorK/URG2BZlqJFkJtcrIbb+ob+cIwR1SUl907XeDK+zIbTbGZf6hIwkjl44ghkGvJLI8B0GAWegXa32CaIH0vH7Q2W9oKj7NzqchZ0apDNsz6ctVBqPDHjoGdx+9fxr2M/ZWvjzUSUgayIw9uonyzAoJYMR/AygfMQCfhabNQi8m0fbFzmAzGW5E0BHgMR0UMXpruz7WBzL4XGcxW9HKL5D+FFb5LZVRE1d/66G0OwumwcvZ4kg4iuEyYXBCoE7LyxJ/u6Hd3JsbyHoi9lvN3ozr0pd98oCbFdTfJPQu6PQzfmiZSsO2w8LgmkSwC26qW3vWSYJ0Xzb/iPdaUGz1zka1ASjceeqNlMYqO3q0KQofXevsmDgZ4DoY311OMw+zEF3DTtlibq7QdO39dMjEYBDb3r/UlNsoICNgZKSUdYEx8GIzoGCpIcoKWOgj0GCkkvUAONfFtDRaXXKwWDB2dQf4xUGRWlxjjDBnvCnWUrwfZFiQ23OLyM3JehQhKddHYRREP75FaiwEsBJAWU9kUQifKW/hnRPwPZVK9/OZmYDZSUtCBcJRKS67/fP9w/2nnd293Z6528/cf+4bEVWK0hpDUO+ngxhCt9jcX4ajy5xuHA5MDYeTrHg5/r+Af/julaMf3Uv5Ixs5s1xhvyLf5dl4WF6Jzi+XpKDCfesBdrJlKZj3zv8O3RG+iBzpF9+n/Iz/YM6SwEZBvc6WagS8E8RM3Th0Pw3xzs7b3e1y69dhOk5RnRQc5APWmKgFHE63W2x6F8Kx3p5EOFJJBzmG72YumzEncjndDZ88BsNMDZ1gg+Lg7Ij8nBXUsKSyYEixp5wYRCsKQSF5ZkCJY0UiPmqKKxsBudOglkedIY3aczLvF0vji8Ba6uZfTo6zIYLMOc3bhMfsurGkKdGWLUQFsR7YyTxq8HOW95NUPDM04nmYEO6uogtf3JINRLQxqrs7UgMMKw3taqqulidbhWVUYc2WlLLaZL3nfn9x/uRuK/ofivkPe9QfK1IiEhC3V19kHpBOUVcTdxYDXWFo5n+qREy2ZmPICnEnnAbOYWX2j3IqtNITI0RJPtnaVduNUt3/qt0mGfLjEqKTHQJQYfhznssu83xKBt3MEGG7txiyM2UMMlOnRrdwYyIpyOAu8G9jvda54w3sYpLCNRroMzqb52ouV6C9T7ybUUAyb5+nYgWIS+BKB/q0Qf1VKf3ZkAK6aNoboxAbk86LaE3hLj/VSzYJgPaJwoWz+YcyqrgaYbqFZggVmG3ti4ruorLGppFVqYQAj96bCwXGiZS+JVNjaw3M+gJNs0dWJmS0f4rkKrLbaB42b5Ih2KqoH+mWPwVP4QNZu6f13Uwf4aNf9TthJjAcSYmvF1hEB5C7ES7cSptlaBn4Peh2gJPh+iWggw5sYlV9/8AXsA+tgq995lPO7dfV1ytcO7kFTv7qu+6cFcr/V9D49pIfY6lAGiBeeV3bQoq5icKY0+cyfXcQ2WpTDHCqSY4+XsGcTNymJFPVGthjVVkCu8vZJBr0MIhgGL1DkcwkUxUOAxhRP61p/PbNUWvwYC7WA5vWKsFmVHw1wmG637ZISZEo7UviMfFbcrfFm+PjmKxA32FpxpuuiOZjbKdL+qEdjeh3vAOQbC5E4MdGwZOMjb/oPXjSnjhEZgHVJWaez5af/MuUdgT+kuf9iukAa+JODP3EKnKqsYaP3JuJ/MmxxFAewmL7q+j9d/msrtZHzbVMWkbipkG5mIuu6i7G1YRFFQ3/yS9i4EC5LtwC9qLBKsMN/n+1VcywOrqPbxfiVlHzCujn+gZka4mRvLgT6ZjjgVBqHeDEqLo4ehV57RCbXSFXVihbv0JPaJqwDElV22oLhRV/0u1y2uulxa3lZJA5g9UAX0UP54OLIvKwJixiq6mGHIwZLgsXw5g3M3cZUYvK4xd5mbMvjh1n6JzNod2S+RJbsDp2Q+xi51Q/df7GLJRcZK7Xy/zz4nN/R555/uZz73EXt5jwokPuv/acMvIkPNXoOvrk19u6ZbIFBP8VF5RQpU4NckvimvBxEMqBZEzUkzWA42n3tV+EdWPkkn77Pe86fB8upjQ+WUSgZG8vsX/cReRy8AZ3gRWlu6QYaYJ1tK4rJN94nY7Vt9fdgs0OEmjQJw5ugDKJbN42qtqmUr2KZavbBFbX+H9vTDaq2ZtS/YnlkCzyijkbbRUzIP/VjeKru8y0jrrKJhAjuLKZHZte0jsd2XNdF5qB1G9IB8PspQN4aXlnXTOPZhhCS9VmjriGgDyZP0D4TjS/ragBwTCyPUx6OlF+swtxgWqdSNODOwK9fl9Aw3N9DNVSpaVnP8jngF3cMNmilQrR2WN7gm9x58NwfNBTd4+lTU5CaVd6W7LEWQU1W//8C19Q9GH3cL8IH94CotH1ztI1xbLasf/PXyQ2Ap/OAucgyqtZB9sNYpt20pOD8EdgUfbIWf1TTr2Act653PGq4rJp1yNK0+eNPjg8O/Hxz+WmNH4bi9R4ePHviBSOcnY+CUb6RPiOekj9t4dcqljDB4qT74geJZc5D87GwESXykgqwuq5dgQkHpwevKoIQfgJlrd+cTmmnPG+oi4526mk5RYrH++WQxHrAwFJ7JjOxZKiWz7NtZwHxbbSMrQ6ZRy1gH1q1SU12gg3j2y4NrrGS/U2EAMKQh2NXCJjxZrNSOVwOv+sY9bCxo4fsIa14Aw8ONnWgDb3Tu6DPtuBHq+Wew+zm9uas63men+vL8/n4VDhnrqOJX2Xg7ulO9uo8bNSkVrFLO5iaGPUtCbgcBYjGfWYDPbSt3BlV2DIOgJIKZzlnwZKKWMzeAkQGjzHENiJaDxrQiUsYK4AKikajQiL3GzBpav52wJNVbIDGWJoZQwzjvG/MLNKzXqVXaDUQbsdtSEi4/Y9s1e19SQgNYJz+WBGqD1IJ083UIsGqjof7zllj39TbO7X24+2x91whJI/GcinlqzsdSizTnyYz7GZG6HDVHo7iScKXoVirmD6Qwu6Zo03ikCMyV/LiKwBDpZDl5eakHEncQIi4oa83BoA5pPURr2nofQNg9dDxjZB0oslp7i0q6gtpZg7BWsQdS9jZEWdwKNG/F/+oQ18e2cr/1UdSFy6Y2eW9rklfja6O6oiFjdZQpdAPHuNTU4SP98NFu/IxhCxjVqqGvKAMt8CRLquE/RAg49bUwOO+CU6ZU+7t5K2j1tqzdlpWbW7er56EKy+dNxWb16OB8qqZwsAgnUo0xUEG+2ChA/ISNkuHme/UVRnoXqkRiew9dLCCwcwlstfFfRcWDKtHzpxbs8IjQNv9jxZ0xwELKT+UmXS7hVm81JAh0Q0wKMEOxY2QtYUlu2vhoZcUx/woMIJrUxhvbJ7ucLg9FJkSdUMtcLXGN2vXIpYxJH0sqddYsWt01juU5Vz3dHZRm77+rmH12FMbTGW6tMMsB3z/BVhPv7kgfzylGS4eNfh92ZXehYN9Iizl/JTOTsdAEQW89NDJg1coicvNoDJjGs7jlvGNuoa57iP3SuIkarznjLtpid+yYD2jgtXQkdfHwXts+ug4e3lvbt7YV8H5ssWsttqnX+FSWMgf7wqNHGG8QPSYcJc4K6rCU7rx95E03Ob4fd3OOcg68odM/jLWbF3hdj+5qJUWUqD5ERX4Bsf7B08mJCwhhPOCwn0ewxr3edDaBjKS0J99UPseb5G+8qXyNxQ8ssKUKbFGBLVVgS3b5fLMPIVhDXu3UgsztuFVdTAaiodcqfBYFlIaetWcQSbYJjeHVP8uVXbbTwkZKPm8JSQAiQRbGiNj0ngzDGDuM+6ArHHpIWBmXHV82NcU0ZaQ3k0lMhJOhAoSktaSorE7OjSSuNi3/L/C8l+Miv2+VfDfB5QQIULAKHuqKJmMpeQsY/i0v4Zoq3+Sj87eo037SgeQnTYti8H6L3jMiwNvHT/GthQS87zwNnd05uFrRPZeh9XRFtEqC5zC4EKbHPLELRLwRUYg/slKmSVHGPPCoO6y7EJyNPbJSrOOiEHtygu5ARLUexX+Hr3ya+BM+PM9tiWZWrKqZDqbypVMcz+0sD0MsgsgghMCHrdiS0nJCaWGntAsp2vR9CRCalsuwfC+t/+dC9TjGN907WfX0a075r8+2253z+79gdPsmA4odkFXwMp1dLeYBGGLLP9OjZ+MVCm487sEAshQETwdcNbFbmYNno4mWTYY2406FdUuuIawUZ09TjAMGiyGrYHhVFb9jNLxvRSfAg6wC40mosSVqSOO5Hry8gKUX9mk0jDIUybh/OVFBJkgNMGGy0WDDvnhngcaRWL4g3ZpVoTuPOh5Wi6qY+N0ciwy2KsPkRvyLS+YF3BDRzBWS3f5i0JtctaLy3qxMBT+VBr3BfMd0kmVVgxXu7j5WV1gxehde9LQzS2BnfIdjNiV1sa2KYlts2aPAw9KvmMFXbkJb9uOmvmJgPlr3/ogmCqB3AU4OgCM6rI9baCDQ41KuavApbYn/FqxTai8hUyJE89vpZB1u3hMsinwfXcMFeDNrc3CHNdOZ1DTZJVyfxYhKHQ8Jp2NYYm/Lh5TCt2k23nbUdtEfpNmzp1KtwR1n43jnzX6EfobRm7eH4u/O4V60t/NLg7LUIDMBVg2Mwr138OrV/tEx+1jazHPdDF/E7ULfbIlCfJIydmE+a4w4fIy+63Jw9krgTF5wmiQ9a345y4rLyXAQnItpUqhUb8+3IFELcYGuFMe1BRCACkxJXwKVT0VtEfDqlI61bBw68M3zUAei9Ui81/A/4fizpp93Spp+xpquLZDDTWyWNPFENWGNA/EGZlrnQdlVpEF5gxI2OeLvVsu55NuKHnQzs+VfB22VbK/trbW7pba30sEtNI+OaMdNZDve4E7X2XiyCZIN8nlvDrYZfZFUTKQ1NVegPIbALpJzaoCuazStSI1I1AYPErBVo86WU8f11fbD6xM6clZ7IfWtp4p4We60LuHJYC4AduVZK9GhrnK+0uRpLS3HSOKqGKFWLP6sasYtuGUH1FCKTagNxvZVLdjFGHxtWZBb1VAbfD5VtuIWlO3IEMNq50NNfPxmSLEQg2xOuGryDt8/eMoM3+E72yO+uedXluQyI8PHIk46tMOdh9nXVOHrVvS1UU5wSmWDr3mcGjAdODudEjS7ZSYI1xUJMX3UjRp/l6BVw2ILO+E+XEBc5bBbby3kaLP6205iU/igVD1j6vEVFr80ZDF98QJWH1iHXjA1wcdN6YrYQgzGhq12J14WYllt51m7TqhguCMNoXdlPLySOML0w/lq8soBACd5nFPWJLiDsphHLlBCLDfzXmmj9zwSBGZnMmaAHuQfTOYQOatpy82yxQeeekm3RFZOL305Id6ldcqTuLD2GYitvmBoo+vv8soQxmxG0kvvUGhbvTdv9/bL+uEu/Es7EqjAe4KiV1IdjI2sB59Smq86XLY+s7STXnHeRamMNzWHsZ2DfJJUqFq25UYPlofPYRpf40kgrK3lqWXDk+mSvoqOh5DpnG8bWfoZ8HqHvHZqbYHjgmgwoRRm77PZ+jQboxCWwPrDCaTrFj1av3UyCOHmFFeJIh/lw2QGDuOwRR2KiTy4jcgk0bbwbxpb6XOyrNaVf1Bl86maZGMjU5dvuKjot6ZaxWCCm6bcKULPuiXnAwHbIrdLkrZvwfloW+cIE3qKhe/w4OU/QBqQJREGzkEa4zE5GCjLIeRc+qVhby5XMoHiUmrZNixzs22ujNcc9eKOft///aFahW9ZL8lBohY6JFso+Ly72jjLm5OslJWzlrayfK5e/dKsroR3KK+r02b5YnlfvkulaNM0EzCQJ980JvbuNNG7x0TuGxM99onakcK7fACymUHCR1mZfkN1+kUATAUv3HWiNpqJu8u02oBo2BzbVGObSmxTjW2qsU0ltqmNbcqwTTW2KcM25dimVnBuwjb1sE0/764YCJ/U3xCnONr/3ffChiil22CHECvsgCuBG+bj0GvteyvhKlbmUFfY7VbC1tNDQ/+yx/2yx/2yx/0fscedXlrCoWJ3O+Wrx7/GvrYW8nxJt7BfeS+7ioR+yA62VneMVmJ15r/jrvXL/vDL/vDL/vBz7g/lDjG7wR2LiuYGd2gwNhpIRAjAP84KcCo1DiaDc5VRHlwGBuftbDSd33ozRN4vNckQCjv09hvcZHwPOSHgvBp/9N7OBhlcGpJfjyZWPtjGsTYr9Y7lbGfv9Cus/C6bCfrMkwsLwp5m2ZMEL/+aF0f67Fy/M6/ULxf/Exh8qAKDqh5emvGWTzASPojX4G9s2oBH89QTWEAvp5lOekcZP3hYPx2Mz00NYIajD5tR6TDQj9nrYBxALG3phFRfNMFGcnkKlL65hW0rQDg9L9So44+JGnT8NptQWHRjQlSCnb5PQ+NK01EN25yGVj8brwjzC2HN5YDhlKSHUoWuZeaUgQIe4+ZJAClwwIwyV4dSFJqWhYvsQTaVXiN+UOUG7X5o2FTuoUbvGH3AkbGOJtc/Huw1luHG4y32LTcI+OKIDyu2YvO9RN2KycCDp8ZLAnaWB1zQwRYKFSbVywheuEEVPPCFFzXh1+IR0CYSfwoN0AqXoOKCUFVJ2MO3vb39k52D1zIogaIKWxkn15Yola5D4nXLRNBUDqRSQzEpk9T1Rj/UZijcRXgw4CY/rJZ0MaYRSphHDVv5gmqGvFBUAQCq83xLS2uLpgalAGrCAqiXFHpAtfMcYpvgcqKK8HCCsPbg25IVhz5qbyS+wOhYJ1TfuytK70+tOmcgZpt5tLERbYHiuekGWcD8KoSmipt+bi9kdZvldVSzf6ndqp1ouQSkmHrzSQ/TseX9ZrCU0Dggt03RFYIOhGwjVsG0X4j/6XDa+Xi+NNdyCRIABwtNFnNKoGYvZePsmlikRZnLiJPtRHKiI4tx/h9iNxQcNLF3QnHY3WR8fjGDxMRyFOCOWcl4d3mzZ27oS4Ii/rYLPcuKZnrb9eh4lY8H3YbAAzJ6xG2YiPMehv5qDmaTKbvKIeH6nKdJYRdzKYrrPjGGTIopiiHfxMEGUKNCFmvsHL784e1Rg5Jeiq7L9PANo4KGWA/As2jGeiDVSgHf1+xIxjJIri4qdq8XY7AjUlQ0SmEJJJVaLGBkGIQiK9HQ4f3AznZHD4065ZhlorCWN/1hlox7YhUkXbspn8WWGeDAv3gIAnKpi7sO4pSG9KLfRyjreB9qthjKzd86hsMmBoku4Z7UXNAbQmCAJmzbhSIU7yO4MItKvAAUKEQepRt0OwuqRPMJ5r4aywavsmyKebYoZRZEVfHbaRYYQgcVULeevaWINiLxQ6xS+72Xbw9fid8nJQBkAjm8KzCeuI0Kvr6NzhfD4a2phxVPLvGwdwjpp0b5TTZQgNCN2YHyyMVOdF6sVwOkFBg+sIJyZrYoc30Jdgro4GR+mc0izHdFNNIYtQU2Gbux9tPB/s+IeREtBEjg58HfozlgTAOdFZfJNAM/6SnmssOWsM0NBAgsOV0IRrd4hVjLX74Uy1kLWPlK5xSU7OlDVWyrS4cWRsygucrqp9Q62ZeWLKinyTmLZQb6DN6M1et35bYQcVHKpx9xjD5LhbRmQDIWuEh5eoMQlEk0dGNnPP0D5s3QDVmflN/9DOwKY+ybnf4IhKCzGi1bgz71yiPIC8uOo/ZYWrtZe7VRwsQ0hMpoeeHbVBOzcgClu0gML2eG9CFH1cbUpuTUUzsuYwhgAC1FXg9Rdk37q+gtTEopPqCTUv5VCd2REEtagoA4TcXuTEuONqema8fXrGewXEZnT9ezSWJX97R1vSz3aJXv4SYY9LOAtuav/KuqavW0+iVIae0tfOpTUbNSS7EoyrSUSoLKxsBY0JtBorYzYnNbzynj5vYomTbvlAa0HXXAjvT2x6OX+/LBcMZ2tHlviBsv77+N0gvO0+8guaVYsaSsEesjcCkuPxBwILC4C10pe5/Bqlbkg8xaVNtcxvTmQomSEx2ljC8Euh7Tl4gZhFVP1tgtU0VPN+YkAfODzyGtKCn6YuKKCd49JXUQo6+7ijS74SkvbelWSTM8s5lJrzuOtMI7KHrVqWIztihQ7luwOFDrJFCZXUXsrFuR9122Qt9js2CjFd5Ar1iM6++na5VVa7q97GFeAdByxM5gdkVXcsx6iCEw8C1b/6nNMwYFbjZS6lG24DkNKKajiNt225WqPjUj9WCJKmvmFAj63iQIc5ut266jV5lhkp1zJ4bYa4qBllsOaZxFcacGUJtk0V6n1EV25qH9Y3QN0pwQq1MvcYzXoLPSBPCpZ+O3uhl76aHh7pMKt3Uasiee+YsJ3Mxi1WxC8tmoS8l1eEknSzvbFoIym82bQoYrkGyB0e1j36ESmgBtpCyJyGxwS1Ci9JcKf7+lsHwJY3WKwbC7POtONeqmpoelLRI5/ylV3WY7R+Sb7U2wSxax/CWxPvl8ErrzJ8wBqlSAZA7+xhBxqiv51GoF+nm2zGCxFmiuRxnRIR84k8GDcy0XtPy1RVoNCW5VMaJKbxqDskV+lLKFVkC2POrvpWukW4KvWCMtTrGUXESsuVC+kOgFyYg6tv/VzdJa4rRkS3UXDZSgGkBwRXF32w9ZP+3l2hodvoWmQAAYsqdYjEbJ7FYI5OIynSSzgWEMbqXCeESl5ioqVWJNYMYJv4w2Tpigwn4h+doySBCHhi0SzEbQnyzGc9Q+MRoR7gZrGwvYJrOePaDjqk5i19NEU4ABVboTAL2Y8GuLQWnGvBu0xe/J6EmwPDT/NjjnZ6tFRhkfs7mzZkD/AHEH44Dx5zx2l5GwJA4e0YUU3ao5XCVNsS8fqDMfYUuxdpvDcuR8BL6CA7oh7tnxIgnsjRb5AONTjZP3ST6EXUDbqyemNTgP0VDc3XvfYTyKWZ8loy7pWHiFAaimCdTqFbRWhDF/NoP1TJ1TVR7t+sHCCJo2GBr2udgqHb39ude7G983Apk0RRViVsmnUvUOsW2paNEHawwGm6jj6WLeS4DDxdzYVJQU2wwcl7HVMhy8zvq4z9zhUScRRroyjF0PBjqJCTgKq0cKdiA0hMGcfOhAJMwydjiaELgQniU10nANhRV4lZXUlN5tixEmbnSQ10nAcftrTHZygGgXaTuKLDXcqa2CgsH2E6U2OpbiCjpg44MSU1rhChNmVODSF1q28k7SGVSlKFVoOLAYNiS6WbaspU2rKKSfqOlAErKlKHjRej89KioZ0Eq4qDDhnxoZShS0EioUWPtjEbEmhry9pKabZtdHhn0e+cP5KEDWR8EcOXS6o+GDjVKrSaKKOnU3T1qCrskYpaj2aulw13iTwSUg8Aw8go/NA5jwcbQDe+afQDCIT1Ic3LeW19v166V16mEMLb8udrSyPlrNY17TIkplXe1B59WXRKys/YrG3KurSF5Z+S0OJHXbgsBHuLrjuHnQXYh2Hu36ZNB6bwko96KAqW6px3I5douXQfV8RKuhOsXLoOqw0WI5lrTnoNVcK6mtA0OHa6vpWVLblaYlUNyZXQMaysMa4FAq1IAnhFoNaEKelI0dVYz26MadNy9Dkk6COnM8KvaSefIKUiA3yVbnuvKqXeX1ZHaVTiZXzWkCFw9X3WHqbSpuWz5y6xozxx/KmSf+qXeGvvzw3Gxk/SKhjWzFyTrzMABDT4XH3flkCGmwutGkaENn2oMco601VSe9rR3V8PcZAsAoucpE/aJJZcD7O8crG46zEGIHfhmCDeAYb/jzLJ+LnaxqU1QcX+TjrNuYTLPx9PZmCA6hheBRKBbY9ukxhhNHOB+EOPdQthUVl1kmHdAbx1SuoXIZOpYvs8PFMFjSpEaJwgpj86puA1ech7Wg3AUr4aM94GHwJW8tacDIhsomVjPWhbelJYRWxqklhIZieoGrQlfKHMVdrnvou2yGMdgn4304j7ZRZXO9N9UFe+dyl4/f2FwhISZlmHS2wc9iWrQg8/p5fkM3xXrz5KKlSmKX7TisyftsgECZe+H7PLvulRnE8CBkMc90ARBqcsuNFcO2Nl2NPrNKJTY1XUF+ZzVo5JQE9+TTLCsgu58WU0bztH3eUk0XJYk5mSTgJe5yehAtmrU4JVouvjzzIqO/8grU5SRTKomJEff1GJ837miY73t3apzvey9f7+8ctm+GxQ3erHO7mwwG8kS5gF4l88msy6+2moVK8WL5MrnGBHJt/I5/fPNm5+iX3t7O8Q+7b3eO9hSubLLYOLsfJE3tOPT8hVmdndHkybgDZLf7bvllnk7RFDMFlYZPGPDSP7PnojGBU7OCb8FnoRly8odLXmh3wR/ZWHzsJWjE8NMgYpHULps6Zd3TTwl/MjNPqS11RllRJBdZOrlpF5eT6+tkBsZ4sT+lH3hlKBtm/Xm0+/bkB6HOiYV+B0Uv/tyli4ztRkgOWugEpqQyaq+M0eEkenn09viYOSwS+Og2m7d/Hf86fjkE55Gvjxbj6N0sg6n4dT1Ma8qSpQZopxf5+HzSBLxBF7UvUuHq0tvp7WLYWYC0DEc5wwKqFI03RtgVnwRv9K8HTZ5euCB2U/UKgQQEj9YSBkpwSDF4iVjV0/rVU7s6CQO6pkiY3Pd2Dvd69JDy5KTJBRxO7k7ml421tcACuVzgoKwJEFlKG0+maFGrFUM4iwlJS1vBDI7zvnSKbbQwWoIQF9u/ju8U6HsTNsJ2LwPFk0fdF/B66LjWzJz1nphRL79cvgQROhB/gevQ5RirLaaUX0NClG694G01Eh+ToTfB8LLzdLgoZCEhCHGpQVkghbpZPbB4AYwcwFtwzCwp/sWx1nIa5H7vGs90hdDNZr2kD4NVtBHLXl/sfQazbNyMK/Y+c53htnfd7qOR2k3xU+1rSAAaVtSMucyHS9wWBewwTqZcaBtu6F0sZllT6FaCZoPuMBmlg2R76eoVr4UxxLtNa18JTLpRhmGiIQmKvm2tx22K8fRFqbW13mKKgXMggSrZYfBObwELGAKq+p8ocbi/c7ROgZN3X799+Y+Dw++jdzsnL38Q39bFdm+ggn5H6XDSvwIrwm+LYi7QmywuLiFgmw4BPs6S2TqVBetPBAGgCsAe7v8KfIlr4Ir0DIHjXYEcMkSD58z7ZJ6/z2SSgGgOYePEDrefTAVfRulskgyix5vrYBFOF/2rbF7U6N7/Gs4Wvb5oO2uOkpsi/z3rPu7A/2IeG19gOMiRiNhFcg2A03gVNcmOO4Qyy4p5b3333OnmggsyuXAI0JJ3DXw8em54VwkBsCheCk7Bw41Gs4FD2JNBRuC99Gm4vRUbl5GQwAMdaIlut/TA7xOdDppQWjWLu0F1d9NUZjc2ySORtDzV7i9vUDyLCtudJ4P7u9Fou7M1sM4Og9X27GqDga7GfFdxMlLEPzSYbQATIDhW6OQyK7JomM0xPqAKDXi7no/XTSU90kXUBFGU3STghRZ1vtnY2trYfPH8RfS+0E9Pmc/tV4yHKdS95GRkWLEaRWDzQDdxyayAL13kQa5qVxLizZ5YvX+B2OREEEm/IEFGg95wIoPb0ODE/NtlLo8E3W8l7f54+PZob/9o38UAWiEkEGhoYI51JMbFGH1qhRCwRsgeILHwiPEpTJH14hrwGUQ0JHhVp/N8o7OFo7C10XleRnUGdzn99aiDXB5OwBJQPRq/WHRxOLuMLqXr1WACc34wbwMKTTOl4vY8H2VyOrfno96tOhDSOiIKOjmcnZb0LhHg1qPNGC6HPt503B7EZFnMxAI6t2UlrdpTUN4n4ge9bDs1cYBAbSbJPp/gbCIhTvHlo2+7jzfFYAmZA251lxSLEyntwCouMa4nvzGBs8AbLQQi6sMgTybRaNG/tPEKjs/jTXtQqEPeUCg/UKLBd1HHt2etAB6o7jWx2l1lGmw7zR/8MdlazFqEhCchrcP9ofSWsf7t6PhW3HlvucL4V3J1w1XwKayBhKu86KL9V4Ix9Ax4bjXcthyP8sGNwgOWOevmrKkf+26aetVF57yKNVkDrw4FGLpZgtcWimwuiJyIHWhTQ21Fp2ex4gDRBSngiO25fxUElOlpYtFZdXGVg/xyX1udwkC/Bbn6IxZilzcquIaL37sReW02oUV4Y4dHg4M/fBt9G20tuxtB/j/DfJRL4cGGngPNi3xczJNxP+P0QCaVaY/VW8/hWH8Rm1JymtFTZ7vETUrjA+uWhVMr2nra6QQcuMpaseXzQxrcfL5Kg2x5flDvtlZszF+TH97HMBd9x+vbwB2mtkIzhPmNTw2rOEyExL7ETawdEAKmGLgoybtaodLYl1b0G3iGiM+nyRlNsdP0zCuH05i826BKTDfOvxWV0dGj+ZvJXqt822gWYsW4xbvWciizgsimwLcsxqwXdzZtmeiuD5HfaImSEly+SWvJ9ORjBHryJ0rzXlJTni8hQPoxBEj/TAKkf86ChmZ7s1bp0cA+xXH014i9TtVrZ70j184AFLDn2H3hCyHUTIM10yU15U6dmpY2fAL3ZTX9spp+1GoKhwTWegr8pT+mf+Rim+NqS0j5kilP9dd0u3KtxOjsH7Ew1jI9opHwaH/n+O2hWPEfr786eL0f7f/z3dujE2aCpDR6FCVKPpiIi93vnJvLfhWTcFRsO8eT8XoAhu2shjDA1DrLZsb8vSGdpGQww0K2ROZiMIO3or1sbzH9q/TEaGmTuD6drUEVZpf046Ebm2FP6EfsaavCRqnChewgHdBIyxNnAF+4KbaHk4u83yaWO1oMVUACRVhuYBOl17/TNJbFsIAxBKFJd5Cfn2domUCDEatEfIDweH3pRGtgiQeApgD1s+jbbvRNWfO8OlR7JIHY1TeXVIdQM5GYumItFc92XWUK4RCIZNkFHEBtRC8hCclGtJuAXosufCb3CWYnycdzyKI3GScY0QZv+g2iS9GGHeqllp3ZiXaog5lvozOZE12QInEHP6kYpRjMtixedWNXj//Ly6x/Fe3lBdyRGTS8cNLppmfy7m2GjN7pll9wq8I6nm6qhTzdWoEa5BYeJkfom6JHPWq8AVcjISXejsX/Z9FxPsh8ktxCNGPx32Cz3Cifygy1txDYWPw32Koou2Xb79F2r9toMRj3/zKEOhTj9hJCX88wwoJHI9GX201YyG+3KKg0PozoYYAPg60/eRJg8H7dpcA4CynSu0jg6mKSFk3Rn3XRHXNvNtgnfBBF/w26zk6wZYTGF4/bHSSOgi1E0VPazkKAe/fTpoyt9oLHzf8YYlWNfCj+vabWOSOX9JDeET3dS27/zp2w91DIFtHureBh2ZF7/NQ8jn0SfxUdY2qnqJhm/TwZrsPVj20mxHH1MYuRPHJoy8r7dAK0TecOcOjzWB8+4FO7ghkHfOBGfKxEI2zcRzDug834s9P/xdNyZrXID+T8e2gooh1BHUG1Y3ky07xTfbmnghtQSAh73OgExsOOCDnf5EcfWiKx2EyYmUSX0HKKlUhuGSmbAHIdqsVtWH2XhoPUtU3Y9ErJAo0MsImtGBUMFlBLY6EKfP4BfbbCgOIY4iD9nd0b4PMJh1G8DM4klyyWPNJsPCKybIbIogp8frI8WZUsTMoQv3tkwdf1CCN5yvWfVu9RN9z+/ETorD7ZceXlXZc4l3OFH8O+XEMoQzusGQS1gj2jYh8U0clkImDNGk5mMZ4PCFYIUooounF5NnRKuVCeeF7F+oYUDvCh42TgxmQyz9tPZTKZx20Z6QWs1Zgp5kt27i/Zub9k5/6SnftzZi77XBnJVF4Uyv+C0D42kZhlfF+WCyyYdawrd1gvgaSH8AuWcmfT5drSWU9CAa9YOrNjSmf2RqYzO+HpzEoX7mW7w6U5v5bn/Vo191e9/F81c4CZdB+hRHX2SFSmxTFr+o4y9P0MW7GfklkOe4eSvHX+uNUeM9QF1Da5vpJQM8ObgV6S4O3l26P9HrjT9vDk4WNSvZXm8/kcOX1Yx6oT+3zJlP4lU/qfnSn9j0sXR2l0EfIFHhPRqlgj0R23Ubtgtt0TaSmXKHArhILMhskNBO9TNUxmM64otSKfHzU2rSiwCyrVJf8VEuPJrG4OLXSJJtsv+LL7S9a8/4JZ89iJyZeseV+yqn/Jqv4lq/qXrOr/tbOqf7FNfLFNfLFNfLFNfLFNfMlw/98jw/0Xa8MXa8MXa8MXa8P/RGvDGsunkN7KmO6YVEGGd0epqRJClGYXYEGdliRk9CZRZy2UWSEYP9/NqwBexBaa6qPJtcD7x3KW2V01H+xuW8kwave9Xn4Jq/cQHw4c5M555hD5Ev45xT91000AWUyXNFF0BmnexfC1OCurguwEoFAzTYaP+ccN55nLInDtBgCbESa+7qXi36vB5BpigTYpWQq+x9TkOMVwKlWNqZsQnA2tAVZjZK0ourJQ99Tg0II0Bi3IQwC/Hu3KeL1nUh/RLHF6apo9g3XxlI3NWVV2DpnfKz5zucqC2PVeqWSE4JDvDhd30PeGmYP9N4i9pFkOmSecVP0j6eT2P8ymDhcqrBlfVvRZcqGJypa/n8D9saYdMxcaxqtR6W3zlPOdacW+0NaGK2dN590Cbhf2ryAc3FBKIecenJXt23yKA1kqFKYWje8MbtvqvpfHAjL7bBzfsyS1OjmbHJAzS2KwaYHthoUFflLJtGQIfnrHBxdwli+NzIBko2ICyA+79oc1q7/Y/owC6ymesvpt+Os+UPW0Bv+dWYkoTcQ/yEhpinGpY6ehlKt5KA8lxdysyuqu5g4gu2ZlfiLBr7Me6ODLTTcxS90cT8uyN5VlJzjOMOgbKEVv32ezZIixyL04/oAuBr/fMTHm2eqsAs3vwI5OjEFVmV0n/nw9DHgugCAOOtJ4FRKmkIOFbbsIoWSX8NAbJhcXaGeidAN26aVEE/NlSYdsgEspXA1wlwOsRQZKfU4Zo8soUWl75DQIa3glKRCW9X9VYJ+n79WGuNp99/M6fETffWChvsvEkFaM/0BqAW181WHiV9ETNDnNzF6iMJiGcHGhvFH8jVIFOn6Wa15y16u7W7uuvdK5GD1yoUrEqaeU1xiifph9npIl29Em2/0x9tqOttSez8Wlp+DqlNL8qyLvGWa0tlBgGcGremsBrFgzbTysoTNje2a652V0thZUE/ktsLDCYspYyKFAnYXX9MhNjWEtv056jPASHEqCUXsV/0SZK0pTZ3zJU/FJ81T8l8ki8SUtw5e0DKumZaAacuT/lIQMZWkKDGY6/SqXr27ehrAMf1DuBitYSFnehsr0DOW5F7w+qM59SUjxgIQU+QiDvPxWTMbq96RQv2aZ+lXc6pfzWdLPQKKpF4tx3hdEGgg1dQ3CNkDE26Q/TIoiE81QGf1Kl8ggQCz7jM+tCP4OsuE8oZ+/w1kTVcnPz4cQ5IdqHGf/sYAThDcyZi6WgZ6xMu9AhOMHoaVB3Ar5fg9u7kWv0f367ZRClAgNCoJZUfHFbCjAtDGkvqr049FrFPqt6IeTk3f40yo8A4QghCgVP6LHFnyGlVhTeiokj1iPxf9PB5qkVxAsZQYv51cSY/lKloADpEGeDCdiJTAh8AWNRPm1nXfveicHJ6/34cALkz3J9BuL+UT8ummsvTva/+lg/+fe64M3Byei1Fans/Zy5+UP+z2ISwTnY1B1nAn5UvSmxUWfAou3gS0aay+PBMf3Xr/93ipNXE7pH0RDPTjWQoq1BZYNCf7Nzj97O9+LLePOL8ei0pOnay/fvnnz4+HByS+9nXcHvd2dYwR2OZ9Pi+2NDWi73R9OFoONZJo31l7tHJ/gkX7v+B8H7wT+P+1DPVEF1KK1w/3vj94e994dvf3p4PDl/rHeHzQ6zzpPnmI4dNhOHmLPorf9Pl4eTZT9odH5pvPkmVdslruFnm3qQsf5fyzy3yZwcffeRUD82H918E9CBCbF6fbTs+0IfklT4gADmLh436+t7f548Prk4LAnP71++3Ln9cHJAe+T34ttvh3q5/M8K8S7U2eDmwilfzLAzWFygYkSXiaD/Hf4sV/0k2EyxmRVjR/yUTJKboeYLKzxjyQVG4MEn2yAr5PoZSLYmhKUHCcXmJZU/BjT+wKf8qF8nU8T+fMkGeYF/fxJzMHJLE8KBpztXxojkCpQs6xTkD0gFU2PqT/DRMil9wk9jAeTcU4/F8ktFdmbjKPjZPg+GYhx2M3GGUQqnrhd2x/PwAIc7bajNwL4UIgGosw4v4BIR/Jhkibr9HAAzA7/FmDwTIKkKubZUJARIb0R/06H9HOSg4R8J8ZiKN8sZv0cf72biDERwnrmQXy3GC5QhowVzSXOSOBJOikmMvvhUHR1mA98AktLhcfuNbkpL4iVbpNrRf3xMCFjxN5ilFwsMuKn7xf55WI4lmN0kox/g7F/0HDvjCRX7twuxhfUFvA1pfnZTSRXCa4YTPDXXrLIsdT/Fkw9FDvvQWBoXudpNpsnAxqYVLDUrRyj3xa3NGPeQQgVyQVA7P89KTyzlXg/h3EWNMxlyVyoZzRAeboYKgLcCjzHcnTE4kVDDSOViSVshCNnQ/73ZCSW9fH/397XNbdxJAm+81f09sQtG2MQImnJY3EG3qMkytaOvoKUPOuhcdgG0CBhgWgMGqBE07yYp3u/uH3Y/RkXse8XsfNP5pdcZWZ9ZH01GqLkj4lxhCl0d1ZWVlZWVlZVVubZarLM4/2oNVKkAzdhtBoAX4NMzlYFDamnOZwq0K/vJ8SjxcQwZSWl0ahGn1Sh46RS+6Z/9C+vjg/74rdQ/0zBKU0lqEoHQ+T/gF4lojXYN9OzVSI/4U9ZQlZD0gnF6aN4MF9IXhE3/baRyu8KXgk1wA/pt4Z/YwT+jfMJUElwhgq0LSESvxxY+KYgzfAB6NEZjaOReqtLwr9JOU5GfLgBOv8F4FDojbYH9IV6sgjSb3UpNoyh2Jl+1OXOrIGunkrnyYBPZ6o2VlJVx2YhqO5cP1pkmte6IJuwoOAb/cj6bWB+D94oXAZS42JTHEpTLjoKHw0RQ4seA6Fx0MQIxSv4ZYHjGwapZ04CnxGySpephkNTVH015XGuxaLwy64JvxlImowJFn870PRdwasZG+CX9NuCl+8YPCp5AoefDjR+VcDGBgD4S/VkFdFvTSmlO7GQfEjIdk1KbbpZqg4HyyVTuohaPfd8PXrqla60blOcks9GZFBHhspOtfokUcEnU05o1HApUrVURvxWJUj3hopcKK0MZeiByxjq6WDjxMfvpBJXMobPWnP40wKfOjyEhYS/ZNOIgUY9UwtRzKIfFvEvhDCAR75yW8OtvEbNGQimKqvQqiT2gYoFv0WKxFB1Bh33Qw8n1eeH/Vcvfn/0nM2jauGiIdPZHfOzY36K1bX+vdJsSmcT81Ob/yl4OeRzdHs0QQ3hrbC7hZJh71azN7PyrS64o3/s4KKJtj4nVX+W9yEvToYnIwfgCdWCiKPgNXmgHEwxmxnsVmjHLYxJhV4+4DolfaYCITRFYe84DVeNtGdZvsWcPl2A6+BT5qFRQLBkU2yuwYm+4EPwp1gUnWo1yBbp6bfVt507337b3+l9AqMLXCMIqbVLo0pOLG9NWwRYb7k95vSa03O6n9zOi3VgoBNFr311dPjo6FhZbQe4jXIqGEZbKfCr1zMyaBxzUePgIz615Te0BCaXxUy9xQf8rOFKcmbRaohdFyA9hs8KnJ7ML0R2Yb6aL2I5t5zAOaTEi/7mGis8qVLwGyfg1UJTlV9MplcaQP3I58X0ajIqNVJy9aWpln62IYrYUv5TzCp5gqsfFCrxQv2Efy1gQsUg6VPf4pO+HUnmrgniKYxJ8W85xlf4WA7QzJYpxWwMfQwtSZzGXwrSPOlfFxf4Fx4VehvXSNoFI1pkqRqdR/wt/x2NiOCrMEKIgIkY8YfCoR/UjyvxH/4rnl1E80V5OSHf5FPz0KYP6l94Z7H3V8nvIbsfREC+84zPFHJjvkhwC+/pl6+FPb6AjUc6qEpgn2xnJKzzoUxQhYY8rQykzcNnHmXph971awHVizsWlJb2s5VaI6EEQ/z2s/KyWMwuhPEkSiwNp2V4Z7lskg/A3sUZ/otz8ZmRPAgqzUb9aqrHB6i4qWi7Hkbi23AK5pp6MyhmxViQmy/06CIAjYZmvGcvnkPw+8OXRtvswc7cLqS2gr/693do75iHFRxhyhdUcB8f97Eg+z0uBtbDQpfcVyU/xcdPsST7fYFCaR6G5+qRit3Fx7tYjP3O5wvrYTJVj1TsHj7ew2Ls90V+pR4I7jN8/Azh2O/vVjProVBPVOo3+PgbLMV+f7eaWg9X6olKfY6Pn2Mp9jtfnVkPq2qpnqncfXy8j+XYbzF6rIel/VRcDIqFeiV7HTdkxV8BUmJwRvNQSmh4IaFRMPZQFmZiePMHjXxPS8YeisMeSsCoGFoPBnyfLJqT148fP/kXWyb/+RggxN+2eu7YL04Q4IQ9d+wXT57As/irn+UL8+ZrfPG1esbHr6WV9d/NqQv+TZ6KkQ7HHc9gS3IBWXOlO6SAO0hgQ1ReyBmeF3hMhDaZvMSVL1dVX549mPeYKgELHKDdJm0roe5G8kWAjuPV7BhPeLWV1x+NDyxfJ3a0Gvxk/BYCH0nd9i/yeaQwuduDSjEtMYW4aaPOh8i82VLpNNBxBtP6ehVAfogd/V8C6XsF7ukcopKa92QDo7nZh8ACs2XlmsHiX+IPWb5dftDW0fepsvT5499DHkcyhLlRmaZ0rjk8p8OHczAvpRlNJjLHKPTzYIIZ1Ifn+oaBubZF7FlLYwNb3YIPcMAGwH+D9rmCMLb2/8h3vt/dua8tbZ8l+MZrG6Ymx0/t5NdtSdNokg8XwkwcViTIoiY6brfaDceoo86kmuUZvzNk9QLFji50auf4kgWAYggmY58w45yksTNuwssWrxz+6SyK+TQfCrn5Fu3JJG2xVwv/1VK+4og0y7+tPqHPbfzkNYot8vB7rHEqh6SA8XpHj7c3xZXkndUFkqZwd7osI3cqJU9NeS+5AeaR7fANLBiALfRt55++JUtaGUnML4CgGgEtzhoAleth5uVAA4nfgnpYP7hQQqtrqApOMYIQtZ9LC0EZhKj5LAxUjcAyVgOAw3kzUGkKEyRula8FtyCaFhT2sybI/whWdWLMatC1Sxu2Fx+b/4ijCu6myoGHPhz5UkzbszaKI6YUZGLp6QI5Rq1CcpQGh/LpX//873/9839s/+t//WdP6s8aYEvVJrWwtTrCH/jL8k0x66PHrBnIgRlHjdBEznLyTg/NF9Vc9BhkZpNIq4lYs+cLWDjRNegcMbaTgcGMN6+t3Oc6mdrAv8EoM06MJosCd3wch5QMZr52AukQhTIF/SUbbJqHm0/hYi4P8pb/btCyEatI3vm7jIhqW3W1zLYbrh/RcRmFVD2FNuDCutUq1kjHqs7CZS4mWDYdLijiw+wnoswa6Q6FGiuERKfL5tp2bDIX0YUqjxjmvddgDjIbls2w2011djTFGIQ0O+BLm+nTxNSfnqGCU7h3mj40QL3INmmnmI0kSjrDieA72LkXRcHpcnfha+l75gM3odNW8DF69/Z7Ma2l35OomMgEWmhseWknEdHWMiRD+keEzs4SlVdFSF7Y5vn7CAvWIPUEJq7ntz7jCsT4UsOt4fxtRMhcKE9qmK+jqxtC1KxRHrVU+SJWQ50tK9wHeDK2+OV1yVgPn2vospuUlfLa55WGIlvu85arm8RaWXnujvq5sCSLywnZzZcY+UksCAaUPkgJXEGb9kbqMANqN2Z6A4QnfXbKT3VPBPOwytSsCzKZmRuqKqdmaemLyjJPYiknHbiPW9AoMInlmWqJlQIWDxnwUofAdbqrbtPjO2000Me9g174BElA1xGAq+oVntpke+1kv5182jpws2giHa3kiy6A2lcGZAMwL+U4vUbI0wMB1rtJrqHym1SRxTS44Y/A+SnL6osnLsu3pd/Ag32ngepCiiqyni6AAjyCsu1kmyPfF9zzCMWtG5ytWI9AxmxADQxTnbdlVRqXPvxsEU8FMIEP/vqHrhFiDCIkGIVfjFmdkPmNnLvr5MYlgrHNrC51Pkdf3TH3cxhm7zHKoPOwb1IuJa01gyosgPtOD0qx0ZghY62miegEQliUwnCxg0/dcuuL7EeKiAqpyWFa6dvpwd1eq/b7vTXfP+u13MFPnyBL7t34EJPld+EOZHDPTgIINYURB85VwIE0LyblKm3ZpH8E9Pd6v5hh/ekHGNbCUFrkfdmU2uEsYTx+gO0luBF2eKRs3jlc2zztsQkDWBTnlEDppPxeXMRlCr5aTZUf1Xp4CS7EixHUQifK0XX2+2yv8dghsn5zQoLNByRq00BazrTZPxHqKiej2apJWEpEaN1u7Ga0Nl9/ceSZtdFihPCvf/4PEMNtZkLyj/8e//iv8U//9Z91OP8PfNyJfPw3/tEevv3heb5gtwWNKrAZgNqgM6ny6fw8zzBxGIFd05DbIeLEn056Exx3VJO+kMauxGJec8Q+mpzh/ASD2/TyptiqokER3Mfmnan1IYeObmN926G21u52dbJ/6p4e7vwx3/n+L3/e+cv/+8v//sv/7WGcjc66fbJf74j/FWNrAbcl4PYaWq732zfrad5WYPXodhTYzvvs96VJZ1uJo7/rN1+UQmX0IcEprWDGi/wMNjYz9SOgmuSo1RDxkxM0xCUYRt3ycxiqzyHlpb6dHuz1dOyuT8xrMXWanRa5i9nnTcLlOw6uaCtqdQ9YMxMhpJUdAWFELj3FVP6Ty3/pCs4Isp3rX+QXKv9W5PjBb9+kl+T/eUl+RiPCRZAjiQZdqCaDGXnbzNItFncWWg13wHDTMcO/gcbyBiNI6P51amILI7NhexyAFes7CylSndSaEgl4wiebIH7z+RSL9KwQcB3a/oMKUSdR1VqzdsjYMGox3ISOMkujgg192pIb+4ulrlPa69A2tJQR6pMkQ7rk9jjblQAwilbJeJHPS8Ghco7H+LQ0sHBvh2FF96Ftd2oYwslzsdoNP78SH2a6PvhXVcenKAYaqE5VORm9ayfnaMgXs9VFAV5VGa+h5YcyMBuk53NrK8uJwsCcOs2wwsln9A6szV0ftUu0mk8sV07+nz8f1aGJy8g5XyiFekvPbDtS3jj+lhdjW/Q8gbl4WlvBQxVnQJO40tCYqP1Ps7wkGHdHFO06tOcoFvF7GHRNbcOu/rXJsXrA6FArbN9MNvOemsQimt53oUC6KV2FPsr+2fBhAw8CPkEGTy93PC+Cbe9Nh7sVhDtBk1TTCf7BpDC9dnfuJx/hZJKC1GgPx8y4OMudfeaerPb6cxtC9bQ3L+odHIPUwsdQKSy94IjlW0Bmu6XnjkpqiyORv+QGmWGGPsgf2FHEm/45qdofla1sMTmJQ9go/1D+K/LoWaybKkueH3kH93LkE2x07PPQRmK6FTgh3iwVsiysPUgADhBOHnDrrONaABzs7o/kMUc8gX3QBwhMXodx4Lr9U3AOVg2SCWAY3G3Ev6urAPua8mBZCiGBALjLDjjqd/DZHEoI3L9Lnj153v/68OmTR5Tvov/N0eExLM3Fxy8IQQcYtobNopuurg5270IvEb8hBJXJq5ThMJKjXOCUvxBzdA2TwWeK54+J6GXm+NZmjAMfOUDUMmcKCubigsFgHQGg0YgBQfU+SB2fcZu+XMzM5+zqSuiuC8GHkTUaNukOeIVY5ct1Cx+vm3auL2hU7VyPRhsOL7npB5aadOQkq873OS2WB3FzwDgvtoIemYRTjis0Cx1PIc94uhZFlE2pV0CVtCVvXF/alydfPrzz+vDhyU61vJoWiQxggO7Hnm+t0MV948mcYfMgIozlPQQvsj5FMeljCLxyelmIZmGUMZzl7TgprS1rAoW/y4kwbyl0ygU6VlNd4E59oJJs4CUStuOm75WosIbMpZDHLmmHon5wVzMewaTth3RwQVUcE35Z3/JGw6lWEQe+we3EugODs7BFvJE3vX++KIbiU2Uv7eTXCbIhGu0Ep1CrxnZyfWNG3VB5VAAWSHCSGbyUYUHGPcC9daY4+Cl7bXknQoKDBwOFYo/tbVmLZE0YcMhQaQ9zl0Nq9XYdyhg1wksaoWxSSuokDIarYl12s/Pwyatvdq4lraQrAmjkNR1NbE1CrYd4y8j/btUrAG3RqYGXdds9HWip9G9141jgvpO9MlY984nXNV7fQxf5AvHTdBUEQ3r45CUeEzXuMo/4mq57Zl0U+wV2odcT4NCYifVKd5pfDEZ5IpYf706JNewYVGvdcM+l65ubrm9harSHAHQpZcGB7f0VtoV7USxzuH9hBW8xotPPl/3VEm4fqRhpnVn5NlOx0TriG2xFlmOwgZeZlzUMJ8BqWKIsoL5zY091ZFzalltW3pLBSO4gSA+Anh24Ck2TjGLLHfKUxqkvDeOAWF7+AJHUCcMRiMt80dE1lhD7bVYkKpimNU1RCh/3tiPbNsAz1E6SBiqQskk3MyHB3awoRtKiGN2BSGPULrwiJU+okguZBG81g7l7KWwPTFcToOF8UizgCuBV8vYcdpPVPeuOTUs4grYeaRWTwcpOmXtRLM4KbYUYYcyI7L652eXYJLJZXYz814ELW1WGP0eri7lVvIVbe+wFbTdf33i2DcHQLGpoxwnUNnqmZflmNYejjBvP6OAGBlPJ8xllhgITVH2mmnCQ4tkk7MoHnJywqJOOwabkFEEw/YL8QKNV8hUSltVZe7oN6jtvi3wX5IrfPBeDUmtee3TOCbslDZvpVhMspRWnC+0q7Zn4qEIg4DIgX4DxrbnQjRDCwmMKxQS+i5aN6KDpiLWJEPl8NRWMZGrXyJeFSswP9mEZ2sAR3wrxwISJnEnaAd9yAsKZVabbq982ankHGwIJHmr4Ld7ys4CyAm4fECUuG+zuJzezn7bNQl6BDCmtVu/45yMBrpiTlmErDg/o0O1F/GjViVXQdvDYsMVU5CkbtFxDUH+p2dtoPy6lem43axiY7ABYfpFLFjMPms1F2lIpuHvQUmlA+AmHgqmc1BS6U46qh1tjoswn6+ZuPa3idDLqpPaFEmiXnG/GBewVw0SRrRZTs28AloVaWP9J1CvDnwJQW96srbrX6euqWOwcnom1KVqm5fcQPeDOvc5ugmFLIWopBC19PHl3Z6+zm0qeYbxzGUc1E/gpNGy5Wnbv72Jwc7Fsn5czyxUefd/Ve7Guz0diUT8qwMjL0tVyvPO5ZLPsPTYNakcrPJWYQGaiGeSByJeY9b0K3GLE2LexIjj7HliDGoYEOA2nUjrSybK4wCN2inxdpc5Q9vGi5IC4ezQFaDsVgNYcEkJGxLQ8gikp/AytBRIyBPQPbXWr1rXGJwLxr2+SlRBNlDDtWuQTYZR8DaYsxubNxunrmUxwL0y1fz558TypzvO5jM1xLcTpJnUOQ/is3gcwsFujmzl9MWggTaMZEWJx50e3vblj9IeMfY6aqHFxuZ8h81pamxONcTh7Gi3piiwXZo3RmKWcSR8b3dlyLtIDB3u2yQdiYpl7QMgB22oZKX0HkFJVwqoNrT3zDgMWT0baCDQa09iMBnqtvRgL0uvu2MISDxqkVpK0W6YWjVi3qMTFcgpgvbazcjzt3dTbYFRdK2zeCVYTWtR5ALi15W25nfa2rNV08/7BZaORWph48JUvjK2fvvfgzAaqgdsjnx2ELVWPgRTtHQ7YZcxmy952oKQtEwv93LROyaAYmlOnWm8j9cqzbi1BDO5kBDYw5DaRzWV86fTTWVGeLfL5+WTYnxaXxdTrBQfxBvtK8T0lkGyrYE+tiBwMsV2mm63gaHFxssHYYyYn8rnl70SpAQc+/z12Vy+w2e1j88aXpYh/WSOIySJnDtakGGQjVwVkcJCG+E9ZB+sOqtG8PH/jup0FvdS2hKDZJqNcockmzSLIwjvLdpvq67P3bH6R24dweLcjxG01FPabMMn44V3ttqFviWy4jfiwvICpSvQQbOjRaeHFBKxEiM0GdCWart8ml8ViMr5KaG2bTyHCUZXkZznYqrARCQHKpuYoUky4YukC1y5uv6cXW3VB3g5llyL5mdnXa1vhmuA8Ew1WM7bM5w5mRaKFPdvro1VKIazlhWB5NZxMVM5tyJM0W3b3IR3NTAwySDumF0/qzkw+skkLkmPsZ9k+tt5iBMI6jejz6jMXrAF6UgkDvajOM2eD08QHQPGXys8scewVOb2zBgptAIQugyKYt/JDTvm+DAonLZlo4MGKwgw1hDBuRX+Emj/Z3T3Y3U19t8344E12TF2/65r0KeD3UHX9JBwN3VGoWXx1pPeO1LUo5LexH51ooGhO9noHkVMPM//i8BfKiw8SPZUEvyqt6X1Vm+JSlKVUYrKvBjvi9iY3iYYsPApEBmCwW5GdcIaodjt8mM9GE+hi30wHu3pwpXf9DsJsNmY745zCiUYFgzZ1GLtDVKFKrq3CKtWIsE1n4r65VUey537I3O3xVsC8VgzwzWRIjggzkgn4WXsyeEovevWHgxG7VFLMbVr5qtZGNd3HfMSdT62Atcn75NRtQM92yAh2ZKzQWounyYa0uTftdLD7ITM2kuSqLutuR7fCNcS73xIBBR86GJdiYFds7G/7vXXOL63r+Pm917zo8X2s2Rsc5Iflt3bdVUOXEWi305yj/TX9gnrJXRyFhHsTAW+KL6j5Tt+/p91Bog8E1fky7KLYWLTpHxgnnERvrLgfMvXCV4YervpRYY0ME+w3DCeFzFQeEzIu9y6pEWjTffYQXFOJFXGn0Sh7z3HznmPHGT9uR/oFbqI9+cEEV5/xeiLSqq98/SAMoWTr2D6dJQSmPOucjb33C0fUSsr3ZJxvUSSOcuE4nE8+imC3IIbglzUILBKCX3wEdIkDLWLiX8xq94tqI7en0urFzFy2HPBMMjsgms4IQIptFg9rYGJaGUV3PcMcnLMbLQXpjV6Kwe3dVKV7YanwvNAcqzn6bjvGnygJxTDus0ZiPyHKwCux7I+ZjAyECsniVNY88Cxk8nQTL8LbrdIJ2TZp02LCSXafnBbJGuLtUQBOaxamLYvapuj0OPE26KHvptdRv1nv22nIKdLjdU6hR1jYEYwfctPyFq6Wbbk+FNttpffGhPqRd/XWxXOTMd9qQgIpTCYOjx38jQf/oBBCesSsDfaH+PHAE5tKAfWGsUB6ob5kGHDosucbiup2YzZT6uIEuqR4sQIb1D7ksdYUFRYBTaMMBhlzsS7S4MYMsgLLRRgVC1ToM8wNVtiAHLdJHufcKHeBMHvoHRmWSVWrbPeNyvDldNUNvoLkUPLxphWLk3fL+p3kGREGhOiRW6oyyGN8yNbFgXT0jglSpQuFh7X+bOsApRlMoeocosQajdQgXplRL3rNJeMJBtA0DDXo47QCK3GMfsQlv7DTm1Yjo1BOK6JwFgH2+Y6ObYXnH3a16ijcrUa9t9HWRdlDMcNXN6ntLN5kuuNXktz14Ye1zG7CWlSvdF3dZPQmB2mmoWD46VJsSELAdevZPN5Yp43wzkQYCrPfKIf5alG+4Xhhl91BbAd28HEYcmec1hkj1NaoLDo79OwsPPRZ/BEDT6Hd00hcPFbEcJ5Cl9TgcsI4RhBSLHcw1SCke2sdaE6g+RpQl0zr1qYzGGLh5FprbMLlApZh+Sh0Yy+WHsUcUAlhdq8VbhnlrY+x0PG04gJnHQ5pnOA+FDmsiPn11523tZzNUH4ERvVtdAgWuqDqs8g/bwXUXarP39Iw1HY1WUS8D2vnnOmO09cVHNhSZH15fIynt3SkK3QGNOKmBZ0/FgQmF+Wo6AQ2sEzSGnRdbgdCyUACm8BHPpXFjtAo+kFVacmIJLF3TgvxnkFYGmpvH3iHhev6CPtHImtvvU/nOB0TOHJ/rNgPfh5w+Ya6bhBxQ45c4ZEODzDDlKulmFEv4egI0hkUEJwXT4A3uteT41H6SAoM3Pqly8f6Co+62uPR8tukFFCLt+BlOlkCB66qZDUbnot6tMd04OzfETc6WN9aK2sqSh1XG9jczTVGvXMrvxPv6ijf7QBJ4B4HrYBcgCtp+qh8O6OWUd+SuuODVnbuskyumepE5x3Gzvphpivk9fmVtIXgLeF24XSER/nQMOx0UiVGGjrpew0mYsuHGEr0+AHEh/Eue338FH2i28lXr169lD9fkQu9fDKO0+3kxQn+QM96geJvWU09BY0iFYmYMoQEjG6lqkLKaNzklqE2fdBlQExlgvE3rY+gVeyACqTuFl7sBNGyP62KPupEWkdQlKBMvF2oUCBmX527F6AlpbOm4SuSICxaH/oaQSynG1NIZmJhlboCBS5PW1LwKQvvKUDT7Xy8ujljxaEGg11+IltIHdHgyRydlxNDbPdohRNrC7ptGpTmzMW8889jzDfLSdN1Trd1IJGm/b8BB49fQ58jMS9n4Go1gfFJYMaZzzK6ibN2jORzIRN6qSkHuediw3Y6kHA7nZ7jsSW73C4R97dSyRBE+czztwkfyBh716pEhnhQUXDoIEXKSQOJb4upBBRIOR0dUBIh0Q27nc8//0UPhVM9DGTaQ1IGZiAQEJ35HSQeegS/4ZLIEayVRw7sSmUQ3xcBdIYNA8r7gC/UM2WCwr6iXE5VMSxnI+u1N9hDTFcFIPWSl2hKM7ZNC9KWvafksNoa64T3C0asY/3Y9BowexXImxr+zNWGPsdRtfNaausn3HpnEvDilqyp/4uuGSr4LcvY1x0LIQZzF13wKQZOsnDsdu7f8x3yACaitGhMqxmcxTxqJ2GvwPcctCzfpDtow7KprRnbZ9A9zj11/SV1shRn2EIaVrcifHuqAXvyAHgp9x9cfSejp0U9GPk5PSpOpv66QuvdbZlWsVpVTcXIPl2mV7yblGqu76a2fZ8kNLP8mJ2ILvISr33DB6bww6dPU7aTRGDj9FqXuvnhWldykzYXC8dRtmfvRTUQCw3Y0wPX5qsxa7iTqicdER8KM+E6MaX4lf21smiqdmRtP7Qnu7lox7xIAuJ9f1cfAdnsay7fyjRZJ991ltPPWPotshM3mtjHHQ6OZ/jth4ORJd1WdodpoxGyuUUakOjouswbKZYhoRCohMy0I+MbMrcbjZ8zJ5D3UiJhJ6gNdcimLIopD2M7bjrUrRU27CrdoQAIdzDUwLQ8mwy9BfeoWBZDFZuxyobldHUxU6sMPaZjackPpJm6mo3iucvBZodRTCLA7ut/dXT46OhYJadh7voUgKVtDlJc2A5e9OfnHJaxubNnvx9idGs9klQ15WJyNhFk4uCUDbdE0suLI3ORq4LOJqQy37fcCAqIJ5SjyC26t7u7tT6ae33eo0Dd/MBUrgVwKSe+aEcPeh9G46014Hc7+fxeKwiOA1ET6KcLWoP2N2qq4//pCJ+hWKLh/pAMIGFC3zATEZpSHF/Tm1Rd4bzpTKpqNYCMSoQ5wo8A0ffvhWLx29WzINSqfnr1YxGgA1+r6uHFj1U5hdqmmrPrVCy4RRXyBDhSP07aqpwJ0gpvRXH26sNSqkMRKzbhC8pKsvzofKJMxZJL0l/JThIcYdT0bPXRWIK4MSu4XctGlbhqcd02Q4NNBEfFK0WwZfliWGv43zhZNnDyOqWm9tRmhsBmecIgEL94OpnNV0uM3pvRhVPl5fJqNZ8Wp/NR55GYcx8vVNJbOVNSbOOy6uCRE6oy2HjCI4TTvZ6XNBnhBfs7w+qSObSOxrDaG9GtVfEJEbSTETh1dXEGflMU874MByVGPM84xm93jtuaGHA2QUcxOl3jFGDerc67afUOMzqJHxfqB8++Bac+U+QJUQdHZ9PHikUtNwuPAe9U50WxDMVC9YLrpM/LBKGpS4C0t+XizaAs3/CUPKToCbAbqUllZvUYiuCSpQa+y3AyVgd5KpYUYbbeJD8kJ4DhILlm+G504mS3uTyUEHIW6u0kL/HmN5zEJA9PvobRj8zu6NvR4CwCotWvigWcxY7GB4ktlELEQ6s3AXSCRfTWOIwstNtLNJRG445nK+mmn4pPvY4gdDrLM/CkEfMLOExqRklIXUt2CneYfk0B78ctuvn9risqwR+K0Wk5+E6Yqbp5kJ9BNo2y8lBeUGm2auyxJoH4OYmEPMLcigEIfAG6YeJVsl/dBPUi1hA3x5FDD1WVwT8YbU38xcRGYd6iy5L2Z7LyiyyKM0ENBcoDoECmEVGJ5cintCjLOOLLj7TAnGA+tuUfYj+JPalrg4PWYcw+k9dFyGIKQ3NzisDRvgkDG9OHQMmqiABLk0N2kmJCjGj1PTVa0y4DowfTxemmqzeKYE8O/fHbtpG2OoKAzITscMInvotmnXTTY+uag1Uqit+/OpPtGRqOEVxDNcHijJ0zh0CMMNyOGiZUtyFH9d3tiNEiextSjDDfjhg2KJqQY6XNiahkpbEBr97vkxPIe2qV2NkJx3TgBDoIjly906PD411FIOmyCOmYs1VMxQjL2A5wFwHUN1zlWBTzZgqhtx7p4DZ4I07m5DBfIbh+n4dtD4M9/fJ1+MODw+PD518eRorpQGAPXzxqSkANrKrMAuETMA/sLA+Wa6JX4BJHmAXFxXx5FbA87FAYUhCl16wrwlwsbqFHtczUVqak6pYVgdjV1iPl8pbVaMmtrYvL923bpXafdYWMpZ2358WiyPibWZFhKGDDDzmI5ParjL6MdyBAM2DwG3iZfS/MHkc+OssSAVttlw72xWGJ/mJcVTB1i9oAXeICkFKUyT/gqOG/df5EY5bQWpNRKJ81XfgMJiRjwUEo+txyeK531/WZtFsND8BuncpRNTgvWDupLPwaVWB95acWiCESmMOmTh/Fuo1m1LWdpimXEItoQquJ9kJEO/VPKtRD+qzCwRMMHdeEyXY1p84VfntG82j0iVA0BjeS5wW7evrh2BrcWw1XV0thsNeDeLa8gDFREbPHQkTEpH+c2xdGahyyfFF3hWaNK5pzlFoXa8PuJ31SG+Se3dhwtAdORX0XhJOt+ELdZHj5eFpb3q7iJuL80/En3Jw4W0JNZ9tR77h0y+iYsZuaroC6gYjovReOKCDBoMLsoa4yNjukaZapbBAW3xUJ1kgzNTgTkaqB2Wg0QZ5mm01pLZgRs6063hu6HKmylLdFl91RgZ4JAVj8aW+91yhpSlJj/RTv3fZWfBBxhI0707sfovq0vyjf4jaC7mQ8qGaHvB/c7NLOFRYNRnc4zvyBtVU7AhFYX8VAzVKqvaY6uf5pXOcaeGdZxToGtaOwxAWX2sj1X1sMUhs/v4LYmMNzsW4uaVOqXAyLRKwZ4ULMoBiWQgFA0rU7PH1XRxZ9hnICwFUhlvf5spheJVUpJEfnvMCv08mbAhL95Uu4XUp3ltSjRFVcFjNKTuTVBnNCLtb147Gw/mfL5FK8NaIukf9RrDhKISGryTJPO3wJAEuD914G6GVF1M4H/EFbv0d/15nxogbXYlcU/5hWu6gyarArEh1DihVx7UlV4iNY6LrWj2acqxp+BLv8Q7Gw1hoPcSxiiDN63NL+BA7fApO4IjwwX0dNHMNybd2oV/VxFi2JwBlQjyfXtDHb8fJ0iR+guptlp1YjN5zEAkqDRbuc5rM3kq3VG1BKBrz4E+xmJP8Y3GWJzkotuQtirYDgNopdVSefXQlKDpyZpCO64tSG1HUpbtRDBebJLWczjiImQ2odFOMx7MlCfpv+Lbd8J3A3PnCDL7IBjOe6VG1HiLdKtGzOMKHn8cIUbGTJozG6duGedvaU4TMtLvMZ7tljXAoAZAXYXq+6PFJ7iMpuddEnqwJLYuFsNbbJePsDob/9czl54L4YRphoDrt4G+Pw7DzKtDIOrs+LeDvj4OxMx5wSPHx6dPi8//jJ8cmrPiSFxsGq2+VIAA0/5BCKIshA/FCObeBTLc+ePHr09MhUY/ixWT3ucZtX0dND3hrFxc0qsY/RvCpOXj8WqgvxG8Y3qAGPzYNHY2qBC3U8O3z18Kv+74++adozZmxidWQJgBN80x4yNTbupYZ11vSWqbRRjzWsMNpzpromvVdX2/pevNWRPZasUSvWsb1lgsjh/PrpUyYzEltz+bdQMllsgrZBJwWc1WO0+2fAp5bxEbRyPJ3mXDevKcSFv3kpI73Ny0gRZAV6zfj+XiwJKpQmxIb1wmYlN2OPN0ptFtGhovKnjZ2Ha3/bloHvX5Qz8bemBEFYZUbxM3f13YK/KvJFbQEAYH49Vkt4gszFRR+/xQ5MeTmy1MzI0x9DyuHBk+NXX/W/OTo8lmk1VU3kzDWdiAXsbju5Gy/77MXzV1/FC99tJ5/FCz+S3gLBop+1k8/r1YNLQZQ3prdd7uCHlmY+Ewt3vRejPV7piB9hmyrFa1ahkql11ek+itenxM2tEN6zGrVUepYGq+/R4SvL80Jyh5aYYdq4x6BcQoqlZ6S3mgEjjwOgW3ytqSnT6tRuQbblrk4jJQLt0EU/SdId8JJchyLQvI1xeK3eYvsNwKIXx0++fPL88KkzGdS4V/J1MPMjtMwP8xTxwxVVa5ieJQbcMNVJZ4ypB9jZYxy9AbLxW2ao3nbJGfH6IY5cgdiombGpEDMzDlCzxzhyA2SjV8Zl3A6039SwnsM57DdiYITk+MUf+s9fP3twRGoDjjuKbL+tdquEHO5rkcLFRNCDjguOtU3SQc8qk1wI38lIVahYrm+cZKo/czc5+3bn1N4D9rqGbqs4Rz+x3pNqs2YXTqPwNt+ksiA1welhG2nW9pk5iNIu9oqlKsRTfXP8/bf3b1eD3UWrDYq6WONie4Ny31ZrA/k72kbYiXz/VkHp+nZIAmLNwI3QLde7RFFvvYg2QZ8Rvn87NIr6xnB6Yi0yJ5ZyI7N8i8ns4Hwb5uAgFYF5rJX8Q7d2IWcgW35FP9SWNWaB1d4mA19v8m4++BsRGNECwiTyOEekfCg30yi7XYpaa7VJnEfNNcoGzIqrljqu3d5fdg3HAmS11qioOOPq1NQGrOL6qo45t3byXcMbpKPVWO/F+bJW923AHE8J1nHow7gmr2GToahlWXNPTvoIQPYca6F3PParpFv/n4CghLcUTvROAio1OXz96oWoIXl2dPzlUfIS9l4g6kOSz+dT8J7WcR7LGd5TvTOHCPnF2zvFO7hEmKwwwiQEIK3A0UtGmJQBUTFABKKjqOKVCZCAbmF3EmV+qmgJyWiyKIbL6RWWgjunlXYMSfTOkDDspvNioQvBYVc1uRAUjyei83HPUZDVgCdbL5487Use9E9eHb56fdJ/dnRyciiYIRbCsjnABODBHM+I4UrnVWqVpOSkgntfyVsJeFoJPLxSwUMljv6yFFJF5zGUGRiDf6Yq1vzZtBwI9sTJaifhitUQG411bJQ19xu8QLnGXIYZ3w7jQX4veti1eLwSvRgQxWJByW33QPBVdI5y9XlqdPnRcq4ymhNYY8MIrveJ7RgXmOQkk5wWtLel6OBgdsKYU/xj+zomjl51I1NUqy9jit90D9NzJjCV0VLSPE7wYqSl3Aw1GOkbeNMJBg7BZoulvtBKAQOvBLcJU1MrWFTGHpFYbnfJwoppUrxVuF26sEEWz295psz/+1XyB0gCLRT08I1ANNtB5wISLNGPdFje8WOdTMSiWhrKinQzDShGwRv1NVj34bQqsS6h5iDMI8Rxv3O4g65kVDX5ooFqVCRJtSUdQyr0jwOafSIRsaKS95jQ/7O8D7UAtZp+cvvw6ZQYTJt/YKhDHj8qBmnY14MNQLYeMIJHR5QgxOar6fye30w95jx6VDArgjAtky+klrq+CX043Tb+BdtAU0AnOHsDvIQY5NvuJtd2K1wRc01oWhMvYqpi+12xurRnQ9OaTAFTj978itVC21pNq5DQBj/tfsWQ6x2txt2iC7BeUQvCWC36/KNpLaaAqUUvHWtrodODjeqRRZyaaBu5vq5RfrVZTVDArefR4Tf1tcBhwWbVYAm3HthZj1WkFrJNq9HwUIm9TI1VAau+pugR1qD21nPRAXm2ajwUBaipQayKon0glxuNe0DBG+xqMbHNJgRYO9zawhHWNUyPdK4/9xLvch16IIiBp+Q5PNnHvpYGFIDP8DEIaTSYgHuaR/BJJSRAjlQMqCCcUSZAnngIQhllIKAe6IcQlBzKCi55ho9BSBiKGu6ReAhC4UjSYN/AkwOnh4KAeql+OzAoz+K75wfvsVdIJTBWiKNLjhIrIEb9DuV5V8KLslHZadmlv2ubzJDJjEu6H3BPR3gE2Jjr9VBweDJb2TfobOHmTz/4hoVNsMrwiyYPs00vwb1uUEy7tszjmMMmiT+tA4LRdtNli85Btre5vfqr5Cscunfy1Wiy1ItWyP4yI5fU6ZUwdy5KSOcCQ12MM1giiwE4F2OQLc/QkBotyrnyFbU9P7ZDXh7bbf6eWxnWB2MWtGM45cxuFTNTsVMM9hjsGytQEN+GL6joz9Z9FBetWvfC3sjDrwTc0SMoqJfD9Eo0/+jpo5Nt4zfiLB1BtwEfVXTK7umQPGbJ4dbhM+yAOou3XjspIJ5S1d2enAkLvOC6tnZPQQhQvlzqW52wsBUNsPN7QJO2MT7BtrMDQdsyo852sDJrGyJckckwwirZdncHIH5PKGWNl0eltqljl/rqzUSMttFBgjlBtkMxF7bMYbYJndWHeJJyKwIGePAD7bmEv3mxt/Qmg1AWLebZva72ekR8U26DzR8ZH21t0/ukQGIMoM91bCCIZsxoJwjcpuQsBrgBk5pUs2WuQiLjN2aXt3slRmlcGp18iUCf0vzX3hyzjQHVxMznNN2/kbuNQdKOy7cwTW4Hvh8XeVXO4Ks9FEKwL2lnVQDX7P5ZxW5aW/WJpTB3G486KBke61xLCHEbb5kPhKWW4y138xiGGthQAxbdUJZT0324kmzTXqUNNlKdciaO3U/iUEKJz8aTswy2nLpx7I2Y67VxUNvGwS+zjWzEr2b9UTFazfuw3w66yHohtY/1LqOjAgyXWB2gpdRGELxkVXXxkt4BF9NYdRamLn+wEJqfKlZUf04hrfuyzCivzgdlvhj1VaxHqa1oVzoftBOoE/6d5sM3+GMohm/VF685wWKiVNk047h1czapRP8KCAzoWh0PEzdBF5miRBgmM1FZ0d0uhXqbX72bioke0hd2t3PxazKWw50Se3a3ZYi/bcyR9hZxOfkf2R0ox9UY/rve/n1xZRTcodBvjyfvTlAUwcTA2JO1Ou2m3RDpQ7BdXkrTJYTX2D8Ozl6rI+YSCsNJbbQCcW7LCuDe17Za/7IAp80VrOoFGDPr5aFWamoQQLm1Et3sSPDo+aPmx4JrT9NknclDEN7k0UoYq0M4ubsj3zxeff/9FXtPB2pEyeFILB9lbFI60XsLM7R0UtxRB4a5+j3SWOioCszTagV3OKsDge5hvpg+XM0no+RRMRUMz5PDQfFdsSiT5Hc7XyT6SX3V8KLs23MIjDpfFFWxuFTnm/o6+qvj10f9R69fPn3y8PDVkaDp8es//pG9GE/Ltx2pdzSRQv+f9cUy+o1MS0EidCmzq+E7mSKtI1igT8zE1HCJ8XqP/uXw4avUO8bbtaEein486Udg92xYQe3R8xNYmz1/LH4HCuzbBU5eHQfBPrXBxLru9+it4wHeDRH7/OjwmHYM/QL37ALA5298qM9CaCOwvzGwYkGXpSd/OHwJl4Lh35dHj9TP/tdPjv6Q+omZPudj/b6aXGiCAtfdhj0sr4WGe9oWr7VMoebWwN/1mVgH/mmUm3Wl9jeVrb0tS4otTqLpMirA0QltmcwwuC1zt6HaplyRFMO7u9vZFfPoZFjMqqJf6c9qPbPclR1hcIXGWwV3rjDxHYUZBxiBWXcR4mnQSelXT778KqVw14pewCXwf9FN7t/v3JOu+M+OHj15/Sx18a/t1TUVfL6ugnVioMpRFADEuS8xPn3xhwi9a8UkgPWui1WBHh/hIFSiUaEbNBpF0DHk5Y5Tw566OLtHV2L35MXbffV+n97r5z6aQ8yKY6/NhUdK+Ty2H1nACCSb7adhWmx2LZguMbqBmcwWuSQe0wcGvyPN7CKwvPtJGO1rdghLqJwPxAnr/jHaD2wLXvDzQkzEE7E+MfcmNEcVC01QdrFwFCvzqRhx+7vm9fxcMGU5GfYvJrPu5+wDMpsxsb1lwnjxxJcu17SLP2s3u1VQvtHGPUkFNmW+KDGEfkww6KSgRkIIYN9Zj5DS2KwGS+T2TY2qBpt6YygscMsgk/WgmOrbRcUyn0zlKxz3WgfSXQRYzlt6T+kHltYGMS0pb0B6Alx9YLJTqJ3sjFcIekUlLkoRXWou3OIwAMUaL0ItgrlV15QcDsGWFtbdH2At83UuTDC8EOBFvx8TjWg/CvCHoJTh1CbJzPlScijkQp8OJSdkOLbw9bVpsMy/JTlL6e3wFxLLL9FL9owFzSf4pnstYdX5OKV/2O4ddHbHN/8NGZexTlAexZgigB2Uy2Itrkla1kUw2WwzElMyncdoOqOap1AluAjMFRv1oRVL2zxOiUd//V//BszRbeAOD0RPGyc5aks7ofM4B1xdAw/A8xqhIlEf1utgoKHtlb9mPL9pJ6/KZT51Si7hnVN0XxSVWaX1UMJMG3TjWNjuq+kSI1aJf1XukmDuVEralu2IMbuzp+Z6WIAsKFGURHHgpptYuLklvLMpidkzEEVJGhnmg5D8tjQ89FdscEoNVoaIFaIF8H+hG3EQSoYqkPkpVIku8XfLTTqqJtulkF9YTHK1hAy9rWJtR/M2WYk25ud9hQt+m6mbPpgHWYF8r36ur4YrK5b8ra2nt7Y1o7V4qoHyjY7E4M6fij2KC2Fs8NTPu7yp4u/AvGg+YXvzdLRKnIdde0VSYjEXSWEcZiaIbDmciDIapAjIjq5tMusrrEZ3mLk4JnXze82562Z10vFi9SUt5v4ZaMwuN0BEEy0rw9ghYHGw4e+i8fQD3W7ui3ntXIzjkRhB0xy233UJNcfrdlgmA+9TTY1tCTiH516qaJXliPSKmsFSpk5V3C8FKVc9v4ak3/fgwmBGZXmttnLa20W1hEX27im5nekMS32cuFxeaIiMDT2raVtWSqffGZw1qbHBZb/OSmYWd42ZjMa4jee97HfL9MYFJ98dQYsBpxGbaJnQ0qJAWQi0qSGVNthUhr8x+9E2DVHO+dSacrMktc2Sll3ROL2m3ze/VUYVk4dt+gY7s8bIe0Z7lNutG8u2MSduUooOqJeN0qZgbQfENvZaWpIHkir2hc2kB97qlsGZvQUBF9tv8MpHdx1kLrLAvoNT5xR0RZD8m/hC4EeYcaPT6taHnz2pgchKENwPsKK3Il+h3S11uRne4FYuWdWSIXf8T/stO4aWjFkMmFyjPoLZUik+YnAhMa3nKkwN48gy0Fr42WynBSCGIuJTD2uDOz1gXZ8ILfRbuaxRI1RY3WVqBf/z9L297gtOPP5Juh7f9+93dn2RUsPc2rcOgEWHfWT4u7tngQLN9ICHiCsC4LzSA9i+ejXQUB34HnYz8NfnCRDv70NlYlxl93/TuddOZIZgGkQATL1H86s1qSY7Ypa+5y+y7Z7TvWZqtmlz+o1t4TtwtS0Mdpq7hegU2aTbPFRRLc7aub4PG/SfDNZLQ9paMziWs9J2tabzz28tQA1bvyKoXwX5Jv/H2MirNfmdHvINf6ulH9v8t5vyN2L/g8J63zVAO/n8Ljt82WgpsG5i1cxfO8MGtlh/aXa0dSj6Hvb0Ot38kfTyh7Su5/lkIfcGid3osSjmUJKuEegS8VcavvIwAiPfywBIuCajN3I40gMqB/ophQAfJFpw8V5eye/6UahSA6R3fSzk+lltkVhVWAhIiCS05BU9zacg8aokulb0SXFJ+mbD6WokA8Y1MOS31D4qwO+RET+mCmQQSU4LMTXl8fH2G5TZd8qoI4VuYgLbWT1JCPbkUN5v283qWk9amVhvdSxOX8+pHfaA6espHXv3L9RULleaPe21cIwl7uQQqsWSz7pqXEBWD9+SCtXBxL6uBhuM4dcTuJyoQ3Xw8VRbiwso65HOW8o41otObi+/76JTn7zbE9GHW4t+1PXoey7z5BQmj+T0URycszknid6t+NiqdKOVabPFJpokNavND7bOdLxA1i0zyWfjA6wImS0R6ImGRkX4BFbbGB9sN8FavjZfqn6wJeqL46MPtEKN7ip+jJ3Fj7Qm1Sd7C0yNWXPSt2WdyH2QDcjINDc/bzwjzj07oQa7a0k1mAk3wm9bZmtnwTrcG26och8aWt/I7ozsIf8YffmBuX0LjiivHd8dgEt/23CuFV8kIIxcI+SC0AG32XN7cZBr4z2XZnuuDfZcLQjg3WQEOy0MEz7KwvQbitMvQmAKyFW00Qq5svNz18i36ugPbGoHmtqBpHagqR1oageS2oFN7YBRO9DUDhi1A07twFA7UNQOPGoHH3dRAozPm69HBtjbf+tLEcOUqM51GLHBAqQWuRE+jr3RsqMWrxJljnWDxUYtbj08NPa/LzH+vsT4+xLj70uMvy8xfs5LjPl5g4luzuf8tYbuOpTcPFqP05h79RNbBN/fzPLhVlz9aZYMCwhPWqh7mPJu72xEwzPz7/oG7xGn0vv7CLHslDMICbOaSoW4A2oyT84W5WqenOdVki8TgbdaJnARxbmPWC4Sx20BIuIoLkGUGVIIogYoLWzvcnaW0KncHYqzAAUgOOFDIF5SALFYk3w6hUg1+WQGFyExU/NkljyClkowpLEi5/RZ6d+VdEmDTL6YD+HKxqQYIsOEYmyO0bhfLvoYz1TztWVlYRQs9kGJ8Rag7AMfq750bYUWhYtOiJqCqeLFMZyI+l9Cc1O01yEsMQH5oYllLEPZlrYE1LIwlkkZQJ4gehGmFbbSi0YDuyIt/RPsNkOHiaCMn4/Lt6+fPAp8Dp4fswBO8g0Gf4WMpRCWWFVmpStpY8RiVZH1qaUTfItxOVN5lCthM+izeSGAOANi+kvP5PH9eVgyZJQ3Ug7QtPmoQxmQM2LxqdVPPcbRswVGuaUeg2iVwQJg5LIqelyIZL8IRG6MXesyAAtOq28BUHhUCNKzkKkpUPJfUYpaE51tBNBd5KKVEgcyJslri0ACKw1MAKSq/9FxNITQEi2XMiNb6OogCPRAWh0IYLWo6OZ1xnuxtY4TQDIyuP9iMSoWLtF2cSvyAes9IW596poSkPRnqwuMki0auyzhSZA5zBCQ12UiNwmLpljoJCT6elv0rn+DumfzzkxqTD+JVbOyWw7opJIeE5QJSsmLGGnH5VQmNI7JBoaFPXz+8KsXxykJiioWFBM0+9z6qePh2grl8fVEBYKmuTdcWh2IODvLs92Why9f5sr3RDYIY/9kJpKBAJB09cUQR7+iabEsZkUlZuD83aTq7nloac42eD0pILbR4H4J/T5b5meh0YYOL3AbxxUT3aaO3yoVn2hRjO1+UtqyQT+Z72JkUXdpxR7qrt0tR5kBRAUslHmC7Ys/V12nOwVJXLzw0RVMfMm6DJ85r/GF1Xp72ZBXgtMjYS10T+XaQ/5D3pP2O4pFaSN4M5mNummF4XPYeoiFWxsKa2syIq8uYMEEtPnuwZ6nrDEcIbDSTLK6rO9KpSarTj4aZW+MJUBXv/TnjWd5KgsBCCGqp4CxpyEzXdA7GhwyKuHiIAHhyRYB4Wph7Gz9SU7F9KlF99g00fYwUozCKwmWzNkGDaVj0NaE972lZxsTMEYF2HMgFSukKVnDhv/5i+WDfmkZkkZ8mAgEgvQTAAoTXuNzSwRkW0aRt2P0WPZM25532wmfR+TWCrSLPz4WNfJnxetALVRYslu9gb0f/dyXJk5IEct3J6RXbPQP5ebDsb5KzA0menhodkcgjMdqcTm5LBdPZriKYQh7XshRnSidcz0eh9+BcwJJltMebPtjVCXpnc8NH6UegjajWew4dUAommUfAxNlUJ/j9u8bGlb+92bNkoyQiwRNZxSeU6xSyadpjIr1+Oz6m1Ltd0iEFsbcOSRZmA3zZXZqtLWLxKG3B72KoT6pG3jXymFtd62vJeIsYLRpIIw/qrdiqqxaDYQMWJFKT7nm04svHa/Ua0IbV/HdFHdpUrMu49NWRNTatrKOQqm5Liax6qY25q2sVhcX+eKKBa8ajbNNQ7ARVGTRz/YQfBi9hyBPnrAWF0i+tvYNcLKKbBywpfywXM2WaBRjDEs0Lxuv6dmSrtmyfdc1RIQVmeGK3aDqRVeRXRkdB2lUbwXI6iJr8TbR6lpZfLADlP16NOZLvqooZtYCXykCaCy0wiE/sGEz9pKqhPVlNCC0azvXzt01Kgnb8gM15hb7H9b6chonzicA8iNN+5KlLNkF52u1GFL4H8QVITCsOGcYsVNVgRaTwtZOdsH/fy9YzpQ5VfC4hA0CI2owohnusVg1QAzn/vXsJhD0A4qQ0El5w63GblD8ovpCH6IyHGz0QYBIDGwKMr6nONkWdUxGwEpeM5xSLYYY6eiQ7RcQjsHGOB54OPBKhsCjqPpE4Q4kwzaU0/Y4jPNFMdcLv35O6EJ0RkoMwiUUVXiHI1xS3iVZXWDeAIf4lvJFwBWg2fWSHUQW+lN40Hbd2r0vNdkqHKQx67e5BsVULPj2PkcGw41aTg+qQdAiYnarsvQBwCZ7nwtahnmlT8xmuTwVkypRkeHgYtSQPs4H5WXR39vdbVL3IQDD2fIHqFxlyrTDxF+nzwpwtoGzxifQSckxgGSHLTAhMKam+CQ7kgXRjJV7ECg3aFDu8JNQSRSmSGkqh7Z84paG1BV6Mq8tj1NhsLw0GGpLPyZ+B8srpVOPYJqfneHpNQysGjLEGGqM1A22GeSOtk8iqPxNd1XcMmOkhnXBY1j9+7P1aF34GF7viLserQNeT61/PakJzW6pWB1MvWi0Sj1FinCtoMtovSIL9bgtb+3AgA5w7G7aBNTaQ1nh48ydQZta2Ots55gyOqHTA2jli0vIoTbFfYDQoEWRPjSdMLiSVoDpiUM4MhaTax3MA6dnmlEgB36cBr1EqSPCADlU2HsdIZKczZY6tdJyodcyTSibNQ2yEa7lcD3CBxxhIzaQ1qbNnhgnah2lOA/YUUWAel8P1rZ/U2Qfp+318Q6aNz6grm/R+gC2j9P8eqepxq33J5VbNN5H9jG7fk3UhA0FIDAD3loMAjhDDJGp5M2URLt17mymPQcVWCTLbfAo4lTz2Gh6bCLSBL8+eUCqFG7HO/ScCrgeLfj4G3Peyc/wvLIPvLIPGpdlVPn1C43rYJWEU0vpRM9K2KbnloNkj92mZjJ3kOyrO9QuLX2FV58L86+KvT08G7ZI0Ieo9+/XtdZCGD7TpPNMiw6r60zfsgNFdhppnT/CAwMLnDm2nG1+lwWtui16tauhm6Ti5B+pEBMsmr2Jia/D6RvbHs9ni9EOJIrhkJhuejTqYzrZvsyYXWKqnQoGJZbDXUg64RXt6FonQyZebHxPUhet25HUuZzYYRqhtCMnW0OztRVK68yNRrJvKcOvEOOhUH2p2QhnmbluTLRC8pfrJuXgO9FXmQns2W+jh5ssCOkgcCeJbZINVwtdXHzDnQrdfHOvQLIMicRwIbrYP3QZDQdutjqddo7S/1J7W1shCKgd8iBNoAUGxGqgrrZ2EdBWqee6jF2eH6OUMOYECrJGeZjWCUksy3gwXzdsg1pOdHY/GlyeA5CA8ty0HIcPH6TG7afe5Uen+Ah5LW1AsOWEFCa4qZ9SA4JJoaJHqkwV2NuM67qsEkRbXWzGER8ZL2aQ8W3xhohUEWMUKJCDQNI/PpWIGUTDtm2dH9Xmm/HQdCr2uC+UtlfDO1RqArsQjUklJsl3GPYXzmXe6aQRuIGckg4dp19e46Yixuh512od7N4d3WzaOUiWkUtfEm9JpUXh2r6mAxrHoTVOtzrrUIQHDz9MnWImBQeDeaxa+rS+VoQzlRq0dr1M+62fnQVcfHp2kgxtefllUEtLao2qBpcsoUbIcWS2tLZfUu0Tor1BqErlEGK5WHDvCoPiZHIxmeYLiGlzovzBbEcPbbuYPX7JvrbuZraDpbQVO+V226COt9nBtp7Z8QO6KzN3HOvIlU7GLfUQPEYETSJEegginS+WFXi0C23TT1vrC2Id12s8b/jRvZEfeHqMqc90OeyP/rFy2QPYV/kZe4HLL+C3efcUD+gMCqqbaqSuuVnTDK5kh5b4wRcpf9V5+VZmtptT9kK0MZeT5bTopq/gQyLTGqaQBfgdirxMYHA5qSBYHOpffqDv5sZbZ2vgtQZKnDoo33WAJkh2mqWm5vR5mYAXI1xqGE2q+TS/6jh+v9Q4M4eXYFgt33RelfNpcVlMhQlVLlv8ewebmeFf+8NZUV6ItcdVlu7t7e6+u/fZrm0v8paHUx3LYdBXkDohnhrhrox7J+bokOKJeWSIudhYkDtvBKStXo3fUQC53drI4K2lukHvmqzKCcpj854ejfsAiKGU4TrrUjNdfAH/Y4kmIy+nqueZ63NBW7+afK/C3d3bbeOkp+QdxPYlpbPpvzz88qh/8uSPeE/n3i4P6E9HqXLBg2dBRJkLAdVVsiq4i5UvzzvDYjLNGII7hihWg1gnLIrZEjGQdD+ZLb/OZXYq7lUNLBZ20oKgTsSkPjszgHz5I1jSn1Di73E6vJ7c0LJsgj5ckGk7w/MnwbgWExzJU/QGwLLy4pTMEItXpoctI0PIeLP2L0eQwGApSKPljRh0LetrZ54P32Sw19BNByU6IBbv5mKOdxzncB1NmF4tCtJgUN6slWT72qjruikQKHhRMQ5cVgOJ4kSs5afTgeAT4RCjVrC7m14WiyVEB4dj3fLiAsiAmjpX8NcgOl+L6Fz8+B6Oiz1U72xU+I5SgK4WRXZVIUJdohp0hHnbTt7Z78/pvYPmbDHBZWh3V3EFflWiRW+uuumsEmPQYkaowB4vkFpNNuB7IfwWduxagAcpN80T4G+Lydn5kouwgSWka8BR1CajdjJSvjTg8KN73xJYxwRALknByBAHZl1FRC0fksghwLeT0fK8u3cXGywU0/DcEdCxmHVw5y4i7PSdi/s78KLIR++6e7v446r7GdMgAsuD1XJZzjIqKWlNf/c7nKiZWJG5f5D0z0qxTgWVke3stVpUVSVMwW46LcbLtAHy55BO54sv6pGHcMuWZJ+BN5Jd0VO87MzruYT75UJvd5XyiuLb2ze4vltdzCV7j2YwZyuU1DWf2ZAc4wJEKF1HVTf9soTJCNp4kLZ8BKqNorc+Y22ky4DC+hLqEPnT8h18J8vigmbYAreo+sPzyXQkCmUBfy8EGhVwkyaDgi1n0wCnBD5D4KZXy80wvIBFRIbgO8kexH3VM43t5zYbYaRiuBoGhT4xcG021dn4h+cr3N+TMx9d5cDyBwJfz2MAbOTh9ADlght5uvU52SX20nVxOuyxFFb0GJl0LD5OhNpbLLMUrA9BmfiHNhW6UI9ji0lpBM2ajdOXwLlr4MVNUo6Tazan3yQ/JCdo38i7tdeKdXs3O9eiGl4CAG5SR17MaBIdvcwdPszEClH2s7IdRPcwAtqB/he1Ey6npzggtEwht8FsCbaJhfHUV/QKE1tMcGwVUHv971I6yy2zDEclUQpus3smveam9/oI6V59M2OME4VdT89I25HgwQT2rn53jKYoKEaLGwY4gMNtCORtZka4YyvjHqI2lpMjeEzphgxYhrSMox1e3He4GIAPvBpB5JsLF/BWYgFPR3qwFdxqug8ctNyFuQ4709JyN4cbtI1+mU+maMFfFUtuujOzvb+aY+SlfDqVR5AY+qKSdMkNa8yS3GRXm/ZK/aWsBNHrWX16eUIObEIBfnn84jXljbVWs7F9GYnRZ3yuOH42LQf5VOYco0+mJ7y3H6UHQP89hojshz/zLviQHTCId8Ag2AGDH6EDHvy8O+CDjQDyThSqBeJeZOaUSbJbOy+CixnjdeDzezAf/RNUkr8fieEByhv3ANH7DFKyJyaOxu0VkjmRHq09boa9sA5cB5qW5TzDwo8hHMnrJ7CDUxzIPRjZBOiz6eRiApGghCUMydTxBhLsfxiJlyusBGTgtwldBq/kraNE3QHvQGZ32ClYzRP0ZZT1Ebh4eJu8XYiKquTh06PD57irf/L62bPD42+SR4cnXz14cXj8CKvobP1/Yz+gkNRhBQA="""


def _decode_original_source() -> str:
    return gzip.decompress(base64.b64decode(EMBEDDED_SOURCE_GZ_B64)).decode("utf-8", errors="ignore")



def _extract_engine_source(full_source: str) -> str:
    """
    Keep the real processing engine and post-engine patches,
    but skip the giant Tkinter window construction block that starts at root = tk.Tk().
    This avoids desktop UI startup while preserving dedup/autofix/export logic.
    """
    lines = full_source.splitlines(keepends=True)
    pre_ui = lines[:3838]
    patch_part_1 = lines[8042:8160]
    patch_part_2 = lines[8182:8204]
    return "".join(pre_ui + patch_part_1 + patch_part_2)


class _DummyVar:
    def __init__(self, value=None):
        self._value = value
    def get(self):
        return self._value
    def set(self, value):
        self._value = value


class _DummyWidget:
    def __init__(self, *args, **kwargs):
        self._value = kwargs.get("value", "")
        self._values = []
        self._children = []
        self._dict = {}
    def __call__(self, *args, **kwargs):
        return self
    def __getattr__(self, name):
        def _method(*args, **kwargs):
            if name == "get":
                return self._value
            if name in ("selection", "curselection"):
                return ()
            if name == "get_children":
                return []
            return self
        return _method
    def __setitem__(self, key, value):
        self._dict[key] = value
    def __getitem__(self, key):
        return self._dict.get(key, None)
    def __iter__(self):
        return iter([])
    def __len__(self):
        return 0
    def cget(self, key):
        return self._dict.get(key, self._value if key == "text" else "")
    def winfo_children(self):
        return []
    def insert(self, *args, **kwargs):
        if args:
            self._value = args[-1]
        return self
    def delete(self, *args, **kwargs):
        self._value = ""
        return self
    def get(self, *args, **kwargs):
        return self._value
    def set(self, value):
        self._value = value
    def mainloop(self, *args, **kwargs):
        return None
    def after(self, *args, **kwargs):
        return None
    def pack(self, *args, **kwargs):
        return self
    def grid(self, *args, **kwargs):
        return self
    def place(self, *args, **kwargs):
        return self
    def bind(self, *args, **kwargs):
        return self
    def bind_all(self, *args, **kwargs):
        return self
    def configure(self, *args, **kwargs):
        return self
    config = configure
    def title(self, *args, **kwargs):
        return self
    def geometry(self, *args, **kwargs):
        return self
    def create_window(self, *args, **kwargs):
        return 1
    def itemconfig(self, *args, **kwargs):
        return self
    def yview(self, *args, **kwargs):
        return self
    def xview(self, *args, **kwargs):
        return self
    def set_values(self, values):
        self._values = list(values)
    def destroy(self):
        return None
    def protocol(self, *args, **kwargs):
        return None
    def update(self):
        return None
    def update_idletasks(self):
        return None
    def winfo_exists(self):
        return False


class _DummyFont:
    def configure(self, *args, **kwargs):
        return None


def _install_tkinter_shims():
    tk = types.ModuleType("tkinter")
    ttk = types.ModuleType("tkinter.ttk")
    messagebox = types.ModuleType("tkinter.messagebox")
    filedialog = types.ModuleType("tkinter.filedialog")
    tkfont = types.ModuleType("tkinter.font")

    widget_names = [
        "Tk", "Toplevel", "Frame", "Label", "LabelFrame", "Entry", "Button",
        "Canvas", "Scrollbar", "Spinbox", "Listbox", "Scale", "Checkbutton",
        "Progressbar", "Treeview", "Combobox"
    ]
    for name in widget_names:
        setattr(tk, name, _DummyWidget)
        setattr(ttk, name, _DummyWidget)

    tk.StringVar = _DummyVar
    tk.IntVar = _DummyVar
    tk.BooleanVar = _DummyVar
    tk.END = "end"

    # Common aliases some patched builds use.
    ttk.Style = _DummyWidget

    def _askopenfilename(*args, **kwargs):
        return ""
    def _asksaveasfilename(*args, **kwargs):
        return ""
    def _askdirectory(*args, **kwargs):
        return os.getcwd()

    filedialog.askopenfilename = _askopenfilename
    filedialog.asksaveasfilename = _asksaveasfilename
    filedialog.askdirectory = _askdirectory

    messagebox.showinfo = lambda *args, **kwargs: None
    messagebox.showwarning = lambda *args, **kwargs: None
    messagebox.showerror = lambda *args, **kwargs: None

    tkfont.nametofont = lambda *args, **kwargs: _DummyFont()

    tk.ttk = ttk
    tk.messagebox = messagebox
    tk.filedialog = filedialog
    tk.font = tkfont

    sys.modules["tkinter"] = tk
    sys.modules["tkinter.ttk"] = ttk
    sys.modules["tkinter.messagebox"] = messagebox
    sys.modules["tkinter.filedialog"] = filedialog
    sys.modules["tkinter.font"] = tkfont


@lru_cache(maxsize=1)


def _install_export_completeness_fix(engine):
    if getattr(engine, "_streamlit_export_completeness_fix_installed", False):
        return

    normalize_col = getattr(engine, "normalize_col", lambda x: str(x).strip().lower())
    export_value_present = getattr(engine, "_export_value_present", None)

    helper_cols = {
        "Match_Group", "Group_Order", "Match_Role",
        "Similarity_Score", "SimilarityScore", "MatchPercentage",
        "Duplicate_Tag", "Duplicate_Reason", "DuplicateReason", "Reason",
        "Match_Type", "DedupType", "Confidence", "ConflictReason",
        "Lack_Reason", "LackReason", "_Dup_Scope",
    }

    def _data_columns_from_iterable(columns_iterable):
        cols = []
        for c in list(columns_iterable):
            c_str = str(c)
            c_norm = normalize_col(c_str)
            if c_str in helper_cols:
                continue
            if c_norm in {
                "matchgroup", "grouporder", "matchrole", "similarityscore", "matchpercentage",
                "duplicatetag", "duplicatereason", "reason", "matchtype", "deduptype",
                "confidence", "conflictreason", "lackreason", "dupscope"
            }:
                continue
            if c_str.startswith("__sort_"):
                continue
            if c_str.startswith("_") and c_str not in {"_Source", "_RowUID"}:
                continue
            cols.append(c)
        return cols

    def _present(value):
        if callable(export_value_present):
            try:
                return bool(export_value_present(value))
            except Exception:
                pass
        try:
            if pd.isna(value):
                return False
        except Exception:
            pass
        return str(value).strip() != ""

    def _patched_export_data_columns_for_completeness(obj):
        if obj is None:
            return []
        if hasattr(obj, "columns"):
            try:
                if obj.empty:
                    return []
            except Exception:
                pass
            return _data_columns_from_iterable(obj.columns)
        if hasattr(obj, "index"):
            data_cols = _data_columns_from_iterable(obj.index)
            score = 0
            for col in data_cols:
                try:
                    if _present(obj.get(col, "")):
                        score += 1
                except Exception:
                    pass
            return score
        return []

    engine._export_data_columns_for_completeness = _patched_export_data_columns_for_completeness
    engine._streamlit_export_completeness_fix_installed = True


def load_engine():
    _install_tkinter_shims()
    source = _decode_original_source()
    source = source.replace(
        "import tkinter as tk\nfrom tkinter import filedialog, messagebox, ttk\nfrom tkinter import font as tkfont",
        _TKINTER_FALLBACK_BLOCK,
    )
    source = source.replace(
        "        import tkinter as _tk\n        from tkinter import messagebox as _mb",
        "        _tk = tk\n        _mb = messagebox",
    )
    source = source.replace("from openpyxl import load_workbook", "", 1)
    source = source.replace("from openpyxl.styles.numbers import is_date_format", "", 1)
    source = _OPENPYXL_FALLBACK_BLOCK + "\n" + source
    source = source.replace("from rapidfuzz import fuzz, distance", _RAPIDFUZZ_FALLBACK_BLOCK, 1)
    module = types.ModuleType("fuel_subsidy_engine")
    module.__file__ = "embedded_fuel_subsidy_engine.py"
    exec(compile(source, module.__file__, "exec"), module.__dict__)
    _install_export_completeness_fix(module)
    return module


def _sanitize_filename(name: str) -> str:
    name = os.path.basename(name or "file.xlsx")
    name = re.sub(r"[^A-Za-z0-9._-]+", "_", name)
    return name or "file.xlsx"


def _get_session_dir() -> Path:
    if "fuel_subsidy_streamlit_dir" not in st.session_state:
        st.session_state["fuel_subsidy_streamlit_dir"] = tempfile.mkdtemp(prefix="fuel_subsidy_streamlit_")
    base = Path(st.session_state["fuel_subsidy_streamlit_dir"])
    (base / "uploads").mkdir(parents=True, exist_ok=True)
    (base / "exports").mkdir(parents=True, exist_ok=True)
    return base


def _save_upload(uploaded_file, slot: str) -> str:
    if uploaded_file is None:
        return ""
    base = _get_session_dir() / "uploads"
    safe_name = f"{slot}_{_sanitize_filename(uploaded_file.name)}"
    path = base / safe_name
    data = uploaded_file.getbuffer()
    with open(path, "wb") as f:
        f.write(data)
    return str(path)


def _safe_list_sheets(engine, path: str):
    if not path:
        return []
    try:
        return engine.list_excel_sheets(path)
    except Exception:
        try:
            xls = pd.ExcelFile(path)
            return list(xls.sheet_names)
        except Exception:
            return []


def _df_copy(df):
    if df is None:
        return None
    try:
        return df.copy()
    except Exception:
        return df



def _prefer_existing(value, fallback_func):
    return value if value is not None else fallback_func()


def _to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Sheet1") -> bytes:
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        (df if df is not None else pd.DataFrame()).to_excel(writer, index=False, sheet_name=sheet_name[:31] or "Sheet1")
    return bio.getvalue()


def _zip_paths(paths, zip_name: str) -> bytes:
    bio = io.BytesIO()
    with zipfile.ZipFile(bio, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for p in paths:
            if p and os.path.exists(p):
                zf.write(p, arcname=os.path.basename(p))
    bio.seek(0)
    return bio.getvalue()


def _build_prefix_and_tag(file_a_name: str, file_b_name: str):
    base_a = Path(file_a_name).stem if file_a_name else ""
    base_b = Path(file_b_name).stem if file_b_name else ""
    if base_a and base_b:
        return f"{base_a}_AND_{base_b}", "Both"
    if base_a:
        return base_a, "Aonly"
    return base_b or "Streamlit_DeDup_App", "Bonly"

FINDING_FIELD_LABELS = {
    "first": "First Name",
    "middle": "Middle Name",
    "last": "Last Name",
    "region": "Region",
    "province": "Province",
    "municipality": "City / Municipality",
    "barangay": "Barangay",
    "contact_no": "Contact No",
    "birth_month": "Birth Month",
    "birth_day": "Birth Day",
    "birth_year": "Birth Year",
    "lgu": "LGU",
    "suffix": "Extension Name",
}
FINDING_FIELD_ORDER = [
    "first", "middle", "last", "region", "province", "municipality", "barangay",
    "contact_no", "birth_month", "birth_day", "birth_year", "lgu", "suffix",
]
DEFAULT_FINDING_FIELDS = {
    "first", "last", "region", "province", "municipality", "barangay",
    "contact_no", "birth_month", "birth_day", "birth_year", "lgu",
}
EXCLUDED_USE_AS_KEYS = {"license_no"}
DEDUPE_FIELD_LABELS = {
    "first": "First Name",
    "middle": "Middle Name",
    "last": "Last Name",
    "suffix": "Extension Name",
    "birth_year": "Birth Year",
    "birth_month": "Birth Month",
    "birth_day": "Birth Day",
}
DEDUPE_FIELD_ORDER = ["first", "middle", "last", "suffix", "birth_year", "birth_month", "birth_day"]
DEFAULT_DEDUPE_FIELDS = {"first", "last", "birth_year", "birth_month", "birth_day"}

LOGIC_MAPPING_ORDER = [
    "first", "middle", "last", "suffix",
    "birthdate", "birth_year", "birth_month", "birth_day",
    "region", "province", "municipality", "barangay",
    "contact_no", "lgu",
    "full_name", "sex", "age",
]
ENGINE_CANONICAL_HEADERS = {
    "first": "First Name",
    "middle": "Middle Name",
    "last": "Last Name",
    "suffix": "Extension Name",
    "birthdate": "Birthdate",
    "birth_year": "Birth Year",
    "birth_month": "Birth Month",
    "birth_day": "Birth Day",
    "region": "Region",
    "province": "Province",
    "municipality": "City/Municipality",
    "barangay": "Barangay",
    "contact_no": "Contact No",
    "license_no": "Driver's License Number",
    "lgu": "LGU",
    "full_name": "Full Name",
    "sex": "Sex",
    "age": "Age",
}
NAME_SPLIT_SUFFIXES = {"jr", "sr", "ii", "iii", "iv", "v", "vi"}


HEADER_LOGIC_META = {
    "first": {
        "label": "First Name",
        "finding": True,
        "dedup": True,
        "finding_help": "Missing first name usually means the beneficiary identity is incomplete, so it can push the row to Finding.",
        "dedup_help": "First name is a core identity field used to compare possible duplicate beneficiaries.",
    },
    "middle": {
        "label": "Middle Name",
        "finding": True,
        "dedup": True,
        "finding_help": "Missing middle name can be reviewed in Finding when you want stricter identity completeness.",
        "dedup_help": "Middle name helps tighten duplicate matching, especially for similar first and last names.",
    },
    "last": {
        "label": "Last Name",
        "finding": True,
        "dedup": True,
        "finding_help": "Missing last name usually means the identity is incomplete, so it can push the row to Finding.",
        "dedup_help": "Last name is a core identity field for duplicate matching.",
    },
    "suffix": {
        "label": "Extension Name",
        "finding": True,
        "dedup": True,
        "finding_help": "Missing suffix is optional in many files, but you can include it in Finding when you want stricter checking.",
        "dedup_help": "Suffix helps separate names like Sr, Jr, II, or III during duplicate review.",
    },
    "region": {
        "label": "Region",
        "finding": True,
        "dedup": False,
        "finding_help": "Missing region can flag location completeness issues, so it can push the row to Finding.",
        "dedup_help": "Logic Not Applicable.",
    },
    "province": {
        "label": "Province",
        "finding": True,
        "dedup": False,
        "finding_help": "Missing province can flag location completeness issues, so it can push the row to Finding.",
        "dedup_help": "Logic Not Applicable.",
    },
    "municipality": {
        "label": "City / Municipality",
        "finding": True,
        "dedup": False,
        "finding_help": "Missing city or municipality can flag location completeness issues, so it can push the row to Finding.",
        "dedup_help": "Logic Not Applicable.",
    },
    "barangay": {
        "label": "Barangay",
        "finding": True,
        "dedup": False,
        "finding_help": "Missing barangay can flag location completeness issues, so it can push the row to Finding.",
        "dedup_help": "Logic Not Applicable.",
    },
    "contact_no": {
        "label": "Contact No",
        "finding": True,
        "dedup": False,
        "finding_help": "Missing contact number can flag contact completeness issues, so it can push the row to Finding.",
        "dedup_help": "Logic Not Applicable.",
    },
    "birth_year": {
        "label": "Birth Year",
        "finding": True,
        "dedup": True,
        "finding_help": "Missing birth year can push the row to Finding because age identity is incomplete.",
        "dedup_help": "Birth year is one of the main fields used to confirm duplicate identity.",
    },
    "birth_month": {
        "label": "Birth Month",
        "finding": True,
        "dedup": True,
        "finding_help": "Missing birth month can push the row to Finding because age identity is incomplete.",
        "dedup_help": "Birth month helps tighten duplicate matching when names are similar.",
    },
    "birth_day": {
        "label": "Birth Day",
        "finding": True,
        "dedup": True,
        "finding_help": "Missing birth day can push the row to Finding because age identity is incomplete.",
        "dedup_help": "Birth day helps tighten duplicate matching when names are similar.",
    },
    "birthdate": {
        "label": "Birthdate",
        "finding": True,
        "dedup": True,
        "expands_to": ["birth_year", "birth_month", "birth_day"],
        "finding_help": "A combined birthdate column can be used as a Finding trigger for birth completeness.",
        "dedup_help": "A combined birthdate column expands to Birth Year, Birth Month, and Birth Day in dedup logic.",
    },
    "license_no": {
        "label": "Driver's License Number",
        "finding": True,
        "dedup": False,
        "finding_help": "Missing driver’s license number can flag record completeness issues, so it can push the row to Finding.",
        "dedup_help": "Logic Not Applicable.",
    },
    "lgu": {
        "label": "LGU",
        "finding": True,
        "dedup": False,
        "finding_help": "Missing LGU can flag local office completeness issues, so it can push the row to Finding.",
        "dedup_help": "Logic Not Applicable.",
    },
    "full_name": {
        "label": "Full Name",
        "finding": False,
        "dedup": False,
        "finding_help": "Logic Not Applicable.",
        "dedup_help": "The current engine uses split name fields instead of a single full name field.",
    },
    "sex": {
        "label": "Sex",
        "finding": False,
        "dedup": False,
        "finding_help": "Logic Not Applicable.",
        "dedup_help": "The current engine does not use sex in dedup matching.",
    },
    "age": {
        "label": "Age",
        "finding": False,
        "dedup": False,
        "finding_help": "Logic Not Applicable.",
        "dedup_help": "The current engine does not use age in dedup matching.",
    },
}

HEADER_EXTRA_ALIASES = {
    "first": [
        "first name", "firstname", "given name", "givenname", "fname",
        "pangalan", "beneficiary first name", "name first",
    ],
    "middle": [
        "middle name", "middlename", "middle initial", "m.i", "m i", "mi",
        "beneficiary middle name", "gitnang pangalan",
    ],
    "last": [
        "last name", "lastname", "surname", "family name", "lname",
        "apelyido", "apelido", "beneficiary last name",
    ],
    "suffix": [
        "suffix", "extension name", "ext", "ext.", "suffix name", "extension",
        "name extension", "jr", "sr",
    ],
    "region": ["region", "reg", "region name"],
    "province": ["province", "prov", "province name"],
    "municipality": [
        "city municipality", "city/municipality", "city / municipality",
        "city", "municipality", "mun", "municity", "munisipyo", "town", "lunsod", "lungsod",
    ],
    "barangay": ["barangay", "brgy", "brgy.", "barangay name", "village"],
    "contact_no": [
        "contact no", "contact number", "contact", "contact #", "cellphone", "cell no",
        "mobile", "mobile no", "mobile number", "phone", "phone number", "cp no", "cp#",
    ],
    "birthdate": ["birthdate", "birth date", "date of birth", "dob"],
    "lgu": ["lgu", "local government unit", "municipal lgu", "city lgu"],
    "full_name": ["full name", "fullname", "full_name", "name of beneficiary", "beneficiary name", "name"],
    "sex": ["sex", "gender"],
    "age": ["age"],
}

HEADER_TOKEN_HINTS = {
    "first": {"first", "firstname", "fname", "given", "pangalan"},
    "middle": {"middle", "middlename", "mi", "initial"},
    "last": {"last", "lastname", "lname", "surname", "family", "apelyido", "apelido"},
    "suffix": {"suffix", "ext", "extension", "jr", "sr"},
    "region": {"region", "reg"},
    "province": {"province", "prov"},
    "municipality": {"city", "municipality", "mun", "munisipyo", "town", "lunsod", "lungsod"},
    "barangay": {"barangay", "brgy", "village"},
    "contact_no": {"contact", "phone", "mobile", "cellphone", "cp"},
    "birthdate": {"birthdate", "birth", "dob", "birthday"},
    "birth_year": {"birth", "year", "dob"},
    "birth_month": {"birth", "month", "dob"},
    "birth_day": {"birth", "day", "dob", "birthday"},
    "lgu": {"lgu", "local", "government", "unit"},
    "full_name": {"full", "name", "beneficiary"},
    "sex": {"sex", "gender"},
    "age": {"age"},
}

ENGINE_TO_LOGIC_KEY = {
    "first": "first",
    "middle": "middle",
    "last": "last",
    "suffix": "suffix",
    "region": "region",
    "province": "province",
    "municipality": "municipality",
    "city": "municipality",
    "barangay": "barangay",
    "contact": "contact_no",
    "contact_no": "contact_no",
    "birth_month": "birth_month",
    "birth_day": "birth_day",
    "birth_year": "birth_year",
    "license_no": "license_no",
    "lgu": "lgu",
}


def _normalize_header_text(engine, text: str) -> str:
    if hasattr(engine, "normalize_col"):
        try:
            return engine.normalize_col(text)
        except Exception:
            pass
    return re.sub(r"[^a-z0-9]+", "", str(text or "").lower())


def _header_tokens(engine, text: str) -> set[str]:
    normalized = _normalize_header_text(engine, text)
    raw = re.findall(r"[a-z0-9]+", str(text or "").lower())
    tokens = {tok for tok in raw if tok}
    if normalized:
        tokens.add(normalized)
    return tokens


def _logic_hint_bonus(engine, header: str, logic_key: str) -> int:
    tokens = _header_tokens(engine, header)
    hints = HEADER_TOKEN_HINTS.get(logic_key, set())
    overlap = len(tokens & hints)
    bonus = overlap * 6
    alias_norms = {_normalize_header_text(engine, a) for a in HEADER_EXTRA_ALIASES.get(logic_key, [])}
    header_norm = _normalize_header_text(engine, header)
    if header_norm in alias_norms:
        bonus = max(bonus, 18)
    return bonus


def _score_confidence_label(score: int) -> str:
    if score >= 95:
        return "High"
    if score >= 82:
        return "Medium"
    return "Low"



def _best_alias_score(engine, header_norm: str, aliases) -> int:
    best = 0
    for alias in aliases or []:
        alias_norm = _normalize_header_text(engine, alias)
        if not alias_norm or not header_norm:
            continue
        if header_norm == alias_norm:
            return 100
        if header_norm.startswith(alias_norm) or header_norm.endswith(alias_norm) or alias_norm.startswith(header_norm) or alias_norm.endswith(header_norm):
            best = max(best, 90)
        elif alias_norm in header_norm or header_norm in alias_norm:
            best = max(best, 80)
    return best



def _match_header_to_logic_key(engine, header: str):
    header_text = str(header or "").strip()
    header_lower = re.sub(r"\s+", " ", header_text.lower())
    header_norm = _normalize_header_text(engine, header)
    if not header_norm:
        return None, 0

    # Hard guards to stop common false matches.
    if any(token in header_lower for token in ["validity date", "expiry", "expiration", "expiry date", "expiration date", "exp date"]):
        return None, 0
    if header_lower in {"no", "no.", "number", "row no", "row no.", "row number"}:
        return None, 0

    # Prefer exact split-birth-part fields before fuzzy alias scoring.
    birth_part_rules = [
        ("birth_month", ["birth month", "birthmonth", "month of birth", "dob month", "month born"]),
        ("birth_day", ["birth day", "birthday ", "day of birth", "birth dd", "birth_day", "dob day", "day born"]),
        ("birth_year", ["birth year", "birthyear", "year of birth", "birth yyyy", "birth_year", "dob year", "year born"]),
        ("birthdate", ["birthdate", "birth date", "date of birth", "dob", "birthday"]),
    ]
    for logic_key, phrases in birth_part_rules:
        for phrase in phrases:
            phrase_norm = re.sub(r"\s+", " ", phrase.strip().lower())
            if header_lower == phrase_norm:
                return logic_key, 100

    best_key = None
    best_score = 0

    for logic_key, aliases in HEADER_EXTRA_ALIASES.items():
        score = _best_alias_score(engine, header_norm, aliases)
        if score:
            score = min(100, int(score) + _logic_hint_bonus(engine, header_text, logic_key))
        if score > best_score:
            best_key = logic_key
            best_score = score

    patterns = getattr(engine, "COLUMN_PATTERNS", {}) or {}
    for engine_key, logic_key in ENGINE_TO_LOGIC_KEY.items():
        score = _best_alias_score(engine, header_norm, patterns.get(engine_key, []))
        if score:
            score = min(100, int(score) + _logic_hint_bonus(engine, header_text, logic_key))
        if score > best_score:
            best_key = logic_key
            best_score = score

    if best_score < 78:
        return None, 0
    return best_key, best_score



def _expand_logic_key(logic_key: str) -> list[str]:
    meta = HEADER_LOGIC_META.get(logic_key, {})
    expanded = meta.get("expands_to")
    if expanded:
        return list(expanded)
    return [logic_key] if logic_key else []



def _logic_choice_values() -> list[str]:
    required_core = [
        "first", "last", "middle", "suffix",
        "birthdate", "birth_year", "birth_month", "birth_day",
        "region", "province", "municipality", "barangay",
        "contact_no", "lgu", "full_name", "sex", "age",
    ]
    ordered = []
    for key in required_core + list(LOGIC_MAPPING_ORDER):
        if key in EXCLUDED_USE_AS_KEYS:
            continue
        if key in HEADER_LOGIC_META and key not in ordered:
            ordered.append(key)
    return ["__ignore__"] + ordered



def _logic_choice_label(logic_key: str) -> str:
    if logic_key == "__ignore__":
        return "Keep Original"
    meta = HEADER_LOGIC_META.get(logic_key, {})
    return meta.get("label", str(logic_key or "Keep Original"))


def _logic_choice_label_for_row(logic_key: str, detected_logic: str) -> str:
    if logic_key == "__ignore__":
        return "Keep Original"
    base_label = _logic_choice_label(logic_key)
    if detected_logic and logic_key == detected_logic:
        return f"Detect Header → {base_label}"
    return base_label



def _default_find_for_logic(logic_key: str) -> bool:
    return any(k in DEFAULT_FINDING_FIELDS for k in _expand_logic_key(logic_key))



def _default_dedup_for_logic(logic_key: str) -> bool:
    return any(k in DEFAULT_DEDUPE_FIELDS for k in _expand_logic_key(logic_key))



def _logic_note_for_row(logic_key: str, score: int) -> str:
    meta = HEADER_LOGIC_META.get(logic_key or "", {})
    if not logic_key or not meta:
        return "This column stays in the file, but it will not affect Finding or Dedup unless you map it."

    parts = []
    if meta.get("finding"):
        parts.append("Missing value can move the row to Finding.")
    else:
        parts.append("Not used to move rows to Finding.")

    if meta.get("dedup"):
        parts.append("Can help compare duplicate rows.")
    else:
        parts.append("Not used to compare duplicate rows.")

    parts.append("If Not Applicable is checked, this column is ignored by the logic.")
    return " ".join(parts)



def _on_simple_checkbox_change():
    _clear_previous_run_state()


def _on_mapping_change(mapping_key: str, na_key: str, finding_key: str, dedup_key: str):
    _clear_previous_run_state()
    selected_logic = st.session_state.get(mapping_key, "__ignore__")
    if selected_logic == "__ignore__":
        st.session_state[na_key] = True
        st.session_state[finding_key] = False
        st.session_state[dedup_key] = False
        return

    meta = HEADER_LOGIC_META.get(selected_logic, {})
    st.session_state[na_key] = False
    st.session_state[finding_key] = _default_find_for_logic(selected_logic) if meta.get("finding") else False
    st.session_state[dedup_key] = _default_dedup_for_logic(selected_logic) if meta.get("dedup") else False



def _on_not_applicable_change(na_key: str, mapping_key: str, finding_key: str, dedup_key: str):
    _clear_previous_run_state()
    selected_logic = st.session_state.get(mapping_key, "__ignore__")
    if selected_logic == "__ignore__" or st.session_state.get(na_key, False):
        st.session_state[finding_key] = False
        st.session_state[dedup_key] = False
        st.session_state[na_key] = True
        return

    meta = HEADER_LOGIC_META.get(selected_logic, {})
    st.session_state[finding_key] = _default_find_for_logic(selected_logic) if meta.get("finding") else False
    st.session_state[dedup_key] = _default_dedup_for_logic(selected_logic) if meta.get("dedup") else False



def _on_logic_checkbox_change():
    _clear_previous_run_state()


def _build_header_row_state(row: dict, key_prefix: str) -> dict:
    mapping_key = f"{key_prefix}_mapping_{row['row_id']}"
    finding_key = f"{key_prefix}_finding_{row['row_id']}"
    dedup_key = f"{key_prefix}_dedup_{row['row_id']}"
    na_key = f"{key_prefix}_na_{row['row_id']}"

    if mapping_key not in st.session_state:
        st.session_state[mapping_key] = row.get("logic_key") or "__ignore__"

    selected_logic = st.session_state.get(mapping_key, "__ignore__")
    meta = HEADER_LOGIC_META.get(selected_logic, {})

    if na_key not in st.session_state:
        st.session_state[na_key] = selected_logic == "__ignore__"

    if selected_logic == "__ignore__" or st.session_state.get(na_key, False):
        st.session_state[na_key] = True
        st.session_state[finding_key] = False
        st.session_state[dedup_key] = False
    else:
        if finding_key not in st.session_state:
            st.session_state[finding_key] = _default_find_for_logic(selected_logic) if meta.get("finding") else False
        if dedup_key not in st.session_state:
            st.session_state[dedup_key] = _default_dedup_for_logic(selected_logic) if meta.get("dedup") else False
        if not meta.get("finding"):
            st.session_state[finding_key] = False
        if not meta.get("dedup"):
            st.session_state[dedup_key] = False

    return {
        "mapping_key": mapping_key,
        "finding_key": finding_key,
        "dedup_key": dedup_key,
        "na_key": na_key,
    }



def _collect_mapping_warnings(rows: list[dict], title: str) -> list[str]:
    warnings = []
    mapped = {}
    for row in rows or []:
        logic_key = row.get("logic_key")
        if not logic_key or logic_key == "__ignore__" or row.get("not_applicable"):
            continue
        mapped.setdefault(logic_key, []).append(str(row.get("header") or "").strip())

    for logic_key, headers in mapped.items():
        if len(headers) > 1:
            label = HEADER_LOGIC_META.get(logic_key, {}).get("label", logic_key)
            joined = ", ".join(h for h in headers if h)
            warnings.append(f"{title}: multiple headers are mapped as {label}: {joined}")

    if not any(k in mapped for k in {"first", "last"}):
        warnings.append(f"{title}: no First Name or Last Name mapping is selected yet.")

    return warnings


def _review_status_for_row(row: dict, selected_logic: str, not_applicable: bool) -> tuple[str, bool]:
    detected_logic = row.get("logic_key") or "__ignore__"
    score = int(row.get("score") or 0)
    note = str(row.get("note") or "").lower()

    if not_applicable or selected_logic == "__ignore__":
        return "⚪ Unused", False
    if selected_logic != detected_logic:
        return "🔵 Manual", False
    if "stronger match" in note:
        return "🟡 Conflict", True
    if score and score < 82:
        return "🟡 Review", True
    return "🟢 OK", False


def _review_style_meta(status_text: str) -> tuple[str, str, str]:
    if "Conflict" in status_text:
        return "#fff1f2", "#e11d48", "#881337"
    if "Review" in status_text:
        return "#fffbeb", "#d97706", "#92400e"
    if "Manual" in status_text:
        return "#eff6ff", "#2563eb", "#1e3a8a"
    if "Unused" in status_text:
        return "#f8fafc", "#94a3b8", "#475569"
    return "#ecfdf5", "#16a34a", "#166534"


def _render_detected_header_controls(title: str, rows: list[dict], key_prefix: str, mode_label: str):
    st.markdown(f"**{title}**")
    st.caption("Use As assigns what each uploaded column means. Smart detection suggests the best match, but you can override it safely.")
    if not rows:
        st.caption("No headers detected yet.")
        return []

    review_only_key = f"{key_prefix}_review_only"
    show_review_only = st.checkbox(
        "Review Only",
        key=review_only_key,
        help="Show only Review and Conflict rows so you can fix them faster.",
    )

    hdr1, hdr2, hdr3, hdr4, hdr5, hdr6, hdr7 = st.columns([2.0, 1.0, 1.7, 0.9, 0.9, 1.1, 4.0])
    hdr1.markdown("**Detected Header**")
    hdr2.markdown("**Review**")
    hdr3.markdown("**Use As**")
    hdr4.markdown("**Finding**")
    hdr5.markdown("**Dedup**")
    hdr6.markdown("**Not Applicable**")
    hdr7.markdown("**Meaning**")

    mapping_options = _logic_choice_values()
    resolved_rows = []

    review_rows = []
    display_rows = []
    for row in rows:
        pre_state = _build_header_row_state(row, key_prefix)
        pre_logic = st.session_state.get(pre_state["mapping_key"], row.get("logic_key") or "__ignore__")
        pre_na = bool(st.session_state.get(pre_state["na_key"], row.get("not_applicable", False)))
        _, pre_needs_review = _review_status_for_row(row, pre_logic, pre_na)
        if pre_needs_review:
            review_rows.append(row)
        if not show_review_only or pre_needs_review:
            display_rows.append(row)

    focus_idx_key = f"{key_prefix}_review_focus_idx"
    focus_active_key = f"{key_prefix}_review_focus_active"
    nav1, nav2 = st.columns([1.2, 1.0])
    with nav1:
        if st.button("Fix Next Review Row", key=f"{key_prefix}_fix_next_review", disabled=not review_rows):
            current_idx = int(st.session_state.get(focus_idx_key, -1))
            st.session_state[focus_idx_key] = (current_idx + 1) % max(len(review_rows), 1)
            st.session_state[focus_active_key] = True
    with nav2:
        if st.button("Clear Review Focus", key=f"{key_prefix}_clear_review_focus", disabled=not review_rows):
            st.session_state[focus_active_key] = False
            st.session_state[focus_idx_key] = 0

    if st.session_state.get(focus_active_key, False) and review_rows:
        current_idx = int(st.session_state.get(focus_idx_key, 0)) % len(review_rows)
        focused_row = review_rows[current_idx]
        display_rows = [focused_row]
        st.info(f"Focused review row {current_idx + 1} of {len(review_rows)}: {focused_row.get('header')}")

    st.caption(f"Showing {len(display_rows)} of {len(rows)} detected header rows.")
    if show_review_only and not display_rows:
        st.success("No Review or Conflict rows right now.")
        return rows

    review_needed_headers = []

    for row in display_rows:
        state = _build_header_row_state(row, key_prefix)
        current_logic = st.session_state.get(state["mapping_key"], row.get("logic_key") or "__ignore__")
        current_na = bool(st.session_state.get(state["na_key"], row.get("not_applicable", False)))
        pre_status_text, _ = _review_status_for_row(row, current_logic, current_na)
        bg_color, border_color, text_color = _review_style_meta(pre_status_text)
        if "Review" in pre_status_text or "Conflict" in pre_status_text:
            st.markdown(
                f"<div style='background:{bg_color}; border-left:6px solid {border_color}; color:{text_color}; padding:8px 12px; border-radius:8px; margin:6px 0 6px 0; font-weight:600;'>"
                f"{pre_status_text} — {row['header']}</div>",
                unsafe_allow_html=True,
            )
        c1, c2, c3, c4, c5, c6, c7 = st.columns([2.0, 1.0, 1.7, 0.9, 0.9, 1.1, 4.0])
        c1.write(row["header"])
        if row.get("logic_key"):
            confidence = _score_confidence_label(int(row.get("score") or 0))
            c1.caption(f"Smart detect: {int(row.get('score') or 0)} | {confidence}")

        c3.selectbox(
            f"Use as {row['header']}",
            options=mapping_options,
            index=mapping_options.index(current_logic) if current_logic in mapping_options else 0,
            format_func=lambda value, detected=row.get("logic_key"): _logic_choice_label_for_row(value, detected),
            key=state["mapping_key"],
            label_visibility="collapsed",
            on_change=_on_mapping_change,
            args=(state["mapping_key"], state["na_key"], state["finding_key"], state["dedup_key"]),
        )

        selected_logic = st.session_state.get(state["mapping_key"], "__ignore__")
        meta = HEADER_LOGIC_META.get(selected_logic, {})
        na_disabled = selected_logic == "__ignore__"
        if na_disabled:
            st.session_state[state["na_key"]] = True
            st.session_state[state["finding_key"]] = False
            st.session_state[state["dedup_key"]] = False

        not_applicable = c6.checkbox(
            f"Not Applicable {row['header']}",
            key=state["na_key"],
            disabled=na_disabled,
            label_visibility="collapsed",
            on_change=_on_not_applicable_change,
            args=(state["na_key"], state["mapping_key"], state["finding_key"], state["dedup_key"]),
        )

        find_disabled = not_applicable or (not meta.get("finding"))
        dedup_disabled = not_applicable or (not meta.get("dedup"))

        find_checked = c4.checkbox(
            f"Finding {row['header']}",
            key=state["finding_key"],
            disabled=find_disabled,
            label_visibility="collapsed",
            on_change=_on_logic_checkbox_change,
        )
        dedup_checked = c5.checkbox(
            f"Dedup {row['header']}",
            key=state["dedup_key"],
            disabled=dedup_disabled,
            label_visibility="collapsed",
            on_change=_on_logic_checkbox_change,
        )

        if not_applicable:
            note = "This column is ignored by the logic. Data can stay, but it will not affect Clean, Finding, or Dedup."
        else:
            note = _logic_note_for_row(selected_logic, 100)
            if mode_label == "Generic" and selected_logic == "__ignore__":
                note = "Map this column first if you want it to affect the result."
        status_text, needs_review = _review_status_for_row(row, selected_logic, bool(not_applicable))
        badge_bg, badge_border, badge_text = _review_style_meta(status_text)
        c2.markdown(
            f"<div style='background:{badge_bg}; border:1px solid {badge_border}; color:{badge_text}; "
            f"padding:4px 8px; border-radius:999px; text-align:center; font-weight:700; font-size:0.85rem;'>"
            f"{status_text}</div>",
            unsafe_allow_html=True,
        )
        if needs_review:
            review_needed_headers.append(str(row.get("header") or ""))
        c7.caption(note)

        resolved_rows.append({
            "row_id": row["row_id"],
            "header": row["header"],
            "logic_key": None if selected_logic == "__ignore__" else selected_logic,
            "logic_label": _logic_choice_label(selected_logic),
            "finding_checked": bool(find_checked) and not bool(not_applicable),
            "dedup_checked": bool(dedup_checked) and not bool(not_applicable),
            "not_applicable": bool(not_applicable) or selected_logic == "__ignore__",
            "expanded_keys": _expand_logic_key(selected_logic) if selected_logic != "__ignore__" else [],
        })

    if review_needed_headers:
        joined = ", ".join(review_needed_headers)
        st.warning(f"{title}: review these headers before Preview: {joined}")

    mapping_warnings = _collect_mapping_warnings(resolved_rows, title)
    for msg in mapping_warnings:
        st.warning(msg)

    return resolved_rows



def _control_rows_signature(control_rows: list[dict]):
    payload = []
    for row in control_rows or []:
        payload.append({
            "header": str(row.get("header", "")),
            "logic_key": row.get("logic_key") or "__ignore__",
            "finding": bool(row.get("finding", False)),
            "dedupe": bool(row.get("dedupe", False)),
            "not_applicable": bool(row.get("not_applicable", False)),
        })
    return json.dumps(payload, ensure_ascii=False, sort_keys=True)


def _collect_logic_selections(control_rows: list[dict]):
    finding_selected = set()
    dedupe_selected = set()
    for row in control_rows or []:
        if row.get("not_applicable"):
            continue
        if row.get("finding_checked"):
            finding_selected.update(row.get("expanded_keys", []))
        if row.get("dedup_checked"):
            dedupe_selected.update(row.get("expanded_keys", []))
    return finding_selected, dedupe_selected



def _unique_headers(headers: list[str]) -> list[str]:
    counts = {}
    result = []
    for header in headers:
        base = str(header or "Unnamed")
        counts[base] = counts.get(base, 0) + 1
        result.append(base if counts[base] == 1 else f"{base}__{counts[base]}")
    return result



def _split_full_name_simple(value):
    text = str(value or "").strip()
    if not text or text.lower() in {"nan", "none", "n/a", "na"}:
        return {"first": "", "middle": "", "last": "", "suffix": ""}

    clean = re.sub(r"\s+", " ", text.replace(";", " ").strip())
    if "," in clean:
        last_part, rest_part = [part.strip() for part in clean.split(",", 1)]
        tokens = rest_part.split()
        first = tokens[0] if tokens else ""
        suffix = ""
        if tokens and tokens[-1].lower().replace(".", "") in NAME_SPLIT_SUFFIXES:
            suffix = tokens[-1]
            tokens = tokens[:-1]
            first = tokens[0] if tokens else first
        middle = " ".join(tokens[1:]) if len(tokens) > 1 else ""
        return {"first": first, "middle": middle, "last": last_part, "suffix": suffix}

    tokens = clean.split()
    suffix = ""
    if tokens and tokens[-1].lower().replace(".", "") in NAME_SPLIT_SUFFIXES:
        suffix = tokens[-1]
        tokens = tokens[:-1]
    if len(tokens) == 1:
        return {"first": tokens[0], "middle": "", "last": "", "suffix": suffix}
    if len(tokens) == 2:
        return {"first": tokens[0], "middle": "", "last": tokens[1], "suffix": suffix}
    return {
        "first": tokens[0],
        "middle": " ".join(tokens[1:-1]),
        "last": tokens[-1],
        "suffix": suffix,
    }



def _apply_header_mapping_to_df(df: pd.DataFrame, control_rows: list[dict]):
    if df is None:
        return pd.DataFrame(), []

    out = df.copy()
    out.columns = [str(c) for c in out.columns]
    notices = []

    rename_map = {}
    used_targets = {}
    for row in control_rows or []:
        header = str(row.get("header", ""))
        logic_key = row.get("logic_key")
        if row.get("not_applicable") or not logic_key or header not in out.columns:
            continue

        target = ENGINE_CANONICAL_HEADERS.get(logic_key)
        if not target:
            continue

        used_targets[target] = used_targets.get(target, 0) + 1
        if used_targets[target] == 1:
            rename_map[header] = target
        else:
            rename_map[header] = f"{header}__extra_{used_targets[target]}"
            notices.append(f"More than one column was mapped to {target}. Only the first one is used for matching.")

    if rename_map:
        out = out.rename(columns=rename_map)
        out.columns = _unique_headers(list(out.columns))

    name_cols_present = any(col in out.columns for col in ["First Name", "Middle Name", "Last Name"])
    if (not name_cols_present) and ("Full Name" in out.columns):
        split_rows = out["Full Name"].apply(_split_full_name_simple)
        split_df = pd.DataFrame(split_rows.tolist(), index=out.index)
        mapping = {
            "first": "First Name",
            "middle": "Middle Name",
            "last": "Last Name",
            "suffix": "Extension Name",
        }
        for src_key, col_name in mapping.items():
            if col_name not in out.columns:
                out[col_name] = split_df[src_key]
        notices.append("Full Name was split into First Name, Middle Name, Last Name, and Extension Name for generic matching.")

    out, birth_notices = _normalize_birth_columns_for_matching(out)
    notices.extend(birth_notices)
    return out, list(dict.fromkeys(notices))


def _blank_like(value) -> bool:
    if pd.isna(value):
        return True
    text = str(value or "").strip()
    return text == "" or text.lower() in {"nan", "none", "n/a", "na", "null"}


def _excel_serial_to_timestamp(value):
    try:
        num = float(value)
    except Exception:
        return None
    if not (1 <= num <= 600000):
        return None
    try:
        ts = pd.to_datetime(num, unit="D", origin="1899-12-30", errors="coerce")
    except Exception:
        return None
    if pd.isna(ts):
        return None
    return ts


def _parse_dateish_value(value):
    if value is None or pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return None if pd.isna(value) else value
    if isinstance(value, (datetime, date)):
        return pd.Timestamp(value)

    serial_ts = _excel_serial_to_timestamp(value)
    if serial_ts is not None:
        return serial_ts

    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "n/a", "na", "null"}:
        return None

    for dayfirst in (False, True):
        try:
            ts = pd.to_datetime(text, errors="coerce", dayfirst=dayfirst)
        except Exception:
            ts = pd.NaT
        if not pd.isna(ts):
            return ts
    return None


def _normalize_birth_year_value(value):
    if _blank_like(value):
        return ""
    if isinstance(value, (int, float)) and not pd.isna(value):
        year_num = int(value)
        if 1000 <= year_num <= 2200:
            return str(year_num)
    text_raw = str(value).strip()
    digits_only = re.sub(r"[^0-9]", "", text_raw)
    if digits_only.isdigit():
        year_num = int(digits_only)
        if 1000 <= year_num <= 2200 and not any(ch in text_raw for ch in ['/', '-', ':']):
            return str(year_num)
    ts = _parse_dateish_value(value)
    if ts is not None and not pd.isna(ts):
        return str(int(ts.year))
    if len(digits_only) >= 4:
        return digits_only[:4]
    return text_raw


def _normalize_birth_month_value(value):
    if _blank_like(value):
        return ""
    if isinstance(value, (int, float)) and not pd.isna(value):
        month_num = int(value)
        if 1 <= month_num <= 12:
            return str(month_num)
    text = str(value).strip()
    digits_direct = re.sub(r"[^0-9]", "", text)
    if digits_direct.isdigit() and not any(ch in text for ch in ['/', '-', ':']):
        month_num = int(digits_direct)
        if 1 <= month_num <= 12:
            return str(month_num)
    ts = _parse_dateish_value(value)
    if ts is not None and not pd.isna(ts):
        return str(int(ts.month))

    text_low = re.sub(r"[^a-z]", "", text.lower())
    month_map = {
        "jan": 1, "january": 1,
        "feb": 2, "february": 2,
        "mar": 3, "march": 3,
        "apr": 4, "april": 4,
        "may": 5,
        "jun": 6, "june": 6,
        "jul": 7, "july": 7,
        "aug": 8, "august": 8,
        "sep": 9, "sept": 9, "september": 9,
        "oct": 10, "october": 10,
        "nov": 11, "november": 11,
        "dec": 12, "december": 12,
    }
    if text_low in month_map:
        return str(month_map[text_low])

    digits = re.sub(r"[^0-9]", "", text)
    if digits:
        try:
            month_num = int(digits)
            if 1 <= month_num <= 12:
                return str(month_num)
        except Exception:
            pass
    return text


def _normalize_birth_day_value(value):
    if _blank_like(value):
        return ""
    if isinstance(value, (int, float)) and not pd.isna(value):
        day_num = int(value)
        if 1 <= day_num <= 31:
            return str(day_num)
    text = str(value).strip()
    digits = re.sub(r"[^0-9]", "", text)
    if digits.isdigit() and not any(ch in text for ch in ['/', '-', ':']):
        day_num = int(digits)
        if 1 <= day_num <= 31:
            return str(day_num)
    ts = _parse_dateish_value(value)
    if ts is not None and not pd.isna(ts):
        return str(int(ts.day))
    return text


def _normalize_birthdate_value(value):
    if _blank_like(value):
        return ""
    ts = _parse_dateish_value(value)
    if ts is None or pd.isna(ts):
        return str(value).strip()
    return f"{int(ts.month):02d}/{int(ts.day):02d}/{int(ts.year):04d}"


def _parsed_birth_part(ts, part: str) -> str:
    if ts is None or pd.isna(ts):
        return ""
    if part == "year":
        return str(int(ts.year))
    if part == "month":
        return str(int(ts.month))
    if part == "day":
        return str(int(ts.day))
    return ""


def _normalize_birth_columns_for_matching(df: pd.DataFrame):
    if df is None or df.empty:
        return df, []

    out = df.copy()
    notices = []
    birth_cols = {"Birthdate", "Birth Year", "Birth Month", "Birth Day"}
    present = birth_cols.intersection(set(out.columns))
    if not present:
        return out, notices

    if "Birthdate" in out.columns:
        parsed_birth = out["Birthdate"].apply(_parse_dateish_value)
        out["Birthdate"] = out["Birthdate"].apply(_normalize_birthdate_value)
    else:
        parsed_birth = pd.Series([None] * len(out), index=out.index, dtype="object")

    if "Birth Year" in out.columns:
        normalized_year = out["Birth Year"].apply(_normalize_birth_year_value)
        if "Birthdate" in out.columns:
            normalized_year = normalized_year.where(~normalized_year.eq(""), parsed_birth.apply(lambda ts: _parsed_birth_part(ts, "year")))
        out["Birth Year"] = normalized_year

    if "Birth Month" in out.columns:
        normalized_month = out["Birth Month"].apply(_normalize_birth_month_value)
        if "Birthdate" in out.columns:
            normalized_month = normalized_month.where(~normalized_month.eq(""), parsed_birth.apply(lambda ts: _parsed_birth_part(ts, "month")))
        out["Birth Month"] = normalized_month

    if "Birth Day" in out.columns:
        normalized_day = out["Birth Day"].apply(_normalize_birth_day_value)
        if "Birthdate" in out.columns:
            normalized_day = normalized_day.where(~normalized_day.eq(""), parsed_birth.apply(lambda ts: _parsed_birth_part(ts, "day")))
        out["Birth Day"] = normalized_day

    if present:
        notices.append("Birthdate columns were cleaned before matching so mixed date formats do not break dedup logic.")

    return out, notices


def _safe_source_row_count(path: str, sheet_name: str, header_row: int) -> int:
    if not path or not sheet_name:
        return 0
    try:
        df = pd.read_excel(path, sheet_name=sheet_name, header=max(int(header_row or 0), 0))
    except Exception:
        return 0
    return 0 if df is None else len(df)


def _prepare_run_file(uploaded_file, original_path: str, sheet_name: str, header_row: int, control_rows: list[dict], slot: str):
    if not original_path or not sheet_name:
        return original_path, sheet_name, int(header_row or 0), []

    source = uploaded_file if uploaded_file is not None else original_path
    try:
        if hasattr(source, "seek"):
            source.seek(0)
    except Exception:
        pass

    df = pd.read_excel(source, sheet_name=sheet_name, header=max(int(header_row or 0), 0))
    mapped_df, notices = _apply_header_mapping_to_df(df, control_rows)

    prepared_dir = _get_session_dir() / "prepared"
    prepared_dir.mkdir(parents=True, exist_ok=True)
    prepared_name = f"{slot}_prepared_{_sanitize_filename(Path(original_path).name)}"
    prepared_path = prepared_dir / prepared_name

    output_sheet = (sheet_name or f"Sheet_{slot}")[:31]
    with pd.ExcelWriter(prepared_path, engine="openpyxl") as writer:
        mapped_df.to_excel(writer, sheet_name=output_sheet, index=False)

    try:
        if hasattr(source, "seek"):
            source.seek(0)
    except Exception:
        pass

    return str(prepared_path), output_sheet, 0, notices




def _resolve_auto_detect_conflicts(rows: list[dict]) -> list[dict]:
    single_value_keys = {
        "first", "middle", "last", "suffix", "birthdate", "birth_year", "birth_month", "birth_day",
        "region", "province", "municipality", "barangay", "contact_no", "lgu", "full_name", "sex", "age",
    }
    best_by_key = {}
    for row in rows or []:
        logic_key = row.get("logic_key")
        if logic_key not in single_value_keys:
            continue
        current_best = best_by_key.get(logic_key)
        challenger = (int(row.get("score") or 0), len(str(row.get("header") or "")))
        if current_best is None:
            best_by_key[logic_key] = row
            continue
        incumbent = (int(current_best.get("score") or 0), len(str(current_best.get("header") or "")))
        if challenger > incumbent:
            best_by_key[logic_key] = row

    for row in rows or []:
        logic_key = row.get("logic_key")
        if logic_key in single_value_keys and best_by_key.get(logic_key) is not row:
            label = HEADER_LOGIC_META.get(logic_key, {}).get("label", logic_key)
            row["logic_key"] = None
            row["logic_label"] = "Keep Original / Not Used"
            row["expanded_keys"] = []
            row["default_finding"] = False
            row["default_dedup"] = False
            row["note"] = f"Smart detect kept another stronger match for {label}. Review manually if needed."
    return rows


def _build_detected_header_rows(engine, headers: list[str], source_label: str):
    rows = []
    for idx, header in enumerate(headers or []):
        logic_key, score = _match_header_to_logic_key(engine, header)
        if logic_key in EXCLUDED_USE_AS_KEYS:
            logic_key = None
        meta = HEADER_LOGIC_META.get(logic_key or "", {})
        expanded_keys = _expand_logic_key(logic_key)
        rows.append({
            "row_id": f"{source_label}_{idx}",
            "source": source_label,
            "header": str(header),
            "logic_key": logic_key,
            "logic_label": meta.get("label", "Keep Original / Not Used"),
            "finding_applicable": bool(meta.get("finding", False)),
            "dedup_applicable": bool(meta.get("dedup", False)),
            "expanded_keys": expanded_keys,
            "default_finding": any(k in DEFAULT_FINDING_FIELDS for k in expanded_keys),
            "default_dedup": any(k in DEFAULT_DEDUPE_FIELDS for k in expanded_keys),
            "score": int(score or 0),
            "note": _logic_note_for_row(logic_key, score),
        })
    rows = _resolve_auto_detect_conflicts(rows)
    return rows


def _format_logic_field_labels(field_keys: set[str], field_labels: dict[str, str], field_order: list[str]) -> str:
    ordered = [field_labels[k] for k in field_order if k in (field_keys or set()) and k in field_labels]
    extras = [k for k in sorted(field_keys or set()) if k not in field_labels]
    ordered.extend(extras)
    return ", ".join(ordered) if ordered else "None"


def _read_headers_from_upload(uploaded_file, sheet_name: str, header_row: int) -> list[str]:
    if uploaded_file is None or not sheet_name:
        return []
    try:
        uploaded_file.seek(0)
    except Exception:
        pass
    try:
        df = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=max(int(header_row) - 1, 0), nrows=0)
        headers = [str(c) for c in df.columns]
    except Exception:
        headers = []
    finally:
        try:
            uploaded_file.seek(0)
        except Exception:
            pass
    return headers




def _sync_global_logic_checkbox_state(finding_selected: set[str], dedupe_selected: set[str]) -> None:
    for field_key in FINDING_FIELD_ORDER:
        st.session_state[f"finding_field_{field_key}"] = field_key in set(finding_selected or set())
    for field_key in DEDUPE_FIELD_ORDER:
        st.session_state[f"dedupe_field_{field_key}"] = field_key in set(dedupe_selected or set())


def _render_checkbox_grid(title: str, field_order: list[str], field_labels: dict[str, str], defaults: set[str], key_prefix: str):
    st.markdown(f"**{title}**")
    cols = st.columns(3)
    selected = set()
    for idx, field_key in enumerate(field_order):
        label = field_labels.get(field_key, field_key)
        checked = cols[idx % 3].checkbox(
            label,
            value=(field_key in defaults),
            key=f"{key_prefix}_{field_key}",
            on_change=_on_simple_checkbox_change,
        )
        if checked:
            selected.add(field_key)
    return selected




def _reason_matches_enabled_finding(reason_text: str, active_fields) -> bool:
    reason = str(reason_text or "").strip()
    if not reason:
        return False
    active = set(active_fields or [])
    lower_reason = reason.lower()

    if "all fields missing" in lower_reason or "no data row" in lower_reason:
        return True
    if "years old" in lower_reason and ("below" in lower_reason or "above" in lower_reason):
        return True

    field_patterns = {
        "first": ("first name",),
        "middle": ("middle name",),
        "last": ("last name",),
        "suffix": ("extension name", "suffix"),
        "region": ("region",),
        "province": ("province",),
        "municipality": ("city/municipality", "municipality", "city"),
        "barangay": ("barangay",),
        "contact_no": ("contact no", "contact number", "phone number"),
        "birth_month": ("birth month",),
        "birth_day": ("birth day",),
        "birth_year": ("birth year",),
        "license_no": ("driver's license", "drivers license", "license number"),
    }

    for field_key, patterns in field_patterns.items():
        if any(pattern in lower_reason for pattern in patterns):
            return field_key in active

    if "birthdate" in lower_reason or "birth date" in lower_reason:
        return bool({"birth_month", "birth_day", "birth_year"} & active)

    return True


def _filter_lack_reason_text(reason_text: str, active_fields) -> str:
    parts = [part.strip() for part in str(reason_text or "").split(";") if part.strip()]
    if not parts:
        return ""
    kept = []
    seen = set()
    for part in parts:
        if _reason_matches_enabled_finding(part, active_fields):
            key = part.lower()
            if key not in seen:
                kept.append(part)
                seen.add(key)
    return "; ".join(kept)


def _install_streamlit_finding_checkbox_fix(engine) -> None:
    if getattr(engine, "_streamlit_finding_checkbox_fix_installed", False):
        return

    original = getattr(engine, "build_identity_lists", None)
    if not callable(original):
        return

    engine._streamlit_original_build_identity_lists = original

    def _build_identity_lists_streamlit_safe(df, include_birth, source_label="A"):
        result = list(engine._streamlit_original_build_identity_lists(df, include_birth, source_label))
        if len(result) < 14:
            return tuple(result)

        active_fields = set(getattr(engine, "_streamlit_active_finding_fields", set()) or set())
        no_detail_series = list(result[12] or [])
        lack_reason_list = list(result[13] or [])

        max_len = max(len(no_detail_series), len(lack_reason_list))
        new_no_detail = []
        new_lack_reason = []

        for idx in range(max_len):
            reason_text = lack_reason_list[idx] if idx < len(lack_reason_list) else ""
            filtered_reason = _filter_lack_reason_text(reason_text, active_fields)
            new_lack_reason.append(filtered_reason)
            new_no_detail.append(bool(filtered_reason))

        result[12] = new_no_detail
        result[13] = new_lack_reason
        return tuple(result)

    engine.build_identity_lists = _build_identity_lists_streamlit_safe
    engine._streamlit_finding_checkbox_fix_installed = True


def _apply_streamlit_field_overrides(engine, finding_fields: set[str], dedupe_fields: set[str]):
    engine._streamlit_active_finding_fields = set(finding_fields)
    _install_streamlit_finding_checkbox_fix(engine)

    required_vars = getattr(engine, "required_field_vars", {}) or {}
    for field_key, var in required_vars.items():
        try:
            var.set(field_key in finding_fields)
        except Exception:
            pass

    if not hasattr(engine, "_streamlit_original_birthdate_match_detail"):
        engine._streamlit_original_birthdate_match_detail = engine._birthdate_match_detail
    if not hasattr(engine, "_streamlit_original_birth_candidate_block_keys"):
        engine._streamlit_original_birth_candidate_block_keys = engine._birth_candidate_block_keys

    selected_name_fields = {k for k in dedupe_fields if k in {"first", "middle", "last", "suffix"}}
    selected_birth_parts = {k for k in dedupe_fields if k in {"birth_year", "birth_month", "birth_day"}}

    def _blankish(value):
        norm = engine.normalize_name(value)
        return "" if norm in {"", "na", "n/a", "none", "null", "unknown", "notapplicable", "notapp"} else norm

    def _name_fields_label():
        ordered = [DEDUPE_FIELD_LABELS[k] for k in DEDUPE_FIELD_ORDER if k in selected_name_fields]
        return ", ".join(ordered) if ordered else "No Name Fields Selected"

    def _birth_fields_label():
        ordered = [DEDUPE_FIELD_LABELS[k] for k in DEDUPE_FIELD_ORDER if k in selected_birth_parts]
        return ", ".join(ordered) if ordered else "Birthdate Check Disabled"

    def _birth_candidate_block_keys_custom(norm_birth, include_birth=True):
        if not include_birth:
            return tuple()
        key = str(norm_birth or "").strip()
        if not key:
            return tuple()
        yy, mm, dd = engine._birth_parts_from_norm(key)
        if None in {yy, mm, dd}:
            return tuple()

        parts = set(selected_birth_parts)
        if parts == {"birth_year", "birth_month", "birth_day"}:
            return engine._streamlit_original_birth_candidate_block_keys(norm_birth, include_birth=include_birth)

        keys = []
        if parts == {"birth_year"}:
            keys.append(("BIRTH_Y", f"{yy:04d}"))
        elif parts == {"birth_month"}:
            keys.append(("BIRTH_M", f"{mm:02d}"))
        elif parts == {"birth_day"}:
            keys.append(("BIRTH_D", f"{dd:02d}"))
        else:
            if {"birth_year", "birth_month"}.issubset(parts):
                keys.append(("BIRTH_YM", f"{yy:04d}{mm:02d}"))
            if {"birth_year", "birth_day"}.issubset(parts):
                keys.append(("BIRTH_YD", f"{yy:04d}{dd:02d}"))
            if {"birth_month", "birth_day"}.issubset(parts):
                keys.append(("BIRTH_MD", f"{mm:02d}{dd:02d}"))
                md_lo, md_hi = min(mm, dd), max(mm, dd)
                keys.append(("BIRTH_MD_UNORDERED", f"{md_lo:02d}{md_hi:02d}"))
        return tuple(dict.fromkeys(keys))

    def _birthdate_match_detail_custom(norm_birth_1, norm_birth_2, include_birth=True):
        if not include_birth:
            return {"accepted": True, "exact": True, "score": 100.0, "reason": "Birthdate Check Disabled"}

        parts = set(selected_birth_parts)
        if not parts:
            return {"accepted": True, "exact": True, "score": 100.0, "reason": "Birthdate Check Disabled"}
        if parts == {"birth_year", "birth_month", "birth_day"}:
            return engine._streamlit_original_birthdate_match_detail(norm_birth_1, norm_birth_2, include_birth=include_birth)

        b1 = str(norm_birth_1 or "").strip()
        b2 = str(norm_birth_2 or "").strip()
        if not b1 or not b2:
            return {"accepted": False, "exact": False, "score": 0.0, "reason": f"{_birth_fields_label()} Missing On One Side"}

        y1, m1, d1 = engine._birth_parts_from_norm(b1)
        y2, m2, d2 = engine._birth_parts_from_norm(b2)
        if None in {y1, m1, d1, y2, m2, d2}:
            return {"accepted": False, "exact": False, "score": 0.0, "reason": f"{_birth_fields_label()} Not Comparable"}

        diffs = []
        if "birth_year" in parts and y1 != y2:
            diffs.append("Birth Year")
        if "birth_month" in parts and m1 != m2:
            diffs.append("Birth Month")
        if "birth_day" in parts and d1 != d2:
            diffs.append("Birth Day")
        if diffs:
            return {"accepted": False, "exact": False, "score": 0.0, "reason": "Different " + " / ".join(diffs)}
        return {"accepted": True, "exact": True, "score": 100.0, "reason": f"Same {_birth_fields_label()}"}



    def _match_result_rank(res):
        if not res:
            return (-1.0, -1.0)
        dt = str(res.get("dedup_type", "")).upper()
        ranks = {
            "TRUE_DUPLICATE": 5.0,
            "SWAP_TRUE_DUPLICATE": 4.0,
            "SWAP_FUZZ_DUPLICATE": 3.5,
            "FUZZ_DUPLICATE": 3.0,
            "LICENSE_CONFLICT": 1.0,
        }
        return (ranks.get(dt, 0.0), float(res.get("score", 0.0) or 0.0))

    def _best_match_result(*results):
        best = None
        best_key = (-1.0, -1.0)
        for res in results:
            key = _match_result_rank(res)
            if key > best_key:
                best = res
                best_key = key
        return best

    def _confidence_label_local(dedup_type, score=0.0):
        dt = str(dedup_type or "").upper()
        sc = float(score or 0.0)
        if dt == "TRUE_DUPLICATE":
            return "HIGH"
        if dt == "SWAP_TRUE_DUPLICATE":
            return "HIGH" if sc >= 97 else "MEDIUM"
        if dt == "SWAP_FUZZ_DUPLICATE":
            return "MEDIUM" if sc >= 92 else "LOW"
        if dt == "FUZZ_DUPLICATE":
            return "MEDIUM" if sc >= 90 else "LOW"
        return "REVIEW"

    def _ordered_selected_name_fields():
        return [k for k in DEDUPE_FIELD_ORDER if k in selected_name_fields]

    def _make_row_maps(first, middle, last, suffix, firstph=None, middleph=None, lastph=None):
        raw = {"first": first, "middle": middle, "last": last, "suffix": suffix}
        norm = {
            "first": engine.canonical_first_name(first) or engine.normalize_name(first),
            "middle": _blankish(middle),
            "last": engine.normalize_name(last),
            "suffix": engine.normalize_ext(suffix),
        }
        ph = {"first": firstph, "middle": middleph, "last": lastph, "suffix": None}
        return raw, norm, ph

    def _mapped_second_row(raw2_map, norm2_map, ph2_map, perm):
        ordered = _ordered_selected_name_fields()
        mapped_raw = dict(raw2_map)
        mapped_norm = dict(norm2_map)
        mapped_ph = dict(ph2_map)
        for target_field, source_field in zip(ordered, perm):
            mapped_raw[target_field] = raw2_map[source_field]
            mapped_norm[target_field] = norm2_map[source_field]
            mapped_ph[target_field] = ph2_map[source_field]
        return mapped_raw, mapped_norm, mapped_ph

    def _evaluate_alignment(raw1_map, norm1_map, ph1_map, raw2_map, norm2_map, ph2_map,
                            birth_detail, threshold, edit_tol, phonetic_min,
                            exact_dedup_type, fuzzy_dedup_type, reason_prefix, swap_mode=False):
        exact_checks = []
        if "first" in selected_name_fields:
            exact_checks.append(bool(norm1_map["first"] and norm2_map["first"] and norm1_map["first"] == norm2_map["first"]))
        if "middle" in selected_name_fields:
            exact_checks.append(norm1_map["middle"] == norm2_map["middle"])
        if "last" in selected_name_fields:
            exact_checks.append(bool(norm1_map["last"] and norm2_map["last"] and norm1_map["last"] == norm2_map["last"]))
        if "suffix" in selected_name_fields:
            exact_checks.append(norm1_map["suffix"] == norm2_map["suffix"])
        exact_name_match = bool(exact_checks) and all(exact_checks)

        if exact_name_match and birth_detail.get("exact", False):
            reason = f"{reason_prefix} + Same Birthdate"
            score = 100.0 if not swap_mode else 99.0
            return {
                "score": score,
                "type": "EXACT" if not swap_mode else "SWAP_EXACT",
                "reason": reason,
                "dedup_type": exact_dedup_type,
                "confidence": _confidence_label_local(exact_dedup_type, score=score),
                "conflict_reason": reason,
            }

        first_ok = True if "first" not in selected_name_fields else engine.first_name_compatible(
            raw1_map["first"], raw2_map["first"], edit_tol, phonetic_min, ph_a=ph1_map["first"], ph_b=ph2_map["first"]
        )
        middle_ok = True if "middle" not in selected_name_fields else engine.middle_name_compatible(
            raw1_map["middle"], raw2_map["middle"], edit_tol, phonetic_min, mode=engine.MIDDLE_NAME_MODE,
            ph_a=ph1_map["middle"], ph_b=ph2_map["middle"]
        )
        last_ok = True if "last" not in selected_name_fields else engine.field_compatible(
            raw1_map["last"], raw2_map["last"], edit_tol, phonetic_min, ph_a=ph1_map["last"], ph_b=ph2_map["last"]
        )
        suffix_ok = True if "suffix" not in selected_name_fields else (norm1_map["suffix"] == norm2_map["suffix"])
        if not (first_ok and middle_ok and last_ok and suffix_ok):
            return None

        profile = engine._dedup_name_profile(
            raw1_map["first"], raw1_map["middle"], raw1_map["last"], raw1_map["suffix"],
            raw2_map["first"], raw2_map["middle"], raw2_map["last"], raw2_map["suffix"]
        )
        weights = {"first": 0.40, "middle": 0.20, "last": 0.35, "suffix": 0.05}
        field_scores = {
            "first": 100.0 if (norm1_map["first"] and norm1_map["first"] == norm2_map["first"]) else float(profile.get("first_score", 0.0) or 0.0),
            "middle": 100.0 if (norm1_map["middle"] == norm2_map["middle"]) else float(profile.get("middle_score", 0.0) or 0.0),
            "last": 100.0 if (norm1_map["last"] and norm1_map["last"] == norm2_map["last"]) else float(profile.get("last_score", 0.0) or 0.0),
            "suffix": 100.0 if (norm1_map["suffix"] == norm2_map["suffix"]) else 0.0,
        }
        total_weight = sum(weights[k] for k in selected_name_fields)
        name_score = sum(field_scores[k] * weights[k] for k in selected_name_fields) / total_weight if total_weight else 0.0
        score = (name_score * 0.85) + (float(birth_detail.get("score", 100.0)) * 0.15)
        min_threshold = max(84.0 if swap_mode else 82.0, float(threshold or 80.0))
        if score < min_threshold:
            return None

        reason = f"{reason_prefix} + {birth_detail.get('reason', 'Accepted Fuzzy Birthdate')}"
        return {
            "score": score,
            "type": "SWAP_FUZZY" if swap_mode else "FUZZY",
            "reason": reason,
            "dedup_type": fuzzy_dedup_type,
            "confidence": _confidence_label_local(fuzzy_dedup_type, score=score),
            "conflict_reason": reason,
        }

    def _standard_match_result(first1, middle1, last1, suffix1, first2, middle2, last2, suffix2,
                               firstph1, middleph1, lastph1, firstph2, middleph2, lastph2,
                               birth_detail, threshold, edit_tol, phonetic_min):
        raw1_map, norm1_map, ph1_map = _make_row_maps(first1, middle1, last1, suffix1, firstph1, middleph1, lastph1)
        raw2_map, norm2_map, ph2_map = _make_row_maps(first2, middle2, last2, suffix2, firstph2, middleph2, lastph2)
        return _evaluate_alignment(
            raw1_map, norm1_map, ph1_map,
            raw2_map, norm2_map, ph2_map,
            birth_detail, threshold, edit_tol, phonetic_min,
            "TRUE_DUPLICATE", "FUZZ_DUPLICATE",
            f"Same Name ({_name_fields_label()})", swap_mode=False
        )

    def _swap_match_result(first1, middle1, last1, suffix1, first2, middle2, last2, suffix2,
                           firstph1, middleph1, lastph1, firstph2, middleph2, lastph2,
                           birth_detail, threshold, edit_tol, phonetic_min):
        ordered = _ordered_selected_name_fields()
        if len(ordered) < 2:
            return None

        raw1_map, norm1_map, ph1_map = _make_row_maps(first1, middle1, last1, suffix1, firstph1, middleph1, lastph1)
        raw2_map, norm2_map, ph2_map = _make_row_maps(first2, middle2, last2, suffix2, firstph2, middleph2, lastph2)

        import itertools
        identity = tuple(ordered)
        best = None

        for perm in itertools.permutations(ordered):
            if tuple(perm) == identity:
                continue
            mapped_raw2, mapped_norm2, mapped_ph2 = _mapped_second_row(raw2_map, norm2_map, ph2_map, perm)
            perm_label = " / ".join(f"{src}->{dst}" for dst, src in zip(ordered, perm))
            candidate = _evaluate_alignment(
                raw1_map, norm1_map, ph1_map,
                mapped_raw2, mapped_norm2, mapped_ph2,
                birth_detail, threshold, edit_tol, phonetic_min,
                "SWAP_TRUE_DUPLICATE", "SWAP_FUZZ_DUPLICATE",
                f"Swapped/Reordered Name Parts ({perm_label})",
                swap_mode=True
            )
            best = _best_match_result(best, candidate)

        return best

    def _evaluate_pair(first1, middle1, last1, suffix1, first2, middle2, last2, suffix2,
                       firstph1, middleph1, lastph1, firstph2, middleph2, lastph2,
                       birth1, birth2, include_birth, threshold, edit_tol, phonetic_min):
        birth_detail = _birthdate_match_detail_custom(birth1, birth2, include_birth=include_birth)
        if include_birth and not birth_detail.get("accepted", False):
            return None

        standard_res = _standard_match_result(
            first1, middle1, last1, suffix1, first2, middle2, last2, suffix2,
            firstph1, middleph1, lastph1, firstph2, middleph2, lastph2,
            birth_detail, threshold, edit_tol, phonetic_min,
        )
        swap_res = _swap_match_result(
            first1, middle1, last1, suffix1, first2, middle2, last2, suffix2,
            firstph1, middleph1, lastph1, firstph2, middleph2, lastph2,
            birth_detail, threshold, edit_tol, phonetic_min,
        )
        return _best_match_result(standard_res, swap_res)

    def _pair_match_detail_within_custom(idx1, idx2,
        first_list, first_canon_list, middle_list, last_list, suffix_list,
        identity_list, identity_ph_list,
        first_ph_list, middle_ph_list, last_ph_list, suffix_ph_list,
        birth_list, license_list, plate_list, support_strict,
        include_birth, threshold, edit_tol, phonetic_min):
        return _evaluate_pair(
            engine._safe_list_value(first_list, idx1, ""), engine._safe_list_value(middle_list, idx1, ""), engine._safe_list_value(last_list, idx1, ""), engine._safe_list_value(suffix_list, idx1, ""),
            engine._safe_list_value(first_list, idx2, ""), engine._safe_list_value(middle_list, idx2, ""), engine._safe_list_value(last_list, idx2, ""), engine._safe_list_value(suffix_list, idx2, ""),
            engine._safe_list_value(first_ph_list, idx1, ""), engine._safe_list_value(middle_ph_list, idx1, ""), engine._safe_list_value(last_ph_list, idx1, ""),
            engine._safe_list_value(first_ph_list, idx2, ""), engine._safe_list_value(middle_ph_list, idx2, ""), engine._safe_list_value(last_ph_list, idx2, ""),
            engine._safe_list_value(birth_list, idx1, ""), engine._safe_list_value(birth_list, idx2, ""),
            include_birth, threshold, edit_tol, phonetic_min,
        )

    def _pair_match_detail_cross_custom(ia, ib,
        first_a, first_canon_a, middle_a, last_a, suffix_a,
        id_a, idph_a,
        firstph_a, middleph_a, lastph_a, suffixph_a,
        birth_a, license_a, plate_a, support_strict_a,
        first_b, first_canon_b, middle_b, last_b, suffix_b,
        id_b, idph_b,
        firstph_b, middleph_b, lastph_b, suffixph_b,
        birth_b, license_b, plate_b, support_strict_b,
        include_birth, threshold, edit_tol, phonetic_min):
        return _evaluate_pair(
            engine._safe_list_value(first_a, ia, ""), engine._safe_list_value(middle_a, ia, ""), engine._safe_list_value(last_a, ia, ""), engine._safe_list_value(suffix_a, ia, ""),
            engine._safe_list_value(first_b, ib, ""), engine._safe_list_value(middle_b, ib, ""), engine._safe_list_value(last_b, ib, ""), engine._safe_list_value(suffix_b, ib, ""),
            engine._safe_list_value(firstph_a, ia, ""), engine._safe_list_value(middleph_a, ia, ""), engine._safe_list_value(lastph_a, ia, ""),
            engine._safe_list_value(firstph_b, ib, ""), engine._safe_list_value(middleph_b, ib, ""), engine._safe_list_value(lastph_b, ib, ""),
            engine._safe_list_value(birth_a, ia, ""), engine._safe_list_value(birth_b, ib, ""),
            include_birth, threshold, edit_tol, phonetic_min,
        )

    engine._birth_candidate_block_keys = _birth_candidate_block_keys_custom
    engine._birthdate_match_detail = _birthdate_match_detail_custom
    engine._pair_match_detail_within = _pair_match_detail_within_custom
    engine._pair_match_detail_cross = _pair_match_detail_cross_custom


def _apply_birthdate_limit_overrides(engine, use_min_age: bool, min_age: int, use_max_age: bool, max_age: int):
    try:
        engine.MIN_VALID_AGE = int(min_age) if use_min_age else 0
    except Exception:
        engine.MIN_VALID_AGE = 0
    try:
        engine.MAX_VALID_AGE = int(max_age) if use_max_age else 999
    except Exception:
        engine.MAX_VALID_AGE = 999


def _replace_reason_text(value, use_min_age: bool, min_age: int, use_max_age: bool, max_age: int):
    text = str(value or "")
    if use_min_age:
        text = text.replace("Below 18 years old", f"Below {int(min_age)} years old")
    if use_max_age:
        text = text.replace("Above 85 years old", f"Above {int(max_age)} years old")
    return text


def _apply_dynamic_age_reason_labels(engine, results: dict, use_min_age: bool, min_age: int, use_max_age: bool, max_age: int):
    candidate_columns = {"FindingReason", "Lack_Reason"}

    def _patch_df(df):
        if df is None or getattr(df, "empty", True):
            return df
        for col in candidate_columns:
            if col in df.columns:
                df[col] = df[col].apply(lambda v: _replace_reason_text(v, use_min_age, min_age, use_max_age, max_age))
        return df

    engine_attrs = [
        "result_lack_ab", "result_no_detail_a", "result_no_detail_b",
        "result_clean_ab", "result_dup_ab", "result_clean_a", "result_dup_a",
        "result_clean_b", "result_dup_b", "result_cross_dup_ab", "result_cross_clean_a", "result_cross_clean_b",
    ]
    for attr in engine_attrs:
        try:
            df = getattr(engine, attr, None)
            if df is not None:
                _patch_df(df)
        except Exception:
            pass

    for key, df in list((results or {}).items()):
        results[key] = _patch_df(df)
    return results


def _patch_engine_for_streamlit(engine, progress_bar, status_box, metrics_box, message_box):
    def _raise_error(exc):
        raise RuntimeError(str(exc)) from exc

    def _ui_set_progress(percent):
        try:
            value = max(0.0, min(100.0, float(percent))) / 100.0
        except Exception:
            value = 0.0
        progress_bar.progress(value)

    def _ui_set_status(text="", metrics=""):
        if text:
            status_box.info(text)
        else:
            status_box.empty()
        if metrics:
            metrics_box.caption(metrics)
        else:
            metrics_box.empty()

    def _ui_message(kind, title, message):
        text = f"**{title}**\n\n{message}"
        if kind == "warning":
            message_box.warning(text)
        elif kind == "error":
            message_box.error(text)
        else:
            message_box.success(text)

    engine._run_on_ui_thread = lambda func, *args, wait=False, **kwargs: func(*args, **kwargs)
    engine._ui_set_progress = _ui_set_progress
    engine._ui_set_status = _ui_set_status
    engine._ui_message = _ui_message
    engine._ui_refresh_counts = lambda: None
    engine._ui_reset_report_all = lambda: getattr(engine, "reset_report", lambda: None)()
    engine.show_error = _raise_error




def _find_shared_id_columns(df_left, df_right):
    candidate_sets = [
        ["_Source", "_RowUID"],
        ["_Source", "SourceRowRef"],
        ["_Source", "SourceRowNumber"],
        ["_Source", "NO."],
        ["SourceFile", "SourceRowRef"],
        ["SourceFile", "SourceRowNumber"],
        ["_RowUID"],
        ["SourceRowRef"],
        ["SourceRowNumber"],
    ]
    left_cols = set(getattr(df_left, "columns", []))
    right_cols = set(getattr(df_right, "columns", []))
    for cols in candidate_sets:
        if all(c in left_cols and c in right_cols for c in cols):
            return cols
    return []

def _normalize_key_part(value):
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()

def _remove_non_true_survivors_from_clean(clean_df, dup_df):
    if clean_df is None or getattr(clean_df, "empty", True):
        return clean_df
    if dup_df is None or getattr(dup_df, "empty", True):
        return clean_df
    if "DedupType" not in dup_df.columns:
        return clean_df

    dup_types = dup_df["DedupType"].astype(str).str.upper().fillna("")
    nontrue_dup = dup_df.loc[(dup_types != "") & (dup_types != "TRUE_DUPLICATE")].copy()
    if nontrue_dup.empty:
        return clean_df

    if "DedupType" in clean_df.columns:
        clean_types = clean_df["DedupType"].astype(str).str.upper().fillna("")
        mask_keep = (clean_types == "") | (clean_types == "TRUE_DUPLICATE")
        return clean_df.loc[mask_keep].copy()

    shared_id_cols = _find_shared_id_columns(clean_df, nontrue_dup)
    if not shared_id_cols:
        return clean_df

    nontrue_keys = set()
    for _, row in nontrue_dup[shared_id_cols].iterrows():
        nontrue_keys.add(tuple(_normalize_key_part(row[c]) for c in shared_id_cols))

    if not nontrue_keys:
        return clean_df

    keep_mask = []
    for _, row in clean_df.iterrows():
        key = tuple(_normalize_key_part(row[c]) for c in shared_id_cols)
        keep_mask.append(key not in nontrue_keys)
    return clean_df.loc[keep_mask].copy()

def _enforce_true_survivor_only(engine):
    try:
        engine.result_clean_ab = _remove_non_true_survivors_from_clean(
            getattr(engine, "result_clean_ab", None),
            getattr(engine, "result_dup_ab", None),
        )
    except Exception:
        pass
    try:
        engine.result_clean_a = _remove_non_true_survivors_from_clean(
            getattr(engine, "result_clean_a", None),
            getattr(engine, "result_dup_a", None),
        )
    except Exception:
        pass
    try:
        engine.result_clean_b = _remove_non_true_survivors_from_clean(
            getattr(engine, "result_clean_b", None),
            getattr(engine, "result_dup_b", None),
        )
    except Exception:
        pass
    try:
        engine.result_cross_clean_a = _remove_non_true_survivors_from_clean(
            getattr(engine, "result_cross_clean_a", None),
            getattr(engine, "result_cross_dup_ab", None),
        )
    except Exception:
        pass
    try:
        engine.result_cross_clean_b = _remove_non_true_survivors_from_clean(
            getattr(engine, "result_cross_clean_b", None),
            getattr(engine, "result_cross_dup_ab", None),
        )
    except Exception:
        pass

def _run_engine(engine, file_a_path, file_b_path, sheet_a, sheet_b, hdr_a, hdr_b, strict_level, include_birth, use_min_age, min_age, use_max_age, max_age):
    progress_box = st.progress(0.0)
    status_box = st.empty()
    metrics_box = st.empty()
    message_box = st.empty()
    _patch_engine_for_streamlit(engine, progress_box, status_box, metrics_box, message_box)
    _apply_birthdate_limit_overrides(engine, use_min_age, min_age, use_max_age, max_age)

    run_inputs = {
        "file_a": file_a_path or "",
        "file_b": file_b_path or "",
        "sheet_a": sheet_a or "",
        "sheet_b": sheet_b or "",
        "hdr_a": int(hdr_a or 0),
        "hdr_b": int(hdr_b or 0),
        "strict_level": int(strict_level or 80),
        "include_birth": bool(include_birth),
    }

    engine.run_dedup_impl(export_files=False, run_inputs=run_inputs)
    _enforce_true_survivor_only(engine)

    results = {
        "clean_ab": _df_copy(_prefer_existing(getattr(engine, "result_clean_ab", None), getattr(engine, "_compute_clean_ab_df", lambda: None))),
        "dup_ab": _df_copy(_prefer_existing(getattr(engine, "result_dup_ab", None), getattr(engine, "_compute_dup_ab_df", lambda: None))),
        "lack_ab": _df_copy(_prefer_existing(getattr(engine, "result_lack_ab", None), getattr(engine, "_compute_lack_ab_df", lambda: None))),
        "clean_a": _df_copy(getattr(engine, "result_clean_a", None)),
        "dup_a": _df_copy(getattr(engine, "result_dup_a", None)),
        "lack_a": _df_copy(getattr(engine, "result_no_detail_a", None)),
        "clean_b": _df_copy(getattr(engine, "result_clean_b", None)),
        "dup_b": _df_copy(getattr(engine, "result_dup_b", None)),
        "lack_b": _df_copy(getattr(engine, "result_no_detail_b", None)),
        "cross_dup_ab": _df_copy(getattr(engine, "result_cross_dup_ab", None)),
        "cross_clean_a": _df_copy(getattr(engine, "result_cross_clean_a", None)),
        "cross_clean_b": _df_copy(getattr(engine, "result_cross_clean_b", None)),
    }
    progress_box.progress(1.0)
    return results




def _dedup_type_badge_css(value):
    val = str(value or "").strip().upper()
    styles = {
        "TRUE_DUPLICATE": "background-color: #dcfce7; color: #166534; font-weight: 700;",
        "FUZZ_DUPLICATE": "background-color: #fef3c7; color: #92400e; font-weight: 700;",
        "SWAP_TRUE_DUPLICATE": "background-color: #dbeafe; color: #1d4ed8; font-weight: 700;",
        "SWAP_FUZZ_DUPLICATE": "background-color: #fde68a; color: #9a3412; font-weight: 700;",
        "LICENSE_CONFLICT": "background-color: #fee2e2; color: #b91c1c; font-weight: 700;",
    }
    return styles.get(val, "")

def _confidence_badge_css(value):
    val = str(value or "").strip().upper()
    styles = {
        "HIGH": "background-color: #dcfce7; color: #166534; font-weight: 700;",
        "MEDIUM": "background-color: #fef3c7; color: #92400e; font-weight: 700;",
        "LOW": "background-color: #fee2e2; color: #b91c1c; font-weight: 700;",
        "REVIEW": "background-color: #e5e7eb; color: #374151; font-weight: 700;",
    }
    return styles.get(val, "")

def _build_preview_styler(df: pd.DataFrame):
    styler = df.style
    if "DedupType" in df.columns:
        styler = styler.applymap(_dedup_type_badge_css, subset=["DedupType"])
    if "Confidence" in df.columns:
        styler = styler.applymap(_confidence_badge_css, subset=["Confidence"])
    return styler

def _display_result_block(label: str, df: pd.DataFrame, key_prefix: str):
    if df is None or getattr(df, "empty", True):
        st.caption(f"{label}: no data")
        return
    st.markdown(f"**{label}** — {len(df):,} rows × {len(df.columns):,} columns")
    preview_limit = 5000
    if len(df) > preview_limit:
        st.caption(f"Preview limited to first {preview_limit:,} rows for browser speed. Download gets the full table.")
        st.dataframe(_build_preview_styler(df.head(preview_limit).copy()), use_container_width=True, height=420)
    else:
        st.dataframe(_build_preview_styler(df.copy()), use_container_width=True, height=420)
    st.download_button(
        f"Download {label} as Excel",
        data=_to_excel_bytes(df, sheet_name=label[:31]),
        file_name=f"{key_prefix}_{re.sub(r'[^A-Za-z0-9_-]+', '_', label)}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"download_{key_prefix}_{label}",
    )




def _template_store_dir() -> Path:
    root = Path(__file__).resolve().parent / "_dedup_header_templates"
    root.mkdir(parents=True, exist_ok=True)
    return root



def _safe_template_name(name: str) -> str:
    clean = re.sub(r"[^A-Za-z0-9._-]+", "_", str(name or "last_used").strip())
    return clean or "last_used"



def _template_file_path(name: str) -> Path:
    return _template_store_dir() / f"{_safe_template_name(name)}.json"



def _build_template_rows_from_state(header_rows: list[dict], key_prefix: str) -> dict:
    payload = {}
    for row in header_rows or []:
        row_id = row["row_id"]
        mapping_key = f"{key_prefix}_mapping_{row_id}"
        finding_key = f"{key_prefix}_finding_{row_id}"
        dedup_key = f"{key_prefix}_dedup_{row_id}"
        na_key = f"{key_prefix}_na_{row_id}"
        logic_key = st.session_state.get(mapping_key, row.get("logic_key") or "__ignore__")
        payload[str(row["header"])] = {
            "logic_key": logic_key,
            "finding": bool(st.session_state.get(finding_key, row.get("default_finding", False))),
            "dedup": bool(st.session_state.get(dedup_key, row.get("default_dedup", False))),
            "not_applicable": bool(st.session_state.get(na_key, logic_key == "__ignore__")),
        }
    return payload



def _build_template_payload(app_mode: str, header_rows_a: list[dict], header_rows_b: list[dict]) -> dict:
    return {
        "app_mode": app_mode,
        "table_a": _build_template_rows_from_state(header_rows_a, "table_a_header_logic"),
        "table_b": _build_template_rows_from_state(header_rows_b, "table_b_header_logic"),
        "table_a_headers": [str(row.get("header", "")) for row in header_rows_a or []],
        "table_b_headers": [str(row.get("header", "")) for row in header_rows_b or []],
    }



def _save_header_template(template_name: str, payload: dict):
    target = _template_file_path(template_name)
    target.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    (_template_store_dir() / "_last_template_name.txt").write_text(_safe_template_name(template_name), encoding="utf-8")
    (_template_store_dir() / "_last_template.json").write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")



def _load_header_template(template_name: str | None = None):
    if template_name:
        target = _template_file_path(template_name)
    else:
        target = _template_store_dir() / "_last_template.json"
    if not target.exists():
        return None
    try:
        return json.loads(target.read_text(encoding="utf-8"))
    except Exception:
        return None



def _last_template_name() -> str:
    target = _template_store_dir() / "_last_template_name.txt"
    if target.exists():
        value = target.read_text(encoding="utf-8").strip()
        if value:
            return value
    return "last_used"



def _safe_streamlit_rerun():
    rerun_fn = getattr(st, "rerun", None) or getattr(st, "experimental_rerun", None)
    if callable(rerun_fn):
        rerun_fn()


def _header_match_stats(current_headers: list[str], saved_headers: list[str]):
    current_clean = [str(h or "").strip() for h in current_headers or [] if str(h or "").strip()]
    saved_clean = [str(h or "").strip() for h in saved_headers or [] if str(h or "").strip()]
    current_set = set(current_clean)
    saved_set = set(saved_clean)
    overlap = len(current_set & saved_set)
    denom = max(len(saved_set), len(current_set), 1)
    ratio = overlap / denom
    return overlap, ratio, len(current_set), len(saved_set)


def _template_is_compatible(header_rows: list[dict], payload: dict, table_name: str, *, require_strong_match: bool) -> bool:
    if not payload:
        return False
    current_headers = [str(row.get("header", "")) for row in header_rows or []]
    saved_headers = payload.get(f"{table_name}_headers") or []
    if not current_headers:
        return False
    if not saved_headers:
        return not require_strong_match
    overlap, ratio, current_count, saved_count = _header_match_stats(current_headers, saved_headers)
    if require_strong_match:
        return overlap >= 3 and ratio >= 0.60
    return overlap >= 1 and (ratio >= 0.35 or overlap >= min(3, saved_count, current_count))


def _sanitize_loaded_logic_key(header: str, logic_key: str):
    header_text = str(header or "").strip()
    header_lower = re.sub(r"\s+", " ", header_text.lower())
    if any(token in header_lower for token in ["validity date", "expiry", "expiration", "expiry date", "expiration date", "exp date"]):
        return "__ignore__"
    if header_lower in {"no", "no.", "number", "row no", "row no.", "row number"} and logic_key == "contact_no":
        return "__ignore__"
    if header_lower == "birth day" and logic_key == "birthdate":
        return "birth_day"
    if header_lower == "birth month" and logic_key == "birthdate":
        return "birth_month"
    if header_lower == "birth year" and logic_key == "birthdate":
        return "birth_year"
    return logic_key


def _apply_template_to_rows(header_rows: list[dict], key_prefix: str, table_payload: dict):
    if not table_payload:
        return
    for row in header_rows or []:
        saved = table_payload.get(str(row["header"]))
        if not saved:
            continue
        row_id = row["row_id"]
        mapping_key = f"{key_prefix}_mapping_{row_id}"
        finding_key = f"{key_prefix}_finding_{row_id}"
        dedup_key = f"{key_prefix}_dedup_{row_id}"
        na_key = f"{key_prefix}_na_{row_id}"
        logic_key = saved.get("logic_key") or "__ignore__"
        logic_key = _sanitize_loaded_logic_key(str(row.get("header", "")), logic_key)
        st.session_state[mapping_key] = logic_key
        st.session_state[na_key] = bool(saved.get("not_applicable", logic_key == "__ignore__")) or logic_key == "__ignore__"
        st.session_state[finding_key] = bool(saved.get("finding", False)) and not st.session_state[na_key]
        st.session_state[dedup_key] = bool(saved.get("dedup", False)) and not st.session_state[na_key]
        if logic_key == "__ignore__":
            st.session_state[finding_key] = False
            st.session_state[dedup_key] = False



def _header_signature(header_rows: list[dict]) -> str:
    return "|".join(str(row.get("header", "")) for row in header_rows or [])



def _auto_apply_last_template(header_rows: list[dict], key_prefix: str, table_name: str):
    signature = _header_signature(header_rows)
    state_key = f"{key_prefix}_auto_template_signature"
    if not signature or st.session_state.get(state_key) == signature:
        return
    payload = _load_header_template(None)
    if payload and _template_is_compatible(header_rows, payload, table_name, require_strong_match=True):
        _apply_template_to_rows(header_rows, key_prefix, payload.get(table_name, {}))
        app_mode = payload.get("app_mode")
        if app_mode in {"Preset", "Generic"} and "dedup_app_mode" not in st.session_state:
            st.session_state["dedup_app_mode"] = app_mode
    st.session_state[state_key] = signature


def _clear_previous_run_state():
    for key in [
        "fuel_results",
        "fuel_export_zip_bytes",
        "fuel_export_zip_name",
        "fuel_config_snapshot",
        "fuel_last_error",
    ]:
        st.session_state.pop(key, None)


def main():
    st.set_page_config(page_title="Streamlit_DeDup_App", layout="wide")
    st.title("Streamlit_DeDup_App")
    st.caption("Uses your latest cross-swap/group-view Tkinter engine inside the newer Streamlit UI, while keeping the safer header-review controls.")

    with st.expander("Run note", expanded=True):
        st.write(
            "Upload File A and optional File B, choose sheets/header rows, then run Preview or build an Export ZIP. "
            "This version keeps your dedup/autofix/export logic but replaces Tkinter popups with in-page Streamlit controls."
        )

    engine = load_engine()

    col1, col2 = st.columns(2)
    with col1:
        file_a = st.file_uploader("Excel File A", type=["xlsx", "xlsm", "xls"], key="file_a_uploader")
        path_a = _save_upload(file_a, "A") if file_a is not None else ""
        sheets_a = _safe_list_sheets(engine, path_a)
        sheet_a = st.selectbox("Sheet A", sheets_a or [""], index=0, key="sheet_a_select")
        hdr_a = st.number_input("Header Row A (0-based)", min_value=0, value=0, step=1, key="hdr_a_input")

    with col2:
        file_b = st.file_uploader("Excel File B (optional)", type=["xlsx", "xlsm", "xls"], key="file_b_uploader")
        path_b = _save_upload(file_b, "B") if file_b is not None else ""
        sheets_b = _safe_list_sheets(engine, path_b)
        sheet_b = st.selectbox("Sheet B", sheets_b or [""], index=0, key="sheet_b_select")
        hdr_b = st.number_input("Header Row B (0-based)", min_value=0, value=0, step=1, key="hdr_b_input")

    opt1, opt2, opt3, opt4 = st.columns([1.2, 1, 1, 2.2])
    with opt1:
        app_mode = st.radio("Mode", ["Preset", "Generic"], horizontal=True, key="dedup_app_mode")
    with opt2:
        strict_level = st.slider("Strictness", min_value=0, max_value=100, value=80, step=1)
    with opt3:
        include_birth = st.checkbox("Include Birthdate", value=True)
    with opt4:
        if app_mode == "Preset":
            st.info("Preset mode keeps the built-in behavior, but you can still adjust each detected header before running.")
        else:
            st.info("Generic mode lets you remap headers so the same app can work on other file layouts too.")

    with st.expander("Birthdate age limit settings", expanded=False):
        st.caption("For Birth Year / Birth Month / Birth Day checking, you can choose a minimum age and a maximum age. Uncheck a rule if you do not want to use it.")
        age_col1, age_col2 = st.columns(2)
        with age_col1:
            use_min_age = st.checkbox("Use below-age limit", value=True, key="use_min_age_limit")
            min_age_limit = st.number_input("Below-age number", min_value=0, max_value=150, value=18, step=1, key="min_age_limit_input", disabled=not use_min_age)
        with age_col2:
            use_max_age = st.checkbox("Use too-old limit", value=True, key="use_max_age_limit")
            max_age_limit = st.number_input("Too-old number", min_value=0, max_value=150, value=85, step=1, key="max_age_limit_input", disabled=not use_max_age)

        if use_min_age and use_max_age and int(min_age_limit) > int(max_age_limit):
            st.error("Below-age limit cannot be greater than the too-old limit.")

    headers_a = _read_headers_from_upload(file_a, sheet_a, hdr_a + 1) if file_a and sheet_a else []
    headers_b = _read_headers_from_upload(file_b, sheet_b, hdr_b + 1) if file_b and sheet_b else []

    header_rows_a = _build_detected_header_rows(engine, headers_a, "A")
    header_rows_b = _build_detected_header_rows(engine, headers_b, "B")
    _auto_apply_last_template(header_rows_a, "table_a_header_logic", "table_a")
    _auto_apply_last_template(header_rows_b, "table_b_header_logic", "table_b")

    control_rows_a = []
    control_rows_b = []
    finding_selected = set()
    dedupe_selected = set()

    with st.expander("Header controls for Finding and Dedup", expanded=False):
        st.caption(
            "Each detected header can be mapped, turned on for Finding, turned on for Dedup, or marked Not Applicable. "
            "Not Applicable means the data may stay in the file, but that header will not affect the logic."
        )
        if headers_a:
            st.markdown("**Detected table headers - A**")
            st.code(" | ".join(headers_a[:60]) + (" | ..." if len(headers_a) > 60 else ""))
        if headers_b:
            st.markdown("**Detected table headers - B**")
            st.code(" | ".join(headers_b[:60]) + (" | ..." if len(headers_b) > 60 else ""))

        if header_rows_a or header_rows_b:
            if header_rows_a:
                control_rows_a = _render_detected_header_controls("Table A header logic", header_rows_a, "table_a_header_logic", app_mode)
                fa, da = _collect_logic_selections(control_rows_a)
                finding_selected.update(fa)
                dedupe_selected.update(da)
            if header_rows_a and header_rows_b:
                st.divider()
            if header_rows_b:
                control_rows_b = _render_detected_header_controls("Table B header logic", header_rows_b, "table_b_header_logic", app_mode)
                fb, db = _collect_logic_selections(control_rows_b)
                finding_selected.update(fb)
                dedupe_selected.update(db)

            _sync_global_logic_checkbox_state(finding_selected, dedupe_selected)

            template_cols = st.columns([1.6, 1, 1.4])
            template_name = template_cols[0].text_input("Save template name", value=st.session_state.get("header_template_name", _last_template_name()), key="header_template_name")
            save_template = template_cols[1].button("Save template", use_container_width=True)
            load_template = template_cols[2].button("Load saved template", use_container_width=True)
            st.caption("Saved template remembers your header mapping, Finding, Dedup, and Not Applicable choices for the next run.")

            if save_template:
                payload = _build_template_payload(app_mode, header_rows_a, header_rows_b)
                _save_header_template(template_name, payload)
                st.success(f"Template saved: {_safe_template_name(template_name)}")

            if load_template:
                payload = _load_header_template(template_name)
                table_a_ok = _template_is_compatible(header_rows_a, payload, "table_a", require_strong_match=False) if header_rows_a else True
                table_b_ok = _template_is_compatible(header_rows_b, payload, "table_b", require_strong_match=False) if header_rows_b else True
                if payload and table_a_ok and table_b_ok:
                    _apply_template_to_rows(header_rows_a, "table_a_header_logic", payload.get("table_a", {}))
                    _apply_template_to_rows(header_rows_b, "table_b_header_logic", payload.get("table_b", {}))
                    if payload.get("app_mode") in {"Preset", "Generic"}:
                        st.session_state["dedup_app_mode"] = payload["app_mode"]
                    st.success(f"Template loaded: {_safe_template_name(template_name)}")
                    _safe_streamlit_rerun()
                elif payload:
                    st.warning("Saved template does not match the current headers closely enough. Load a better template or save a new one for this file.")
                else:
                    st.warning("Saved template not found yet.")
        else:
            st.caption("No headers detected yet, so the fallback canonical controls are shown.")
            finding_selected = _render_checkbox_grid(
                "Finding: mark row as Finding when these fields are missing",
                FINDING_FIELD_ORDER,
                FINDING_FIELD_LABELS,
                DEFAULT_FINDING_FIELDS,
                "finding_field",
            )
            st.divider()
            dedupe_selected = _render_checkbox_grid(
                "Dedup: use these fields in match logic and reason text",
                DEDUPE_FIELD_ORDER,
                DEDUPE_FIELD_LABELS,
                DEFAULT_DEDUPE_FIELDS,
                "dedupe_field",
            )

        age_rules_text = []
        if use_min_age:
            age_rules_text.append(f"Below-age < {int(min_age_limit)}")
        if use_max_age:
            age_rules_text.append(f"Too-old > {int(max_age_limit)}")
        if not age_rules_text:
            age_rules_text.append("Age limits disabled")

        st.info(
            "Finding watches: "
            + _format_logic_field_labels(finding_selected, FINDING_FIELD_LABELS, FINDING_FIELD_ORDER)
            + "\n\nDedup compares: "
            + _format_logic_field_labels(dedupe_selected, DEDUPE_FIELD_LABELS, DEDUPE_FIELD_ORDER)
            + "\n\nBirthdate age rules: "
            + ", ".join(age_rules_text)
        )
        st.caption("Tip: for stable dedup, keep at least First Name, Last Name, and Birth Year checked.")

    btn1, btn2 = st.columns(2)
    run_preview = btn1.button("Run Preview", type="primary", use_container_width=True)
    run_export = btn2.button("Build Export ZIP", use_container_width=True)

    results = st.session_state.get("fuel_results")
    config_snapshot = {
        "file_a_name": file_a.name if file_a is not None else "",
        "file_b_name": file_b.name if file_b is not None else "",
        "path_a": path_a,
        "path_b": path_b,
        "sheet_a": sheet_a,
        "sheet_b": sheet_b,
        "hdr_a": int(hdr_a),
        "hdr_b": int(hdr_b),
        "strict_level": int(strict_level),
        "include_birth": bool(include_birth),
        "use_min_age": bool(use_min_age),
        "min_age_limit": int(min_age_limit),
        "use_max_age": bool(use_max_age),
        "max_age_limit": int(max_age_limit),
        "finding_fields": sorted(finding_selected),
        "dedupe_fields": sorted(dedupe_selected),
        "mode": app_mode,
        "table_a_logic_signature": _control_rows_signature(control_rows_a),
        "table_b_logic_signature": _control_rows_signature(control_rows_b),
    }

    if run_preview or run_export:
        if use_min_age and use_max_age and int(min_age_limit) > int(max_age_limit):
            st.error("Below-age limit cannot be greater than the too-old limit.")
        elif not path_a and not path_b:
            st.error("Upload at least File A or File B first.")
        elif not ({"first", "middle", "last", "suffix"} & set(dedupe_selected)):
            st.error("Check at least one name field in Dedup: First Name, Middle Name, Last Name, or Extension Name.")
        else:
            _clear_previous_run_state()
            try:
                prepared_a, prepared_sheet_a, prepared_hdr_a, notices_a = _prepare_run_file(
                    file_a, path_a, sheet_a, hdr_a, control_rows_a, "A"
                ) if path_a and sheet_a else (path_a, sheet_a, hdr_a, [])
                prepared_b, prepared_sheet_b, prepared_hdr_b, notices_b = _prepare_run_file(
                    file_b, path_b, sheet_b, hdr_b, control_rows_b, "B"
                ) if path_b and sheet_b else (path_b, sheet_b, hdr_b, [])

                for notice in list(dict.fromkeys((notices_a or []) + (notices_b or []))):
                    st.caption(f"• {notice}")

                _apply_streamlit_field_overrides(engine, finding_selected, dedupe_selected)
                results = _run_engine(
                    engine,
                    prepared_a,
                    prepared_b,
                    prepared_sheet_a,
                    prepared_sheet_b,
                    prepared_hdr_a,
                    prepared_hdr_b,
                    strict_level,
                    include_birth,
                    use_min_age,
                    int(min_age_limit),
                    use_max_age,
                    int(max_age_limit),
                )
                results = _apply_dynamic_age_reason_labels(
                    engine,
                    results,
                    use_min_age,
                    int(min_age_limit),
                    use_max_age,
                    int(max_age_limit),
                )
                st.session_state["fuel_results"] = results
                st.session_state["fuel_config_snapshot"] = config_snapshot

                if run_export:
                    export_dir = _get_session_dir() / "exports"
                    for old in export_dir.glob("*"):
                        try:
                            old.unlink()
                        except Exception:
                            pass
                    prefix, tag = _build_prefix_and_tag(config_snapshot["file_a_name"], config_snapshot["file_b_name"])
                    paths = engine.export_outputs(str(export_dir), prefix, tag, export_cross=False)
                    zip_name = f"{prefix}_{tag}_EXPORT.zip"
                    zip_bytes = _zip_paths(paths, zip_name)
                    st.session_state["fuel_export_zip_bytes"] = zip_bytes
                    st.session_state["fuel_export_zip_name"] = zip_name
                    st.success("Export ZIP is ready.")
                else:
                    st.success("Preview complete.")
            except Exception as exc:
                st.session_state["fuel_last_error"] = str(exc)
                st.exception(exc)
                results = None

    if st.session_state.get("fuel_export_zip_bytes"):
        st.download_button(
            "Download Export ZIP",
            data=st.session_state["fuel_export_zip_bytes"],
            file_name=st.session_state.get("fuel_export_zip_name", "Streamlit_DeDup_App_Export.zip"),
            mime="application/zip",
            use_container_width=True,
        )

    if results:
        last_config_snapshot = st.session_state.get("fuel_config_snapshot")
        if last_config_snapshot != config_snapshot:
            st.warning("Header/file settings changed after the last run. These results are outdated — rerun needed.")
        clean_ab_df = results.get("clean_ab")
        lack_ab_df = results.get("lack_ab")
        if (clean_ab_df is None or len(clean_ab_df) == 0) and lack_ab_df is not None and len(lack_ab_df) > 0:
            reason_col = "Lack_Reason" if "Lack_Reason" in lack_ab_df.columns else None
            if reason_col:
                top_reasons = lack_ab_df[reason_col].astype(str).value_counts().head(3)
                summary = "; ".join(f"{reason} ({count})" for reason, count in top_reasons.items())
                st.warning("Clean A+B is empty because rows were pushed to Finding A+B. Top reasons: " + summary)
            else:
                st.warning("Clean A+B is empty because rows were pushed to Finding A+B.")

        clean_ab_count = 0 if results.get("clean_ab") is None else len(results.get("clean_ab"))
        dedup_ab_count = 0 if results.get("dup_ab") is None else len(results.get("dup_ab"))
        finding_ab_count = 0 if results.get("lack_ab") is None else len(results.get("lack_ab"))
        cross_dup_ab_count = 0 if results.get("cross_dup_ab") is None else len(results.get("cross_dup_ab"))
        overall_file_ab_count = _safe_source_row_count(path_a, sheet_a, hdr_a) + _safe_source_row_count(path_b, sheet_b, hdr_b)
        total_main_ab_count = clean_ab_count + dedup_ab_count + finding_ab_count
        total_dedup_finding_ab_count = dedup_ab_count + finding_ab_count

        counts = [
            {"View": "Overall File A+B", "Rows": overall_file_ab_count},
            {"View": "Total Clean+Dedup+Finding A+B", "Rows": total_main_ab_count},
            {"View": "Total Dedup+Finding A+B", "Rows": total_dedup_finding_ab_count},
            {"View": "Clean A+B", "Rows": clean_ab_count},
            {"View": "Dedup A+B", "Rows": dedup_ab_count},
            {"View": "Finding A+B", "Rows": finding_ab_count},
            {"View": "Cross Duplicate A+B", "Rows": cross_dup_ab_count},
        ]
        st.subheader("Result counts")
        st.dataframe(pd.DataFrame(counts), use_container_width=True, hide_index=True)

        tabs = st.tabs(["Clean A+B", "Dedup A+B", "Finding A+B", "Cross Duplicate A+B"])
        tab_map = [
            ("Clean A+B", "clean_ab"),
            ("Dedup A+B", "dup_ab"),
            ("Finding A+B", "lack_ab"),
            ("Cross Duplicate A+B", "cross_dup_ab"),
        ]
        for tab, (label, key) in zip(tabs, tab_map):
            with tab:
                _display_result_block(label, results.get(key), key)


if __name__ == "__main__":
    main()
